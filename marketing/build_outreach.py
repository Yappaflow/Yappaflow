"""Build Yappaflow organic-outreach workbook for Turkish web agencies + freelancers."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

OUT = "/sessions/inspiring-busy-dirac/mnt/Yappaflow/marketing/Yappaflow_TR_Outreach_Leads.xlsx"

# ---- Styling ----
FONT = "Arial"
HDR_FILL = PatternFill("solid", start_color="1F3A5F")  # navy
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT, bold=True, size=16, color="1F3A5F")
H2_FONT = Font(name=FONT, bold=True, size=12, color="1F3A5F")
BODY_FONT = Font(name=FONT, size=11)
BODY_BOLD = Font(name=FONT, size=11, bold=True)
ZEBRA = PatternFill("solid", start_color="F2F5FA")
PRIORITY_HIGH = PatternFill("solid", start_color="FFE1A8")
PRIORITY_MED = PatternFill("solid", start_color="F5F5F5")
THIN = Side(border_style="thin", color="C7CED8")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

wb = Workbook()

# =========================================================
# Sheet 1: Plan
# =========================================================
plan = wb.active
plan.title = "Plan"
plan.sheet_view.showGridLines = False

plan["A1"] = "Yappaflow — Turkish organic-outreach plan"
plan["A1"].font = TITLE_FONT
plan.merge_cells("A1:F1")

plan["A2"] = "Prepared 2026-04-20 · Region: Turkey · Audience: small web/branding agencies + freelance web devs/designers"
plan["A2"].font = Font(name=FONT, italic=True, size=10, color="666666")
plan.merge_cells("A2:F2")

sections = [
    ("1. What we are selling them", [
        "Yappaflow takes a two-party chat (agency ↔ client) and auto-generates + packages a complete, deploy-ready website as a ZIP.",
        "Supported platforms today: custom static site (ready); Shopify export (in-flight).",
        "Every generated site ships with a light default theme + working dark-mode toggle, WCAG AA contrast.",
        "For agencies: cuts the 'brief → first draft' cycle from days to minutes. For freelancers: lets one person deliver the volume of a small studio.",
    ]),
    ("2. Two angles, one list", [
        "CLIENT angle (default) — 'Use Yappaflow to deliver client websites 5-10x faster.' Use this for agencies + freelancers who build small-business sites.",
        "PARTNERSHIP angle — Use this for Shopify/ikas partner agencies. Frame Yappaflow as a funnel filler: when their store clients also need a marketing site, we generate it + hand back a ZIP they can host.",
        "Priority in the Leads tab reflects which angle likely lands better: High = warm fit (custom-site shops + small studios), Med = partner agencies where the partnership angle needs a meeting, Low = larger full-service shops where we're a long shot.",
    ]),
    ("3. First-touch outreach sequence (suggested)", [
        "Day 0 — personalized cold email (120-180 words). Reference something from their site ('saw your project for X'). Offer a 3-minute Loom walking through Yappaflow with their own brand as the demo.",
        "Day 3 — LinkedIn connection request with no pitch, just 'sent you a note about a tool I'm building.'",
        "Day 7 — reply in-thread with a short clip or a link to a generated demo site using one of their past clients' copy.",
        "Day 14 — final bump, 'closing the loop — worth a 15-min chat?'",
        "Track status in the Leads tab 'Status' column so you can see pipeline at a glance.",
    ]),
    ("4. What to promise (and not promise)", [
        "DO say: free for the first 5 sites, white-label ZIP export, light+dark theme by default, you keep the client relationship.",
        "DO NOT say: 'fully automated agency in a box' — agencies resist tools that replace them. Position as a delivery accelerator, not a replacement.",
        "Always point to: live demo site (recommend building one using a real Turkish café/atelier as a public showcase before sending).",
    ]),
    ("5. Social / organic parallel track", [
        "Open X (Twitter) + LinkedIn company pages in the same week. Post 3-4x/week: before/after builds, a generated site per day, bite-size threads on 'how we turned a WhatsApp chat into a live site'.",
        "Join Turkish dev + agency communities: 'Türkiye Dijital Ajanslar' LinkedIn group, kodluyoruz Discord, Indie Hackers Turkey, armut/bionluk forums.",
        "Reach out to 2-3 Turkish indie-hacker YouTubers/podcasters (Birkan Ulusoy, Can Arslan, etc.) for a feature once you have a public demo.",
    ]),
    ("6. Lead-sheet legend", [
        "Type: AGENCY (≥3 people) · FREELANCER (solo) · PLATFORM PARTNER (Shopify/ikas/Webflow certified agency).",
        "Priority: HIGH = strong fit and easy to reach · MED = fit but needs warmer intro · LOW = larger/full-service, use as stretch.",
        "Status column left blank for you to fill: 'Queued', 'Sent', 'Replied', 'Meeting', 'Passed', 'Closed'.",
        "Some rows have no public email — use the website contact form or LinkedIn DM as the first touch.",
    ]),
]

row = 4
for title, bullets in sections:
    plan.cell(row=row, column=1, value=title).font = H2_FONT
    plan.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    for b in bullets:
        plan.cell(row=row, column=1, value="  • " + b).font = BODY_FONT
        plan.cell(row=row, column=1).alignment = WRAP
        plan.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1
    row += 1  # blank spacer

plan.column_dimensions["A"].width = 22
plan.column_dimensions["B"].width = 22
plan.column_dimensions["C"].width = 22
plan.column_dimensions["D"].width = 22
plan.column_dimensions["E"].width = 22
plan.column_dimensions["F"].width = 22
# row heights auto with wrap, but give content rows a bit more space
for r in range(4, row):
    plan.row_dimensions[r].height = 32

# =========================================================
# Sheet 2: Leads
# =========================================================
leads = wb.create_sheet("Leads")
leads.sheet_view.showGridLines = False

HEADERS = [
    "#", "Company / Name", "Type", "City", "Website", "Email", "Phone",
    "LinkedIn", "Decision-maker", "Fit note", "Outreach angle", "Priority", "Status",
]

# Row data. (company, type, city, website, email, phone, linkedin, dm, fit, angle, priority)
LEADS = [
    # --- Istanbul small/mid agencies (CLIENT angle) ---
    ("Grimor Web Tasarım", "Agency", "Istanbul (Mecidiyeköy)", "https://www.grimor.com/", "info@grimor.com",
     "+90 212 272 46 00", "https://www.linkedin.com/company/grimor-ajans/", "",
     "Custom corporate sites since 2004 — exactly the 'every project is bespoke' shop Yappaflow saves days for.",
     "Client", "High"),
    ("Kumsal Agency", "Agency", "Istanbul (Maltepe)", "https://www.kumsalajans.com/", "hello@kumsalajans.com",
     "+90 216 706 60 64", "https://www.linkedin.com/company/kumsalajans/", "",
     "Full-service with web as a core service. Good candidate for agency-delivery angle.",
     "Client", "High"),
    ("WebStudio (Web Studio Dijital Ajans)", "Agency", "Istanbul (Ataşehir)", "https://www.ws.com.tr/", "info@ws.com.tr",
     "+90 216 469 2011", "https://www.linkedin.com/company/web-studio-dijital-ajans/", "",
     "Delivery-focused agency, also runs destek@ support queue — signals volume.",
     "Client", "Medium"),
    ("Penta Yazılım", "Agency", "Istanbul", "https://www.pentayazilim.com/", "satis@pentayazilim.com",
     "+90 850 302 6950", "https://tr.linkedin.com/company/pentayazilim", "",
     "Self-described 'Turkey's Web Design Agency', 50+ staff, 1000+ projects — stretch lead but prestige if landed.",
     "Client", "Low"),
    ("Real Web Tasarım", "Agency", "Istanbul", "https://www.realwebtasarim.com/", "info@realwebtasarim.com",
     "+90 532 069 94 35", "", "",
     "Mid-size corporate web shop since 2012. Standard delivery pipeline — clear time-savings pitch.",
     "Client", "Medium"),
    ("MediaClick", "Agency", "Istanbul (Ataşehir)", "https://www.mediaclick.com.tr/", "destek@mediaclick.com.tr",
     "+90 850 811 2530", "", "",
     "Award-winning Istanbul agency. destek@ is a support address — try reaching founder on LinkedIn first.",
     "Client", "Medium"),
    ("UA Creative Agency", "Agency", "Istanbul", "https://umayajans.com/", "murat@umayajans.com",
     "+90 850 242 56 40", "https://www.linkedin.com/in/murattekmen/", "Murat Tekmen (Founder)",
     "Founded 2015, award-winning. Founder is active on LinkedIn — warm up there before emailing.",
     "Client", "High"),
    ("Sentez Bilişim", "Agency", "Istanbul (Başakşehir)", "https://www.sentezbilisim.com/", "info@sentezbilisim.com",
     "+90 212 470 00 03", "", "",
     "E-commerce + corporate sites, WooCommerce/OpenCart shop — could love a faster draft tool.",
     "Client", "High"),
    ("Onex Software", "Agency", "Istanbul", "https://onexyazilim.com/", "info@onexyazilim.com",
     "+90 545 862 15 65", "", "",
     "Listed as a web-design company in TR directories — low-friction cold email target.",
     "Client", "Medium"),
    ("432 Design Studio", "Agency", "Istanbul", "https://432designstudio.com/", "iletisim@432tasarim.com",
     "+90 212 243 53 63", "", "",
     "Design-forward studio, the type that will judge generator OUTPUT — lead with the demo site.",
     "Client", "Medium"),
    ("1MM", "Agency", "Istanbul", "https://1mm.com.tr/", "info@1mm.com.tr",
     "0850 888 8520", "", "",
     "Small/mid web shop — standard delivery pitch.",
     "Client", "Medium"),
    ("Istanbul Webmaster", "Agency", "Istanbul", "https://istanbulwebmaster.com/", "info@istanbulwebmaster.com",
     "+90 212 576 3000", "", "",
     "Generalist TR web firm — low-effort cold outreach.",
     "Client", "Low"),
    ("Betatek", "Agency", "Istanbul (Beşiktaş)", "https://www.betatek.com/", "info@betatek.com",
     "+90 212 401 4122", "", "",
     "Serves Yapı Kredi, Deloitte, Cartoon Network — big-ticket shop. Partnership angle may fit better than client.",
     "Partnership", "Low"),
    ("TekTurkey", "Agency", "Istanbul", "https://tekturkey.com/", "sales@tekturkey.com",
     "+90 212 452 5057", "", "",
     "Web design + marketing shop. sales@ = warm-ish inbound address.",
     "Client", "Medium"),
    ("VantaWorks", "Agency", "Istanbul (Kağıthane)", "https://vantaworks.com/", "info@vantaworks.com",
     "+90 850 885 30 71", "https://www.linkedin.com/company/vantaworks/", "",
     "~6-person team doing corporate web + CRM/ERP — perfect size for Yappaflow to feel immediate.",
     "Client", "High"),
    ("DIJICREA", "Agency", "Istanbul (Kadıköy)", "https://dijicrea.com/", "info@dijicrea.com",
     "+90 216 771 7796", "https://www.linkedin.com/company/dijicrea", "",
     "23-person agency with an 'AI-forward' positioning — warm audience for an AI-driven generator.",
     "Client", "High"),
    ("Webintek", "Agency", "Istanbul (Kartal)", "https://webintek.com.tr/", "bilgi@webintek.com.tr",
     "+90 216 606 03 99", "https://tr.linkedin.com/company/webintekcomtr", "",
     "Corporate transformation services, remote delivery model — less friction to try a new tool.",
     "Client", "High"),
    ("Dijitanya", "Agency", "Istanbul", "https://www.dijitanya.com/", "",
     "+90 539 923 7385", "", "Kürşad Sualp",
     "Boutique performance-marketing agency, 10+ experts. No public email — reach Kürşad via WhatsApp/LinkedIn.",
     "Client", "Medium"),
    ("Hoops Agency", "Agency", "Istanbul (Beşiktaş)", "https://hoops.com.tr/", "",
     "", "", "",
     "Creative + social + digital (360). Partnership framing lands better than client here. Use contact form.",
     "Partnership", "Medium"),
    ("Artnova Creative", "Platform Partner", "Istanbul (Pendik)", "https://www.artnovacreative.com/", "",
     "+90 547 180 18 80", "https://tr.linkedin.com/company/artnovacreative", "",
     "Official ikas partner agency. PARTNERSHIP angle: Yappaflow's ikas export could be their new default.",
     "Partnership", "High"),
    ("PATH", "Platform Partner", "Istanbul (Ataşehir)", "https://path.com.tr/", "info@path.com.tr",
     "+90 536 240 24 24", "https://www.linkedin.com/company/pathdevelop/", "",
     "Shopify Plus Partner — pitch Yappaflow's Shopify export as a client-onboarding accelerator.",
     "Partnership", "High"),
    ("Nodus Works", "Platform Partner", "Istanbul", "https://nodusworks.com/", "hub@nodusworks.com",
     "+90 850 888 2815", "https://linkedin.com/company/nodusworks", "Efe Gündüz (Founder)",
     "Shopify partner, one of the few in TR — partnership integration is the lead hook.",
     "Partnership", "High"),
    ("DigitalPals", "Platform Partner", "Istanbul (Maslak)", "https://digitalpals.com/", "hello@digitalpals.com",
     "+90 538 646 16 79", "https://www.linkedin.com/company/digitalpals/", "",
     "Shopify + growth-marketing partner; Maslak Kolektif House address = startup-friendly culture.",
     "Partnership", "High"),
    ("Shopiuzman", "Platform Partner", "Istanbul", "https://shopiuzman.com/", "info@shopiuzman.com",
     "+90 538 036 86 52", "", "",
     "Shopify-only certified partner, small team — ideal early adopter for Shopify export.",
     "Partnership", "High"),
    ("Qreate Dijital", "Platform Partner", "Istanbul (Şişli) + Ankara", "https://qreatedijital.com/", "merhaba@qreatedijital.com",
     "+90 850 308 6002", "https://www.linkedin.com/company/qreate-dijital/", "",
     "360° e-commerce + ikas partner — pitch partnership on ikas export once it ships.",
     "Partnership", "High"),

    # --- Ankara ---
    ("Blue Ajans", "Agency", "Ankara (Yenimahalle)", "https://www.blueajans.com.tr/", "",
     "+90 553 939 25 83", "https://www.linkedin.com/company/blue-ajans/", "",
     "Ankara-based web + digital marketing. Use LinkedIn DM + contact form, no public email.",
     "Client", "Medium"),
    ("medyANKA", "Agency", "Ankara (Söğütözü)", "https://www.medyanka.com/", "bilgi@medyanka.com",
     "+90 312 955 05 91", "https://www.linkedin.com/company/medyanka", "Tolga Aksun (Founder)",
     "Founder-led shop, 15+ years — named founder means you can write a personal first-email.",
     "Client", "High"),
    ("Nano Medya", "Agency", "Ankara (Çankaya)", "https://www.nanomedya.com.tr/", "info@nanomedya.com",
     "+90 532 720 32 25", "https://tr.linkedin.com/company/nano-medya-yazilim-ve-reklam-ajansi", "",
     "15+ years, 500+ clients, strong volume — classic 'delivery pipeline' target.",
     "Client", "High"),
    ("Mui Medya", "Agency", "Ankara", "https://muimedya.com/", "info@muimedya.com",
     "+90 312 684 8000", "", "",
     "Dijital ajans, WhatsApp-first — pitch via WhatsApp + email together.",
     "Client", "Medium"),

    # --- İzmir ---
    ("İzmir Web Ajans", "Agency", "İzmir (Balçova)", "https://www.izmirwebajans.com/", "info@izmirwebajans.com",
     "+90 554 363 14 91", "", "",
     "Local İzmir shop — first-message warmth: 'İzmir'li bir tool olarak...'",
     "Client", "Medium"),
    ("Arce Web Ajans", "Agency", "Manisa (Celal Bayar Teknopark)", "https://www.arcewebajans.com/", "info@arceyazilim.com",
     "+90 533 259 96 97", "", "",
     "Teknopark-based — early-adopter leaning. Likely small team.",
     "Client", "Medium"),
    ("İzmir Web", "Agency", "İzmir (Bayraklı)", "https://www.izmirweb.com/", "info@izmirweb.com",
     "+90 535 421 41 33", "", "",
     "Serves internationally from Bayraklı Tower. Delivery-speed pitch.",
     "Client", "Medium"),
    ("Mudiweb", "Agency", "İzmir (Bayraklı)", "https://www.mudiweb.com/", "info@mudiweb.com",
     "+90 532 770 46 23", "", "",
     "Corporate web + custom software since 2008, Google Partner. Decent fit.",
     "Client", "Medium"),
    ("Digital Karınca", "Agency", "İzmir (Karşıyaka)", "https://digitalkarinca.com/", "bilgi@digitalkarinca.com",
     "+90 553 315 62 10", "", "",
     "Notable logos (BCG, BASF, Siemens) — bigger clients = less agile, partnership angle.",
     "Partnership", "Low"),
    ("1007 Medya", "Agency", "İzmir", "https://www.1007medya.com/", "info@1007medya.com",
     "+90 533 260 51 39", "https://www.linkedin.com/company/1007medya", "",
     "10+ years of SMB-focused web work in İzmir — clear fit for freelancer-like delivery.",
     "Client", "High"),
    ("OVARSA Yazılım", "Agency", "İzmir (Bayraklı)", "https://www.ovarsa.com/", "",
     "+90 232 332 37 76", "https://tr.linkedin.com/in/ovarsa-yazılım-teknolojileri-976a8a186", "",
     "İzmir + Manisa offices. No public email — LinkedIn DM first.",
     "Client", "Low"),

    # --- Antalya ---
    ("AWA Web Ajans", "Agency", "Antalya (Muratpaşa)", "https://awawebajans.com/", "info@awawebajans.com",
     "+90 242 324 03 54", "https://www.linkedin.com/company/awa-web-ajans", "",
     "100+ clients, London office — modest size, good outreach target.",
     "Client", "Medium"),
    ("EC Tasarım", "Agency", "Antalya (Muratpaşa)", "https://www.ectasarim.com/", "info@ectasarim.com",
     "+90 532 214 54 63", "", "",
     "SEO + Google Ads + web. Typical small agency workload Yappaflow compresses.",
     "Client", "Medium"),
    ("Doubleyou Agency", "Agency", "Antalya", "https://doubleyou.agency/", "hello@doubleyou.agency",
     "+90 242 247 66 66", "", "",
     "Digital + AI-solutions positioning — AI generator pitch should resonate.",
     "Client", "High"),
    ("Maxantalya", "Agency", "Antalya (Muratpaşa)", "https://maxantalya.com/", "info@maxantalya.com",
     "+90 530 285 08 87", "", "",
     "SMB-focused Antalya shop. Straightforward client pitch.",
     "Client", "Medium"),
    ("Antalya Web Tasarımcı (dijitalmedya07)", "Agency", "Antalya (Muratpaşa)", "https://www.antalyawebtasarimci.com/", "info@dijitalmedya07.com",
     "+90 534 320 58 07", "", "",
     "Google Ads partner, focused on local SMBs — fast-turnaround shop.",
     "Client", "Medium"),
    ("Creamake", "Platform Partner", "Antalya (HQ) + Istanbul + London", "https://creamake.com/", "info@creamake.com",
     "+90 553 663 07 43", "https://www.linkedin.com/company/creamake", "",
     "Authorized ikas partner, 100+ stores launched. PARTNERSHIP angle on ikas export.",
     "Partnership", "High"),

    # --- Bursa ---
    ("GEON Creative Workshop", "Agency", "Bursa", "https://www.geonajans.com/", "",
     "+90 224 443 22 23", "", "",
     "Bursa web + software shop. No public email — call or use contact form.",
     "Client", "Low"),

    # --- Freelancers ---
    ("Doruk Sucuka", "Freelancer", "Istanbul", "https://www.doruksucuka.com.tr/", "hello@doruksucuka.com.tr",
     "+90 532 157 20 66", "https://www.linkedin.com/in/doruk-sucuka-78b66151", "Doruk Sucuka",
     "Freelance since 2004 + explicitly positions as 'AI-destekli dijital çözümler' — warmest possible audience.",
     "Client", "High"),
    ("Ferdi Tarakçı", "Freelancer", "Istanbul (Kadıköy)", "https://ferditarakci.com/", "bilgi@ferditarakci.com",
     "+90 546 831 20 73", "https://www.linkedin.com/in/ferditarakci/", "Ferdi Tarakçı",
     "Vue/React frontend freelancer — technical enough to self-serve Yappaflow output.",
     "Client", "High"),
    ("Tezel Zenginoğlu", "Freelancer", "İzmir", "https://www.tezelzenginoglu.com/", "tezelzenginoglu@gmail.com",
     "+90 553 558 90 41", "https://www.linkedin.com/in/tezel-zenginoğlu/", "Tezel Zenginoğlu",
     "Freelance web + graphic designer, portfolio-led. Care about output quality, so lead with demo.",
     "Client", "High"),
    ("Emrah Özyürek (ankarawt)", "Freelancer", "Ankara", "https://www.ankarawt.com/", "info@ankarawt.com",
     "+90 543 584 76 21", "", "Emrah Özyürek",
     "SEO-focused freelance web designer — Yappaflow saves him the boring parts of each build.",
     "Client", "High"),
    ("Emre Alkaç", "Freelancer", "İzmir (Bornova)", "https://www.emrealkac.com/", "ivan@emrealkac.com",
     "+90 505 257 59 25", "https://www.linkedin.com/in/emre-alkac/", "Emre Alkaç",
     "Freelance graphic + web. Design-side freelancer — will judge on aesthetics.",
     "Client", "Medium"),
    ("Gökmen Bekar", "Freelancer", "Turkey", "https://www.gokmenbekar.com/", "info@gokmenbekar.com",
     "", "https://www.linkedin.com/in/gokmenbekar/", "Gökmen Bekar",
     "Senior UI/UX designer + frontend. High craft expectation; use as design-focused beta user.",
     "Client", "Medium"),
    ("Celalettin Bedir", "Freelancer", "Istanbul (Beyoğlu)", "https://www.celalettinbedir.com/", "info@celalettinbedir.com",
     "+90 535 344 58 74", "https://www.linkedin.com/in/celalettin-bedir-a38284ab/", "Celalettin Bedir",
     "Freelance web dev — clean target.",
     "Client", "High"),
    ("Cem Sevinç", "Freelancer", "Istanbul (Beyoğlu)", "https://cemsevinc.com/", "",
     "+90 532 631 70 83", "", "Cem Sevinç",
     "Freelance web software + design. No public email — WhatsApp/SMS first.",
     "Client", "Medium"),
    ("Onur Turkay", "Freelancer", "Turkey", "https://www.onurturkay.com.tr/", "info@onurturkay.com.tr",
     "+90 544 221 4161", "", "Onur Turkay",
     "Freelance web designer — standard SMB-site workload.",
     "Client", "Medium"),
    ("Uğur Topuz", "Freelancer", "İzmir", "https://www.ugurtopuz.com/", "",
     "", "", "Uğur Topuz",
     "100+ websites across hotel/university/automotive — high volume, time-saving pitch. Email via form.",
     "Client", "High"),
    ("Adnan Saykı", "Freelancer", "Istanbul (Ataşehir)", "https://www.adnansayki.com/", "asayki@gmail.com",
     "+90 537 419 43 23", "", "Adnan Saykı",
     "Self-labeled 'Turkey's first SEO expert', freelance web + SEO — pitch Yappaflow sites' SEO readiness.",
     "Client", "Medium"),
    ("Ferdi Taraçcı (Armut.com top-rated)", "Freelancer", "Istanbul", "https://armut.com/istanbul-web-tasarim", "",
     "", "", "",
     "Armut.com itself is the distribution channel for many Turkish solo web pros — also worth a partnership DM.",
     "Partnership", "Low"),
    ("Bionluk freelancer pool (web yazılım)", "Freelancer", "Online (TR)", "https://bionluk.com/freelancer-bul/web-yazilim", "",
     "", "", "",
     "Bionluk is Turkey's largest freelance marketplace. Not a single lead — earmark for paid-ad + community post.",
     "Partnership", "Low"),

    # ================== ROUND 2 — additional boutique Istanbul shops ==================
    ("Ait Istanbul", "Agency", "Istanbul", "https://aitistanbul.com/", "hello@aitistanbul.com",
     "", "https://www.linkedin.com/company/aitistanbul/", "Eylül Günaydınlar (Founder & ECD)",
     "Independent creative studio, brand-first. Founder-led — write a personal note, not a mass email.",
     "Client", "High"),
    ("PEP Brands", "Agency", "Istanbul", "https://pepistanbul.com/", "hello@pepistanbul.com",
     "", "https://www.linkedin.com/company/pep-brand-agency/", "",
     "Brand agency; small team. Pitch Yappaflow as 'first draft of every client microsite'.",
     "Client", "High"),
    ("Kreatif (Yeni Nesil İletişim)", "Agency", "Istanbul (Şişli)", "https://kreatif.net/", "contact@kreatif.net",
     "+90 212 213 65 55", "https://www.linkedin.com/company/1926746/", "",
     "Multidisciplinary strategy + marketing + design shop — good partnership-angle candidate for microsite work.",
     "Partnership", "Medium"),
    ("MARK-A", "Agency", "Istanbul + Afyon + Kuşadası", "https://mark-a.com.tr/", "info@mark-a.com.tr",
     "+90 212 982 82 89", "https://www.linkedin.com/company/markacomtr", "Mehmet Dinler (Founder)",
     "Founder wrote a digital-marketing book; receptive to AI tooling. Warm founder-level intro.",
     "Client", "High"),
    ("Fikir Tasarım Atölyesi", "Agency", "Istanbul (Levent)", "https://fikirtasarimatolyesi.com/", "",
     "+90 533 252 5691", "https://www.linkedin.com/company/fikir-tasar%C4%B1m-at%C3%B6lyesi", "",
     "360° ad agency since 2014. No public email on site — use LinkedIn company DM + WhatsApp.",
     "Client", "Medium"),
    ("Poligon Interactive", "Agency", "Istanbul + Canada", "https://www.poligoninteractive.com/", "",
     "", "https://www.linkedin.com/company/poligoninteractive-tr/", "",
     "Renault/Nissan/Barilla clients — larger shop, partnership angle for microsite delivery.",
     "Partnership", "Low"),
    ("Ceviz Bilişim", "Agency", "Istanbul (Esenyurt)", "https://cevizbilisim.com.tr/", "info@cevizbilisim.com.tr",
     "+90 212 807 00 06", "https://www.linkedin.com/showcase/ceviz-bilisim/", "",
     "Web + e-commerce + SEO shop. Classic mid-size delivery target for the client angle.",
     "Client", "Medium"),
    ("Master Tasarım", "Agency", "Istanbul + Amsterdam", "https://mastertasarim.com/", "",
     "+90 850 242 4 393", "", "",
     "Google Partner, 750+ references, multi-office. Partnership angle — pitch microsite pipeline.",
     "Partnership", "Low"),
    ("Vagonmedya", "Agency", "Kocaeli (Başiskele)", "https://vagonmedya.com/", "iletisim@vagonmedya.com",
     "", "", "",
     "Runs multiple internet projects, ~8M monthly uniques. Partnership angle — they scale content, we scale sites.",
     "Partnership", "Medium"),
    ("Ali Tarhan Dijital Atölye", "Agency", "Istanbul", "https://www.alitarhan.com.tr/", "ad@alitarhan.com.tr",
     "+90 532 523 82 83", "", "Ali Tarhan (Director)",
     "Director-led boutique ad shop; ad@ is his direct address. Personal first-email possible.",
     "Client", "High"),

    # ================== ROUND 2 — Gaziantep ==================
    ("VEMEDYA", "Agency", "Gaziantep (Teknopark)", "https://vemedya.com/", "info@vemedya.com",
     "+90 342 221 03 70", "", "",
     "20 years, Teknopark-based Gaziantep shop. Teknopark tenants are usually early-adopter friendly.",
     "Client", "High"),
    ("Aydıngüler Medya", "Agency", "Gaziantep (Şehitkamil)", "https://aydingulermedya.com/", "info@aydingulermedya.com",
     "+90 538 581 22 12", "", "",
     "Small web + SEO + social agency. Straightforward client pitch; easy to land a meeting.",
     "Client", "Medium"),
    ("Kaktüs Creative (Kaktüs Medya)", "Agency", "Gaziantep", "https://kaktusmedya.com/", "brain@kaktusmedya.com",
     "+90 342 255 55 43", "https://tr.linkedin.com/company/kaktusmedya", "Servet (Founder)",
     "Since 2013. brain@ is a friendly inbound inbox — Yappaflow is a natural 'speed-up delivery' fit.",
     "Client", "High"),
    ("Memsidea Dijital Ajans", "Agency", "Gaziantep", "https://memsidea.com/", "hello@memsidea.com",
     "+90 507 487 2098", "", "",
     "187+ brands, 8 years, web+mobile+e-commerce — active delivery shop.",
     "Client", "High"),
    ("Gazisoft Yazılım", "Agency", "Gaziantep (OSB Teknokent)", "https://gazisoft.com.tr/", "info@gazisoft.com.tr",
     "+90 532 503 29 09", "", "",
     "Teknokent-based software + web shop. Leans dev-heavy — they'll appreciate the ZIP export.",
     "Client", "Medium"),

    # ================== ROUND 2 — Kayseri ==================
    ("Seven Ajans (Seven Creative Design)", "Agency", "Kayseri (Melikgazi)", "https://www.sevenajans.com/", "info@sevenajans.com",
     "+90 532 163 38 82", "", "",
     "SMB + manufacturing-client focus, offers 1-year support guarantee — values reliable delivery.",
     "Client", "High"),
    ("Galaksi Medya", "Agency", "Kayseri (Kocasinan)", "https://galaksimedya.com/", "info@galaksimedya.com",
     "+90 546 847 69 20", "", "",
     "Since 2008. Small, stable regional shop — easy cold-email target.",
     "Client", "Medium"),
    ("Medyatör İnteraktif", "Agency", "Kayseri (Erciyes Teknopark)", "https://medyator.net/", "info@medyator.net",
     "+90 352 224 42 20", "https://www.linkedin.com/in/medyator/", "",
     "Erciyes Teknopark tenant — academic-adjacent, curious crowd; solid demo-first target.",
     "Client", "High"),

    # ================== ROUND 2 — Mersin / Konya / Sakarya / Eskişehir / Denizli ==================
    ("Çağ Medya", "Agency", "Mersin (HQ) + Istanbul", "https://cagmedya.com/", "info@cagmedya.com",
     "+90 324 336 68 50", "", "",
     "20-yr Mersin shop with Istanbul branch — multi-city delivery where ZIP hand-off shines.",
     "Client", "High"),
    ("Öncü Web Tasarım", "Agency", "Denizli", "https://www.oncuweb.com/", "info@oncuweb.com",
     "+90 542 385 66 20", "", "",
     "Since 2002, small Denizli shop. Simple, reliable client pitch.",
     "Client", "Medium"),
    ("AWT Bilişim (Denizli Web / Ankara Web / İzmir Web)", "Agency", "Denizli + Ankara + Istanbul + İzmir", "https://denizliweb.com.tr/", "info@denizliweb.com.tr",
     "+90 532 482 31 20", "", "",
     "Multi-city 'city+web' branding — built for volume. Huge fit for a generator pipeline.",
     "Client", "High"),
    ("BenefitAd Reklam Ajansı", "Agency", "Denizli", "https://benefitads.com/", "benefitreklam@gmail.com",
     "+90 532 420 21 51", "", "",
     "Small Denizli agency, direct email. Low-friction first outreach.",
     "Client", "Medium"),
    ("Yobisi 360° Reklam", "Agency", "Denizli", "https://yobisi.com/", "",
     "+90 258 244 01 23", "https://www.linkedin.com/company/18034199", "",
     "8-year Denizli 360° agency. Reach via LinkedIn DM first, no public email.",
     "Client", "Medium"),

    # ================== ROUND 2 — İzmir ==================
    ("İmgesel Tasarım", "Agency", "İzmir (Karşıyaka)", "https://www.imgeseltasarim.com/", "satis@imgeseltasarim.com",
     "+90 537 381 73 57", "https://www.linkedin.com/company/imgesel-tasarim", "",
     "Brand + ad agency. satis@ = warm sales inbox.",
     "Client", "Medium"),
    ("Deniz Web Ajans", "Agency", "İzmir (Çiğli)", "https://deniz-web.com/", "info@denizweb.com.tr",
     "+90 507 868 27 75", "", "",
     "360° digital agency since 2018 — young, hungry, willing to try new tools.",
     "Client", "High"),

    # ================== ROUND 2 — Ankara / Polatlı ==================
    ("101 Medya Web Stüdyosu", "Agency", "Polatlı (Ankara)", "https://101medya.com/", "info@101medya.com",
     "+90 533 1919 0101", "", "",
     "Small web studio serving TR+abroad — right size for Yappaflow to feel immediate.",
     "Client", "Medium"),
    ("Sıradışı Digital", "Agency", "Ankara + Eskişehir", "https://siradisidigital.com/", "info@siradisidigital.com",
     "+90 850 304 33 34", "", "",
     "Two-city 360° agency. Enough scale for a partnership conversation on microsite delivery.",
     "Partnership", "Medium"),

    # ================== ROUND 2 — Konya ==================
    ("INVIVA Medya", "Agency", "Konya (HQ) + Istanbul", "https://inviva.com.tr/", "info@inviva.com.tr",
     "+90 332 321 1125", "", "",
     "Two-office web + software + ad agency. Mid-size; offer free first-5-sites as the hook.",
     "Client", "Medium"),
    ("Post Ajans", "Agency", "Konya (Meram)", "https://www.postajans.com.tr/", "postajans@postajans.com.tr",
     "+90 332 238 32 60", "https://www.linkedin.com/company/post-ajans/", "",
     "25 years, 4000+ projects, 9 awards. Bigger shop — partnership angle for microsite add-on.",
     "Partnership", "Medium"),
    ("Konya Web Tasarım (A&G / Ali Gökalp)", "Freelancer", "Konya + İzmir", "https://aligokalp.com.tr/", "",
     "+90 507 736 26 88", "https://www.linkedin.com/in/ali-g%C3%B6kalp/", "Ali Gökalp",
     "Solo operator running two-city brand. Exactly the 'deliver like a studio' pitch.",
     "Client", "High"),

    # ================== ROUND 2 — Eskişehir ==================
    ("Pill Digital", "Agency", "Eskişehir (Odunpazarı)", "https://www.pill.com.tr/", "",
     "+90 554 860 97 60", "https://linkedin.com/company/pill-digital", "",
     "Web + e-commerce + mobile shop since 2016. No public email — LinkedIn DM first.",
     "Client", "Medium"),
    ("Webmarka", "Agency", "Eskişehir (Odunpazarı)", "https://webmarka.com/", "destek@webmarka.com",
     "+90 552 009 20 29", "", "",
     "Small Eskişehir web+software+ads shop. Cold email should convert.",
     "Client", "Medium"),
    ("Forse Reklam", "Agency", "Eskişehir (Odunpazarı)", "https://www.forse.web.tr/", "forse@forse.web.tr",
     "+90 850 302 55 99", "https://www.linkedin.com/in/forse-reklam-b617b5155/", "",
     "Since 1996 — old-guard Eskişehir shop. Pitch modernization via AI-driven generation.",
     "Client", "Medium"),
    ("Saruhan Web Ajans (SW)", "Agency", "Eskişehir", "https://www.saruhanweb.com/", "bilgi@saruhanweb.com",
     "+90 222 220 03 77", "https://tr.linkedin.com/company/saruhanwebajans", "",
     "Already positions itself as 'AI-powered design' — perfect warm audience.",
     "Client", "High"),

    # ================== ROUND 2 — Sakarya ==================
    ("Trigon Software", "Agency", "Sakarya (Adapazarı)", "https://trigonsoftware.com/", "",
     "+90 540 272 80 80", "https://www.linkedin.com/company/trigonsoftware", "",
     "Sakarya software shop. LinkedIn DM first; no public email.",
     "Client", "Medium"),

    # ================== ROUND 2 — additional freelancers ==================
    ("İbrahim Etem Kanbur (Özgür Tasarımcı)", "Freelancer", "Turkey", "https://ozgurtasarimci.com/", "",
     "+90 536 322 73 82", "https://www.linkedin.com/in/ozgurtasarimci/", "İbrahim Etem Kanbur",
     "8 yrs, freelance graphic + web + brand. Reach via LinkedIn DM.",
     "Client", "Medium"),
    ("Çağatay Demir", "Freelancer", "Turkey (works globally)", "https://cagataydemir.com.tr/", "merhaba@cagataydemir.com.tr",
     "+90 530 393 01 71", "https://www.linkedin.com/in/cagatay-demir/", "Çağatay Demir",
     "Multi-country WordPress + SEO freelancer. Will appreciate white-label ZIP export.",
     "Client", "High"),
    ("Dream Office Team", "Freelancer", "Turkey", "https://dreamoffice.com.tr/", "bilgi@dreamoffice.com.tr",
     "+90 544 543 13 33", "", "",
     "Solo/small team, custom web + SEO — the 'save the boring parts' pitch lands.",
     "Client", "Medium"),

    # ================== ROUND 2 — more regional agencies ==================
    ("Eskişehir Web Ajans", "Agency", "Eskişehir", "https://eskisehirwebajans.com/", "iletisim@eskisehirwebajans.com",
     "+90 505 387 05 80", "", "",
     "250 projects, 80 clients, 20 yrs. Steady regional shop — standard client pitch.",
     "Client", "Medium"),
    ("Vizyoner Ajans Danışmanlık", "Agency", "Konya (Selçuklu)", "https://www.vizyoner.com.tr/", "info@vizyoner.com.tr",
     "+90 541 924 24 22", "https://www.linkedin.com/company/vizyoner-danışmanlık-&-ajans/", "",
     "Since 2005. Consulting + agency — partnership-angle friendly.",
     "Partnership", "Medium"),
    ("Webacil", "Agency", "Istanbul", "https://webacil.com/", "",
     "+90 216 330 10 50", "", "",
     "Freelance-style 'fast web' brand. No public email — call or contact form.",
     "Client", "Medium"),
    ("WebAcil freelance pool", "Freelancer", "Istanbul", "https://webacil.com/freelance-web-tasarimci/", "",
     "+90 216 330 10 50", "", "",
     "Curated freelance web pool. Partnership pitch — Yappaflow as their delivery layer.",
     "Partnership", "Low"),

    # ================== ROUND 2 — directory / community leverage (not a single lead) ==================
    ("Kolektif House Maslak community", "Partnership", "Istanbul (Maslak)", "https://kolektifhouse.co/", "",
     "", "", "",
     "Shared office popular with small agencies + freelancers. Organic: pitch a free lunch-&-learn.",
     "Partnership", "Low"),
    ("Kodluyoruz Discord / community", "Partnership", "Online (TR)", "https://kodluyoruz.org/", "",
     "", "https://www.linkedin.com/company/kodluyoruzorg/", "",
     "Turkey's biggest free-coding community — bootcamp grads regularly go freelance. Post a Yappaflow build-along.",
     "Partnership", "Medium"),
    ("Türkiye Dijital Ajanslar (LinkedIn group)", "Partnership", "Online (TR)", "https://www.linkedin.com/groups/3903174/", "",
     "", "", "",
     "LinkedIn group for Turkish digital agencies — post a short Yappaflow demo there as a 'tool I built' note.",
     "Partnership", "Medium"),
    ("Indie Hackers Turkey (community)", "Partnership", "Online (TR)", "https://indiehackers.com.tr/", "",
     "", "", "",
     "Turkish indie founders + freelancers. Write a build-in-public thread about Yappaflow; drives warm inbound.",
     "Partnership", "Medium"),

    # ================== ROUND 3 — Trabzon + Black Sea region ==================
    ("WebTrabzon", "Agency", "Trabzon (Ortahisar)", "https://www.webtrabzon.com/", "bilgi@webtrabzon.com",
     "+90 554 496 00 61", "https://www.linkedin.com/company/webtrabzon/", "",
     "Small Trabzon web shop — direct email, solid cold-outreach target.",
     "Client", "Medium"),
    ("61Medya", "Agency", "Trabzon (Ortahisar)", "https://www.61medya.com.tr/", "",
     "+90 553 416 52 70", "", "",
     "Trabzon web + matbaa + reklam ajansı. Try WhatsApp first; no public email.",
     "Client", "Medium"),
    ("Emre Domaç", "Freelancer", "Giresun (Tirebolu)", "https://www.emrewebtasarim.com/", "bilgi@emrewebtasarim.com",
     "+90 552 239 82 85", "https://www.linkedin.com/in/emre-domac/", "Emre Domaç",
     "16 yrs + 400 sites. Solo veteran — white-label ZIP is an easy sell.",
     "Client", "High"),

    # ================== ROUND 3 — Samsun ==================
    ("Dijitasyon", "Agency", "Samsun", "https://dijitasyon.com.tr/", "web@dijitasyon.com.tr",
     "+90 530 642 69 55", "", "Muratcan & Hazel Karataş (Founders)",
     "Husband-and-wife founder team. Write to Hazel on LinkedIn first.",
     "Client", "High"),
    ("Redia Ajans (medya.red)", "Agency", "Samsun", "https://www.medya.red/", "bilgi@medya.red",
     "+90 555 888 55 24", "https://linkedin.com/company/redia-ajans/", "",
     "Since 2006, 200+ client sites. Stable regional shop.",
     "Client", "Medium"),
    ("SIP Yazılım (Samsun Web Tasarım)", "Agency", "Samsun (Teknopark)", "https://www.samsunwebtasarimi.com/", "",
     "+90 850 302 43 55", "", "",
     "OMÜ Teknopark tenant. Call to reach; no public email.",
     "Client", "Medium"),
    ("bw/a (better with agency)", "Agency", "Samsun", "https://bw.agency/", "",
     "", "", "",
     "Boutique 'local-entrepreneur-only' agency. Warm fit for pitching Yappaflow via demo, no email — use contact form.",
     "Client", "Medium"),

    # ================== ROUND 3 — Adana / Mersin ==================
    ("Adana Tasarım", "Agency", "Adana (Seyhan)", "https://www.adanatasarim.com/", "info@adanatasarim.com",
     "+90 546 787 49 32", "", "",
     "Full-stack Adana web+branding shop. Direct email inbox is warm.",
     "Client", "Medium"),
    ("Gerçek Ajans (Gerçek Bilişim)", "Agency", "Adana (Çukurova)", "https://www.gercekbilisim.com/", "info@gercekbilisim.com",
     "+90 505 020 10 10", "https://tr.linkedin.com/company/adana-web-tasarim", "",
     "Operates from Go Ofis coworking — early-adopter friendly.",
     "Client", "High"),
    ("Mag-Net Web", "Agency", "Adana (Çukurova)", "https://www.magnetweb.com.tr/", "posta@mag-net.com.tr",
     "+90 322 235 57 53", "", "",
     "24 yrs, 5000+ clients. Mid-size shop — partnership angle viable.",
     "Partnership", "Medium"),
    ("SerNis Dijital Ajans", "Agency", "Adana (Seyhan)", "https://www.sernisdijitalajans.com/", "sernisdijitalajans@gmail.com",
     "+90 551 036 60 63", "https://www.linkedin.com/in/sernis-dijital-ajans-undefined-694912317/", "",
     "Boutique Adana agency on gmail — almost certainly solo/duo, fast decision path.",
     "Client", "Medium"),
    ("EfeDizayn", "Agency", "Mersin (Akdeniz)", "https://www.efedizayn.com/", "destek@efedizayn.com",
     "+90 551 816 88 88", "https://linkedin.com/in/efedizayn-webtasarim/", "",
     "Small Mersin web+branding shop. Direct outreach works.",
     "Client", "Medium"),

    # ================== ROUND 3 — Bodrum / Muğla ==================
    ("Ajans Bukalemun", "Agency", "Muğla (Milas)", "https://ajansbukalemun.com/", "info@ajansbukalemun.com",
     "+90 252 512 39 39", "", "",
     "Milas + Bodrum — hospitality/tourism client base. Pitch seasonal-site speed.",
     "Client", "Medium"),
    ("UçarSoft", "Agency", "Bodrum (HQ) + Istanbul", "https://www.ucarsoft.com/", "info@ucarsoft.com",
     "+90 252 313 27 25", "", "",
     "14 yrs, two-office shop. Multi-city clients — ZIP deploy is ideal.",
     "Client", "Medium"),
    ("Toprak Ajans", "Agency", "Muğla (Bodrum)", "https://www.toprakajans.com/", "info@toprakajans.com",
     "+90 252 316 24 94", "", "",
     "Bodrum-area boutique. Hotel/restaurant clients — fast seasonal turnarounds.",
     "Client", "Medium"),
    ("BuDRoM Ajans", "Agency", "Muğla (Bodrum)", "https://budrom.com.tr/", "info@budrom.com.tr",
     "+90 505 335 55 75", "", "",
     "Bodrum 360° ad agency. Social-video lead + web design.",
     "Client", "Medium"),

    # ================== ROUND 3 — Bursa ==================
    ("Bursa Dijital Ajans", "Agency", "Bursa (Nilüfer)", "https://www.bursadijitalajans.com/", "",
     "+90 532 730 66 86", "", "",
     "13 yrs. WhatsApp-first shop. No public email — reach via phone/LinkedIn.",
     "Client", "Medium"),
    ("Ajans Ay", "Agency", "Bursa (Nilüfer)", "https://ajansay.com/", "info@ajansay.com",
     "+90 540 007 77 16", "", "Bedrettin Ayvaz",
     "700+ clients, 16 yrs. Mid-size Bursa shop. Partnership angle fits.",
     "Partnership", "Medium"),
    ("Webkod", "Agency", "Bursa (Nilüfer)", "https://www.webkod.com.tr/", "hello@webkod.com.tr",
     "+90 224 532 12 53", "", "",
     "Bursa web+SEO+domain shop. hello@ inbox = warm target.",
     "Client", "Medium"),
    ("Aven Dijital", "Agency", "Bursa (İnegöl) + London", "https://www.avendijital.com/", "hello@avendijital.com",
     "+90 506 116 14 53", "", "",
     "İnegöl + UK office, many direct numbers. Small-team structure, likely open to new tooling.",
     "Client", "High"),

    # ================== ROUND 3 — Balıkesir ==================
    ("Derin Ajans", "Agency", "Balıkesir (Altıeylül)", "https://derinajans.com.tr/", "bilgi@derinajans.com",
     "+90 266 241 44 11", "https://tr.linkedin.com/in/derin-ajans-80b2b2205", "",
     "Small Balıkesir agency — local SMB focus. Easy cold email.",
     "Client", "Medium"),
    ("Creamive", "Agency", "Balıkesir + Istanbul", "https://www.creamive.com/", "hello@creamive.com",
     "", "", "",
     "Two-city creative shop. hello@ inbox = warm.",
     "Client", "Medium"),

    # ================== ROUND 3 — Gaziantep / Malatya ==================
    ("Yonkasoft", "Agency", "Gaziantep", "https://yonkasoft.com/", "",
     "+90 537 311 98 17", "https://linkedin.com/company/yonkasoft", "",
     "Gaziantep software+web shop. LinkedIn DM or phone first.",
     "Client", "Medium"),

    # ================== ROUND 3 — Diyarbakır ==================
    ("İnotek Ajans", "Agency", "Diyarbakır + Istanbul + İzmir", "https://inotekajans.com/", "info@inotek.com.tr",
     "+90 850 840 54 12", "", "",
     "Three-city creative+software shop. Multi-site volume is classic Yappaflow fit.",
     "Client", "Medium"),

    # ================== ROUND 3 — KKTC (Northern Cyprus) ==================
    ("Ata Bilişim", "Agency", "Lefkoşa (KKTC)", "https://www.atabilisim.pro/", "info@atabilisim.pro",
     "+90 533 882 30 30", "https://cy.linkedin.com/in/atabilisim", "",
     "650+ clients across TR/CY/Japan/UK/DE. Cross-border delivery = ZIP hand-off wins.",
     "Client", "High"),

    # ================== ROUND 3 — Istanbul branding studios (boutique) ==================
    ("MarkaWorks", "Agency", "Istanbul + London + Dubai + Antalya", "https://markaworks.com/", "contact@markaworks.com",
     "+44 20 3885 80 18", "https://linkedin.com/company/markaworks/", "",
     "Global branding agency. Partnership angle for microsite delivery layer.",
     "Partnership", "Medium"),
    ("UBF (Unique Brand Factory)", "Agency", "Istanbul (Beyoğlu/Şişhane)", "https://www.ubf.com.tr/", "",
     "", "https://linkedin.com/company/ubf---unique-brand-factory", "",
     "Boutique strategy+comms+design. Use LinkedIn DM; no public email.",
     "Client", "Medium"),
    ("Starter", "Agency", "Istanbul", "https://starterdesign.co/", "info@starterdesign.co",
     "", "https://linkedin.com/company/starterdesign-co", "Koray Şahan (Founder)",
     "Founder-led boutique brand agency. Personal note to Koray on LinkedIn.",
     "Client", "High"),
    ("brain.work", "Agency", "Istanbul (Şişli)", "https://brain.work/", "merhaba@brain.work",
     "+90 532 236 38 80", "https://linkedin.com/company/braindotwork", "",
     "UX+web+SEO+brand. merhaba@ inbox is inviting.",
     "Client", "High"),
    ("Userspots", "Agency", "Istanbul (Turkey)", "https://userspots.com/", "",
     "", "https://linkedin.com/company/userspots", "",
     "Enterprise UX clients (Arçelik/Boyner/Migros). Partnership-angle only; they build, Yappaflow supplies microsites.",
     "Partnership", "Low"),

    # ================== ROUND 3 — Shopify / ikas / Webflow partners ==================
    ("Prix Studio", "Platform Partner", "Istanbul (İTÜ Arı 3 Teknokent)", "https://prix-studio.com/", "info@prix-studio.com",
     "", "https://linkedin.com/company/prix-studio/", "",
     "Webflow + Shopify + ikas — certified on all three. Strongest partnership target.",
     "Partnership", "High"),
    ("Artelio Creative", "Platform Partner", "Istanbul", "https://www.arteliocreative.com/", "",
     "+90 216 606 70 18", "https://linkedin.com/company/arteliocreative/", "",
     "ikas partner + e-com growth focus. Pitch Yappaflow ikas export.",
     "Partnership", "High"),
    ("YCF Digital", "Platform Partner", "İzmir (Bornova) + USA", "https://ycfdigital.com/", "sales@ycfdigital.com",
     "+90 531 400 39 39", "https://linkedin.com/company/ycfdigital", "Mansur Sarı (Founder)",
     "Official ikas partner + Shopify + Amazon. Two-office, named founder — premium partnership target.",
     "Partnership", "High"),
    ("Alis Dijital", "Platform Partner", "Kayseri + Istanbul", "https://alisdijital.com/", "ajans@alisdijital.com",
     "+90 850 308 80 52", "https://linkedin.com/company/alisdijital", "",
     "200+ stores, verified on ikas+Shopify+Meta+Google. Big partnership prize.",
     "Partnership", "High"),
    ("DigiFist", "Platform Partner", "Türkiye + Belgium + UAE", "https://digifist.com/", "",
     "", "https://linkedin.com/company/digifist", "",
     "First Shopify Premier Partner in Turkey (granted 2025). High-profile partnership if landed.",
     "Partnership", "High"),
    ("Roicool", "Platform Partner", "Istanbul (Maslak)", "https://roicool.com/", "info@roicool.com",
     "+90 531 407 28 45", "https://www.linkedin.com/company/roicool/", "Mahmud Filoğlu & Zehra Adaş",
     "Webflow + performance-marketing shop. Co-founders — good two-person intro.",
     "Partnership", "High"),
    ("Commerwise", "Platform Partner", "Bursa", "https://commerwise.com/", "info@commerwise.com",
     "+90 546 814 25 34", "", "",
     "Shopify/WordPress + CRO focus. Boutique partner; pitch Shopify export early.",
     "Partnership", "Medium"),
    ("Drupart Dijital", "Platform Partner", "Kocaeli (Gebze) + Frankfurt + Dublin", "https://drupart.com.tr/", "",
     "+90 262 678 88 72", "https://linkedin.com/company/drupart", "",
     "Drupal agency, teknopark-based. Partnership-angle for non-Drupal client work.",
     "Partnership", "Medium"),

    # ================== ROUND 3 — other Istanbul mid agencies ==================
    ("MyFC Yazılım", "Agency", "Istanbul (Ataşehir)", "https://www.myfcyazilim.com/", "info@myfcyazilim.com",
     "+90 531 208 98 01", "https://linkedin.com/company/myfcagency", "",
     "15+ yrs, Palladium Tower address. Standard client pitch.",
     "Client", "Medium"),
    ("Magna Dijital", "Agency", "Istanbul (Beşiktaş) + Atlanta", "https://www.magnadijital.com.tr/", "info@magnadijital.com.tr",
     "+90 850 333 80 91", "https://tr.linkedin.com/company/magna-dijital-pazarlama-ajans%C4%B1-ve-dan%C4%B1%C5%9Fmanl%C4%B1k-hizmetleri-ltd-%C5%9Fti", "",
     "Two-country agency. Cross-border client sites = ZIP export hand-off wins.",
     "Client", "Medium"),

    # ================== ROUND 3 — Senior freelancer ==================
    ("Ali Çınaroğlu", "Freelancer", "Istanbul", "https://alicinaroglu.com/", "me@alicinaroglu.dev",
     "", "", "Ali Çınaroğlu",
     "19+ yrs senior SaaS/iOS engineer, remote-open. Pitch as white-label build partner.",
     "Client", "Medium"),

    # ================== ROUND 3b — Kocaeli / Ankara / more Istanbul partners ==================
    ("PAM Ajans", "Agency", "Ankara + Dubai", "https://pamajans.com/", "info@pamajans.com",
     "+90 850 305 77 26", "", "",
     "Full-service digital agency with Dubai office. Export-to-client workflow = good fit.",
     "Client", "Medium"),
    ("Digipeak", "Agency", "Istanbul + London + Texas", "https://digipeak.org/", "info@digipeak.org",
     "", "https://www.linkedin.com/company/digipeakagency/", "Ufuk Yalvaç",
     "Founder Ufuk Yalvaç is reachable on LinkedIn — multi-country shop, open to tooling.",
     "Client", "Medium"),
    ("Kocaeli Dijital", "Agency", "Kocaeli (İzmit)", "https://kocaelidijital.com/", "info@kocaelidijital.com",
     "+90 850 303 02 41", "", "",
     "Regional SMB agency; Yappaflow cuts their delivery timeline dramatically.",
     "Client", "High"),
    ("Essente Bilişim", "Agency", "Kocaeli (İzmit) + Canada", "https://essentebilisim.com/", "info@essentebilisim.com",
     "+90 262 322 41 12", "", "",
     "Cross-border SMB shop. Export ZIP hand-off = clean deliverable for Canadian SMB clients.",
     "Client", "Medium"),
    ("Piyetra", "Agency", "Kocaeli", "https://piyetra.com/", "merhaba@piyetra.com",
     "+90 535 816 5388", "", "",
     "Small Kocaeli boutique. Personal email + founder-led — warm intro works.",
     "Client", "Medium"),
    ("LF Dijital", "Agency", "Kocaeli (İzmit)", "https://lf.com.tr/", "info@lf.com.tr",
     "+90 533 704 74 83", "", "",
     "Regional SMB agency. Pitch as light, rapid build/deploy alternative.",
     "Client", "Medium"),

    # ================== ROUND 3c — more Shopify / ikas / Istanbul boutiques ==================
    ("Swonie Creative", "Platform Partner", "Izmir", "https://swonie.com/", "hello@swonie.com",
     "", "https://www.linkedin.com/company/swonie/", "",
     "Shopify + branding boutique in Izmir. Great fit for Shopify-export beta once shipped.",
     "Partnership", "High"),
    ("Nutima", "Platform Partner", "Istanbul", "https://nutima.co/", "info@nutima.co",
     "", "https://www.linkedin.com/company/nutima/", "",
     "Shopify Plus-minded boutique. Partnership pitch: Yappaflow = fast prototype to Shopify.",
     "Partnership", "Medium"),
    ("Ufuk Yılmaz (Dribbble)", "Freelancer", "Istanbul", "https://dribbble.com/ufukyilmaz", "",
     "", "https://www.linkedin.com/in/ufukyilmaz/", "Ufuk Yılmaz",
     "Dribbble-visible product designer. LinkedIn DM — pitch as design hand-off accelerator.",
     "Client", "Medium"),
    ("Murat Tekmen (Dribbble)", "Freelancer", "Istanbul", "https://dribbble.com/murattekmen", "",
     "", "https://www.linkedin.com/in/murattekmen/", "Murat Tekmen",
     "Product/web designer on Dribbble. LinkedIn outreach — Yappaflow turns his designs into sites.",
     "Client", "Medium"),
    ("Tolga Taşçı (Dribbble)", "Freelancer", "Istanbul", "https://dribbble.com/tolgatasci", "",
     "", "https://www.linkedin.com/in/tolgatasci/", "Tolga Taşçı",
     "Visual designer on Dribbble. Outreach via LinkedIn for small-business delivery partner.",
     "Client", "Low"),

    # ================== ROUND 3d — Kayseri / Eskişehir / Anatolia agencies ==================
    ("Medyatör İnteraktif", "Agency", "Kayseri", "https://www.medyator.net/", "info@medyator.net",
     "+90 850 888 00 38", "", "Fatih Bey (manager)",
     "Erciyes Üniversitesi Teknopark; award-winning regional shop. Perfect client fit for SMB sites.",
     "Client", "High"),
    ("Seven Ajans", "Agency", "Kayseri", "https://www.sevenajans.com/", "info@sevenajans.com",
     "+90 532 163 38 82", "", "",
     "Corporate + e-commerce expert; 1-year support guarantee. Client pitch.",
     "Client", "Medium"),
    ("A2A Dijital", "Agency", "Kayseri", "https://a2adijital.com/", "",
     "+90 352 320 70 78", "", "",
     "13 yrs, 237+ clients. No public email — use phone or contact form.",
     "Client", "Medium"),
    ("Themedya", "Agency", "Kayseri", "https://themedya.com/", "info@themedya.com",
     "+90 539 724 72 89", "", "",
     "Erciyes Teknopark kuluçka-based web/SEO shop. Regional SMB target.",
     "Client", "Medium"),
    ("Eskişehir Web Ajans", "Agency", "Eskişehir", "https://www.eskisehirwebajans.com/", "iletisim@eskisehirwebajans.com",
     "+90 505 387 05 80", "", "",
     "Eskişehir leader per own claim, 250 projects. Good regional SMB pitch.",
     "Client", "Medium"),
    ("Master Web Tasarım", "Agency", "Istanbul (Kağıthane) + Amsterdam", "https://mastertasarim.com/", "",
     "+90 850 242 43 93", "", "",
     "Google Partner; 750+ refs; NL office. Cross-border client ZIP hand-off fit.",
     "Client", "Medium"),
    ("Enteresan Çizgi", "Agency", "Istanbul", "https://www.enteresancizgi.com/", "",
     "", "", "",
     "Automotive-heavy portfolio (Hyundai/Nissan/Renault). Outreach via website contact form.",
     "Client", "Low"),

    # ================== ROUND 3d — More verified Shopify partners ==================
    ("Nodus Works", "Platform Partner", "Istanbul", "https://nodusworks.com/", "hub@nodusworks.com",
     "+90 850 888 28 15", "https://www.linkedin.com/company/nodusworks/", "Eren & Efe",
     "Shopify Partner since 2020, 250+ Shopify/Plus projects, multi-country coverage.",
     "Partnership", "High"),
    ("Digital Pals", "Platform Partner", "Istanbul", "https://digitalpals.com/", "partners@digitalpals.com",
     "", "https://www.linkedin.com/company/digital-pals/", "Kaan",
     "Plus-tier Shopify Partner (since 2019), 150+ brands. partners@ email = warm partnership intro.",
     "Partnership", "High"),
    ("Shopiuzman", "Platform Partner", "Istanbul", "https://shopiuzman.com/", "info@shopiuzman.com",
     "+90 538 036 86 52", "https://www.linkedin.com/company/shopiuzman/", "Uğur",
     "Officially certified Shopify-only agency. Small team, good partnership fit.",
     "Partnership", "High"),

    # ================== ROUND 3e — Antalya / Izmir / more regional ==================
    ("Marinos Ajans", "Agency", "Antalya (Muratpaşa)", "https://www.marinosajans.com.tr/", "",
     "+90 540 571 07 07", "", "",
     "Antalya web + Google Ads shop. Phone-first outreach — SMB target.",
     "Client", "Medium"),
    ("Ceviz Bilişim", "Agency", "Istanbul (Esenyurt)", "https://www.cevizbilisim.com.tr/", "info@cevizbilisim.com.tr",
     "+90 212 807 00 06", "", "",
     "22 years in market (since 2003). Full-service — standard SMB delivery fit.",
     "Client", "Medium"),
    ("OVARSA Yazılım", "Agency", "Izmir (Bayraklı) + Manisa", "https://www.ovarsa.com/", "",
     "+90 232 332 37 76", "", "",
     "Since 2010, serves 30+ countries. Folkart Towers office. Cross-border client fit.",
     "Client", "Medium"),
    ("Goiz Dijital", "Agency", "Izmir (Bornova)", "https://www.goizdijital.com/", "info@goizdijital.com",
     "+90 535 363 35 00", "", "",
     "Since 2004 — corporate web + Google Ads. Regional SMB client target.",
     "Client", "Medium"),
    ("Localveri Yazılım", "Agency", "Izmir (Alsancak) + Istanbul", "https://www.localveri.com.tr/", "",
     "+90 232 265 18 18", "", "",
     "22 yrs, 850+ clients, 4000+ projects. Volume shop — client angle, Yappaflow speeds delivery.",
     "Client", "Medium"),
    ("Digital Karınca", "Agency", "Izmir (Karşıyaka)", "https://digitalkarinca.com/", "bilgi@digitalkarinca.com",
     "+90 553 315 62 10", "", "",
     "Boutique Karşıyaka shop. Good SMB client fit.",
     "Client", "Medium"),

    # ================== ROUND 3f — Ankara digital + more ikas partners ==================
    ("Prix Studio", "Platform Partner", "Istanbul (Sarıyer, İTÜ Arı 3)", "https://prix-studio.com/", "",
     "", "https://www.linkedin.com/company/prixstudio/", "",
     "ikas + Shopify + Webflow partner, 12 yrs. Teknopark address. Use website Teklif Al form.",
     "Partnership", "High"),
    ("StreetArt Digital", "Agency", "Ankara", "https://www.streetartdigital.com/", "bilgi@streetartdigital.com",
     "+90 533 411 59 54", "", "",
     "400+ projects, 25+ sectors. Ankara client target.",
     "Client", "Medium"),
    ("Mui Medya", "Agency", "Ankara", "https://muimedya.com/", "info@muimedya.com",
     "+90 312 684 80 00", "", "",
     "Ankara dijital reklam ajansı. Web + brand identity. SMB client target.",
     "Client", "Medium"),
    ("SupDijital", "Agency", "Ankara (Çankaya, YDA Center)", "https://supdijital.com/", "info@supdijital.com",
     "+90 544 247 16 07", "", "",
     "70+ brands. Video + web. Çankaya premium location.",
     "Client", "Medium"),
    ("Creamake", "Platform Partner", "Antalya (HQ) + Istanbul + London", "https://creamake.com/", "",
     "+90 553 663 07 43", "https://tr.linkedin.com/company/creamake", "Özgür Yaşacan (Founder)",
     "ikas partner + e-commerce shop. Founder + GM both named — good warm intro.",
     "Partnership", "High"),
    ("Qreate Dijital", "Platform Partner", "Istanbul (Şişli) + Ankara", "https://qreatedijital.com/", "merhaba@qreatedijital.com",
     "+90 850 308 60 02", "https://www.linkedin.com/company/qreate-dijital", "",
     "ikas partner, 360 e-commerce shop. Multi-office. Good partnership angle.",
     "Partnership", "High"),

    # ================== ROUND 3g — More Ankara + Istanbul studios ==================
    ("Blue Ajans", "Agency", "Ankara (Yenimahalle, İvedik)", "https://www.blueajans.com.tr/", "",
     "+90 553 939 25 83", "", "",
     "Ankara industrial-zone SMB agency. Phone-first outreach.",
     "Client", "Medium"),
    ("WMS 360 Digital", "Agency", "Ankara (Yenimahalle, İvedik)", "https://www.wms360.com.tr/", "iletisim@wms360.com.tr",
     "+90 312 911 30 63", "", "Gökhan Bey (lead contact)",
     "10 yrs, Arıkanlar Plaza. Target regional SMB client work.",
     "Client", "Medium"),
    ("FUHA Design Studio", "Agency", "Istanbul", "https://fuha.co/", "",
     "", "https://www.linkedin.com/company/fuha-design-studio/", "",
     "Strategic design studio since 2018, no-code focus. Use LinkedIn or site contact form.",
     "Partnership", "Medium"),
    ("432 Design Studio", "Agency", "Istanbul (Bebek)", "https://www.432designstudio.com/", "",
     "+90 530 063 22 83", "", "",
     "Premium Bebek studio; Coca-Cola/Lenovo/Aygaz portfolio. Pitch Yappaflow as delivery-speed tool.",
     "Client", "Medium"),
    ("Hop Design Studio", "Agency", "Istanbul", "https://www.hopistanbul.com/", "",
     "", "https://www.linkedin.com/company/hopistanbul/", "",
     "Interdisciplinary design studio since 2014. Use LinkedIn or site form.",
     "Client", "Low"),

    # ================== ROUND 4 — LinkedIn-first small agencies ==================
    ("Wedevo", "Agency", "Istanbul", "https://wedevo.net/", "hello@wedevo.net",
     "", "https://www.linkedin.com/company/79526388/", "",
     "Premium web design/dev boutique. hello@ email + LinkedIn.",
     "Client", "Medium"),
    ("Lime or Lemon Digital Studios", "Agency", "Istanbul (Kadıköy)", "https://limeorlemon.com/", "",
     "", "https://www.linkedin.com/company/lime-or-lemon", "",
     "Founded 2014, 11-50 ppl, Muhurdar Cad Kadıköy. Award-winning. LinkedIn-first.",
     "Client", "High"),
    ("EKAS TASARIM", "Agency", "Istanbul (Maltepe) + Orlando", "https://www.ekastasarim.com/", "info@ekastasarim.com",
     "", "https://www.linkedin.com/company/ekas-tasarim/", "",
     "100+ projects; cross-border (US office). Client pitch.",
     "Client", "Medium"),
    ("Epigra", "Agency", "Istanbul (Şişli) + London", "https://epigra.com/", "",
     "", "https://www.linkedin.com/company/epigra", "",
     "Global design+software agency since 2009. Use LinkedIn DM.",
     "Client", "High"),
    ("Babel", "Agency", "Bursa (Osmangazi)", "https://babel.com.tr/", "",
     "+90 224 235 37 37", "https://www.linkedin.com/company/babeltr/", "",
     "17+ yrs, 250+ clients, 360 agency. Regional heavy-hitter.",
     "Client", "Medium"),
    ("Office701", "Agency", "Izmir (Konak)", "https://office701.com/", "destek@office701.com",
     "+90 850 441 47 01", "https://www.linkedin.com/company/office701", "",
     "İzmir-based, 10+ yrs, 587 projects. Solid regional SMB target.",
     "Client", "High"),
    ("Clockwork Agency", "Agency", "Istanbul (Sarıyer Maslak)", "https://clockwork.com.tr/", "clock@clockwork.com.tr",
     "+90 212 275 10 84", "https://www.linkedin.com/company/clockwork-agency", "",
     "Serves Turkcell, Ziraat, BİM. Premium Maslak office. Enterprise clients.",
     "Client", "Medium"),
    ("Brandaft Digital Agency", "Agency", "Istanbul", "https://brandaft.com/", "",
     "+90 552 845 11 80", "https://www.linkedin.com/company/brandaft-digital-agency/", "Şahan Muratoğlu + Benay Özcan (Co-founders)",
     "Two co-founders public on LinkedIn. Warm DM to founders works well.",
     "Client", "High"),
    ("Blakfy", "Agency", "Turkey", "https://blakfy.com/", "",
     "+90 505 979 61 34", "https://www.linkedin.com/company/blakfy", "",
     "500+ companies in 5 years. Web + SEO + ads. Use LinkedIn or WhatsApp.",
     "Client", "Medium"),
    ("Moda Kreatif", "Agency", "Istanbul (Kadıköy)", "https://modakreatif.com/", "hello@modakreatif.com",
     "", "https://www.linkedin.com/company/modakreatif/", "",
     "Kadıköy boutique. Multi-industry portfolio. LinkedIn company + hello@.",
     "Client", "Medium"),
    ("Studio Sour", "Agency", "Istanbul (Şişli)", "https://studiosour.co/", "hello@studiosour.co",
     "", "https://www.linkedin.com/company/studio-sour", "",
     "B2B SaaS growth partner. Yappaflow landing-page speed = direct synergy.",
     "Partnership", "High"),
    ("Gricreative", "Agency", "Istanbul (Kadıköy, Kozyatağı)", "https://gricreative.com/", "info@gricreative.com",
     "+90 216 450 02 30", "https://www.linkedin.com/company/gri-creative/", "",
     "22 yrs, 200+ brands, 1500+ projects. Established mid-market.",
     "Client", "Medium"),
    ("Karen Dijital", "Agency", "Bursa (Osmangazi)", "https://karendijital.com/", "hello@karendijital.com",
     "+90 224 334 11 21", "https://www.linkedin.com/company/karen-dijital/", "",
     "KOBİ→kurumsal positioning. Bursa presence. AI focus.",
     "Client", "Medium"),
    ("Sentez Bilişim", "Agency", "Istanbul (Başakşehir)", "https://sentezbilisim.com/", "destek@sentezbilisim.com",
     "+90 212 470 00 03", "", "",
     "Mall of Istanbul address. Web + e-com + SEO. SMB delivery shop.",
     "Client", "Medium"),
    ("BWA Digital", "Agency", "Istanbul", "https://www.bwa.com.tr/", "",
     "", "https://www.linkedin.com/company/bwa-digital/", "",
     "Corporate web agency. Use LinkedIn DM.",
     "Client", "Low"),
    ("Pigu Creative Agency", "Agency", "Istanbul (Kadıköy)", "https://pigu.com.tr/", "",
     "+90 212 706 74 48", "https://tr.linkedin.com/in/bybayezit", "Yıldırım Bayezit (Founder)",
     "Founded 2017, Kadıköy. Direct founder LinkedIn — best outreach path.",
     "Client", "High"),
    ("kreatiFabrika", "Agency", "Istanbul", "https://www.kreatifabrika.com/", "",
     "", "https://www.linkedin.com/company/kreatifabrika/", "",
     "1,494 LinkedIn followers. Full-stack creative. LinkedIn DM.",
     "Client", "Medium"),
    ("Koala Project", "Agency", "Istanbul (Kadıköy)", "https://koalaproject.com.tr/", "info@koalaproject.com.tr",
     "+90 216 574 51 60", "https://www.linkedin.com/company/11120555/", "",
     "Indep creative ad agency since 2014. Data-driven. Client pitch.",
     "Client", "Medium"),
    ("Kontra Creative", "Agency", "Istanbul", "https://www.kontraist.com/", "",
     "", "https://www.behance.net/kontracr", "",
     "Creative collective. Behance-first. Pitch as delivery accelerator.",
     "Client", "Low"),
    ("BBCo Studio", "Agency", "Istanbul + London + Miami + Berlin", "https://www.bbcostudio.com/", "",
     "", "https://www.linkedin.com/company/bbcostudio/", "",
     "Founded 2022, 1000+ brands. Multi-country strategy+creative+tech.",
     "Client", "High"),
    ("Genau Media", "Agency", "Istanbul (Kadıköy)", "https://genaumedia.com/", "",
     "", "https://www.linkedin.com/company/genau-media/", "",
     "Social media agency. Kadıköy, Zühtüpaşa. LinkedIn-first outreach.",
     "Client", "Medium"),
    ("Hoops", "Agency", "Istanbul (Beşiktaş) + Budapest + Malmö", "https://hoops.com.tr/", "",
     "", "https://www.linkedin.com/company/hoopstheagency", "",
     "360 marcomm; multi-country. Etiler office.",
     "Client", "Medium"),

    # ================== ROUND 4b — More LinkedIn-first mid-small agencies ==================
    ("BrandEnn", "Agency", "Istanbul (Şişli, Esentepe) + NY", "https://www.brandenn.com/", "contact@brandenn.com",
     "+90 212 800 79 26", "https://www.linkedin.com/company/brandenn/", "",
     "12-person team, 20 yrs. partners@ email = partnership DM. Cross-border.",
     "Partnership", "High"),
    ("BrandTeam Group", "Agency", "Istanbul (Kadıköy) + Izmir", "https://www.brandteamgroup.com/", "istanbul@brandteamgroup.com",
     "+90 216 504 20 58", "https://www.linkedin.com/company/brandteamgroup-ist/", "",
     "Folkart Towers. Multi-office. Standard client pitch.",
     "Client", "Medium"),
    ("Starter Design", "Agency", "Istanbul", "https://starterdesign.co/", "info@starterdesign.co",
     "", "https://www.linkedin.com/in/korayshahan/", "Koray Şahan (Founder)",
     "Boutique brand+identity+web shop. Founder on LinkedIn — direct DM.",
     "Client", "High"),
    ("MarkaWorks", "Agency", "Istanbul (Nişantaşı) + London + Dubai + Antalya", "https://markaworks.com/", "contact@markaworks.com",
     "+44 20 3885 8018", "https://www.linkedin.com/company/markaworks/", "",
     "International brand agency, 4 offices. Partnership angle.",
     "Partnership", "Medium"),
    ("Brand Agents", "Agency", "Istanbul (Caddebostan)", "https://www.brand-agents.com/", "info@brand-agents.com",
     "", "https://www.linkedin.com/company/brand-agents/", "Fırat Şaka + Gökhan Balyemez (Co-founders)",
     "200+ international brands. Two founders public — warm DM path.",
     "Client", "High"),
    ("Lein Digital", "Agency", "Istanbul", "https://leindigital.com/", "",
     "+90 530 219 30 72", "https://www.linkedin.com/company/lein-digital/", "Can Bey (senior)",
     "10+ yrs, Turkey's first GEO agency claim. AI-focused. Client angle.",
     "Client", "Medium"),
    ("Brandistanbul PR", "Agency", "Istanbul (Nişantaşı)", "https://brandistanbulpr.com/", "",
     "+90 212 224 02 25", "https://tr.linkedin.com/company/brandistanbul-public-relations/", "Hatice Kumalar (President)",
     "PR boutique with 300+ brands. Pitch Yappaflow as quick campaign landing pages.",
     "Client", "Medium"),
    ("Zeo Agency", "Agency", "Istanbul (Kadıköy) + Ankara + London", "https://www.zeo.org/", "hello@zeo.org",
     "+90 216 336 90 37", "https://www.linkedin.com/company/zeo/", "Ceyhun Burak Akgül (Co-founder/Co-CEO)",
     "50+ consultants, SEO/AI heavy. Co-founder LinkedIn reachable.",
     "Partnership", "High"),
    ("Sempeak", "Agency", "Istanbul (Avcılar + Üsküdar)", "https://www.sempeak.com/", "info@sempeak.com",
     "+90 216 550 01 35", "https://www.linkedin.com/company/sempeak/", "",
     "Since 2011. Digital performance boutique. SMB/mid target.",
     "Client", "Medium"),
    ("Reklam5 Digital Agency", "Agency", "Istanbul (Zeytinburnu) + NY", "https://reklam5.com/", "info@reklam5.com",
     "+90 212 356 21 66", "https://www.linkedin.com/company/reklam5/", "",
     "83 awards; NYC office. Cross-border delivery target.",
     "Client", "Medium"),

    # ================== ROUND 4c — SEO/performance agencies + regional ==================
    ("Mobitek SEO Agency", "Agency", "Istanbul (Kadıköy)", "https://mobitek.com/", "iletisim@mobitek.com",
     "+90 216 418 08 84", "https://www.linkedin.com/company/mobitek/", "",
     "Since 2003, SEO-heavy. Business Istanbul office.",
     "Client", "Medium"),
    ("AORA Digital Agency", "Agency", "Istanbul (Ataşehir)", "https://www.aora.com.tr/", "info@aora.com.tr",
     "+90 216 580 97 20", "https://www.linkedin.com/company/aora-digital-agency/", "",
     "25 yrs; 179 projects. UK phone too. Cross-border.",
     "Client", "Medium"),
    ("ROIBLE", "Agency", "Istanbul (Kadıköy, Bağdat Cad)", "https://roible.com/", "",
     "", "https://www.linkedin.com/company/roible/", "Ibrahim (lead)",
     "Result-oriented turnkey SEO; US/UK/EU markets. LinkedIn-first.",
     "Client", "Medium"),
    ("Adverpeak", "Agency", "Istanbul (Maltepe)", "https://adverpeak.com/", "digital@adverpeak.com",
     "+90 216 606 85 58", "https://www.linkedin.com/company/adverpeak/", "",
     "Since 2010, conversion-focused. Small-mid shop.",
     "Client", "Medium"),
    ("Seobaz", "Agency", "Mersin + Istanbul (Sarıyer)", "https://seobaz.com/", "info@seobaz.com",
     "+90 850 840 95 39", "https://tr.linkedin.com/company/seobaz", "",
     "Mersin HQ = rare regional mid-shop. AI/SEO focus.",
     "Client", "Medium"),
    ("İmza SEO Agency", "Agency", "Konya", "https://www.imza.com/", "",
     "", "https://www.linkedin.com/company/imzainternet/", "",
     "Konya since 2011, 10-49 ppl. Regional SEO shop; Yappaflow pairs with SEO delivery.",
     "Client", "Medium"),

    # ================== ROUND 4d — more ikas partners + freelancers ==================
    ("Artelio Creative", "Platform Partner", "Istanbul", "https://www.arteliocreative.com/", "",
     "+90 850 346 60 71", "https://www.linkedin.com/company/artelio-creative/", "",
     "Official ikas business partner. Training + live support focus. Partnership fit.",
     "Partnership", "High"),
    ("Artnova Creative", "Platform Partner", "Istanbul (Pendik)", "https://www.artnovacreative.com/", "",
     "+90 547 180 18 80", "https://tr.linkedin.com/in/artnovacreativecom", "",
     "ikas partner, API integration expertise. Marketplace specialist.",
     "Partnership", "High"),
    ("CreaTwins", "Platform Partner", "Istanbul (Ümraniye)", "https://creatwins.com/", "",
     "+90 850 303 12 26", "https://www.linkedin.com/company/creatwins/", "",
     "ikas consultant agency. E-com product photography focus. Warm partnership.",
     "Partnership", "Medium"),
    ("Mimoza Bilişim", "Agency", "Zonguldak (Ereğli)", "https://mimozabilisim.com/", "onurkalafat67@gmail.com",
     "+90 530 571 40 67", "https://www.linkedin.com/in/onurkalafat/", "Onur Kalafat (Founder)",
     "ikas partner, AI/GEO focus. Founder direct email. Regional = rare.",
     "Partnership", "High"),
    ("Ajans Max", "Platform Partner", "Turkey", "https://ajansmax.com.tr/", "",
     "", "https://www.linkedin.com/company/ajans-max/", "",
     "ikas partner with e-commerce support. Partnership DM via LinkedIn.",
     "Partnership", "Medium"),
    ("Perest", "Agency", "Istanbul", "https://perest.com/", "",
     "+90 850 242 33 32", "https://www.linkedin.com/company/perest/", "",
     "360 performance agency, Meta/Google certified. Phone + LinkedIn.",
     "Client", "Medium"),
    ("GNS Ajans", "Agency", "Istanbul (Küçükçekmece)", "https://www.gnsajans.com/", "",
     "+90 212 909 04 34", "https://www.linkedin.com/company/GNSAjans", "",
     "Mid-shop, currently rebranding. LinkedIn DM.",
     "Client", "Low"),
    ("Grimor Web Design", "Agency", "Istanbul (Mecidiyeköy)", "https://www.grimor.com/", "info@grimor.com",
     "+90 212 272 46 00", "https://www.linkedin.com/company/grimor-ajans/", "",
     "Since 2004, web + e-com + SEO. 20 yrs, good baseline agency.",
     "Client", "Medium"),
    ("Taksim Ajans", "Agency", "Edirne + Istanbul", "https://taksimajans.com.tr/", "",
     "", "https://www.linkedin.com/company/taksim-ajans/", "",
     "Edirne regional coverage — rare. Small shop.",
     "Client", "Low"),
    ("Ferdi Tarakçı", "Freelancer", "Istanbul (Kadıköy)", "https://ferditarakci.com/", "bilgi@ferditarakci.com",
     "+90 546 831 20 73", "https://www.linkedin.com/in/ferditarakci/", "Ferdi Tarakçı",
     "Full-stack Vue/React/Laravel freelancer since 2008. Perfect white-label partner.",
     "Client", "High"),
    ("Gökhan Çınar", "Freelancer", "Eskişehir", "", "",
     "", "https://www.linkedin.com/in/gokhaanc/", "Gökhan Çınar",
     "Indie maker, engagement+sales design focus. LinkedIn-first outreach.",
     "Client", "Medium"),
    ("Ayşe Akar", "Freelancer", "Ankara / Izmir", "", "",
     "", "https://www.linkedin.com/in/ayse-akar/", "Ayşe Akar",
     "Front-end dev, WordPress + modern FE. LinkedIn DM.",
     "Client", "Medium"),
    ("Yasin Yalçın", "Freelancer", "Istanbul", "", "",
     "", "https://www.linkedin.com/in/yasinyalccin/", "Yasin Yalçın",
     "Startup freelance web dev at Masspace. LinkedIn DM.",
     "Client", "Low"),
    ("Ibrahim Tigrek", "Freelancer", "Ankara / Istanbul", "", "",
     "", "https://www.linkedin.com/in/ibrahim-tigrek/", "Ibrahim Tigrek",
     "Software developer, Ankara/Istanbul. LinkedIn DM.",
     "Client", "Low"),

    # ================== ROUND 4e — LinkedIn-first small agencies + regional + freelancers ==================
    ("Grow Dijital Ajans", "Agency", "Istanbul (Kadıköy)", "https://growdijitalajans.com/", "info@growdijitalajans.com",
     "+90 530 024 38 67", "", "",
     "Boutique digital agency, Kadıköy-based. Social + web. Small team. Email primary.",
     "Client", "Medium"),
    ("AWT Dijital (Erzurum Web)", "Agency", "Erzurum + Ankara + Istanbul + Izmir", "https://www.erzurumweb.com.tr/", "info@erzurumweb.com.tr",
     "+90 444 7298", "", "",
     "Regional shop with multi-city presence. Web + SEO + social. Eastern TR coverage is a differentiator.",
     "Client", "Medium"),
    ("Workuid", "Agency", "Istanbul (Pendik)", "https://workuid.com/", "hello@workuid.com",
     "+90 532 313 59 91", "https://www.linkedin.com/company/workuid/", "",
     "Multi-platform e-commerce setup: Shopify, ikas, Ticimax, Tsoft. LinkedIn-first outreach; partnership pitch fits.",
     "Partnership", "High"),
    ("Zemedya", "Agency", "Istanbul (Ataşehir)", "https://www.zemedya.com/", "info@zemedya.com",
     "+90 216 606 74 69", "https://www.linkedin.com/company/zemedya-internet-hizmetleri/", "",
     "18+ yrs, healthcare/tourism/travel web. LinkedIn company page active.",
     "Client", "Medium"),
    ("WegaBT", "Agency", "Antalya (Kepez) + Trabzon", "https://www.wegabt.com/", "bilgi@wegabt.com",
     "+90 554 898 80 89", "https://www.linkedin.com/company/wegabilgiteknolojileri/", "",
     "15+ yrs. Dual-city (Antalya/Trabzon). LinkedIn active. Web + e-com + SEO.",
     "Client", "Medium"),
    ("SEO Turuncu", "Agency", "Istanbul (Şişli) + Miami", "https://www.seoturuncu.com/", "",
     "+90 537 522 49 06", "https://www.linkedin.com/company/seo-turuncu/", "",
     "E-com consultancy + ikas partner. LinkedIn-first — no public email, DM via LinkedIn.",
     "Partnership", "Medium"),
    ("Arisdot Digital", "Agency", "Eskişehir + Istanbul + Berlin", "https://www.arisdot.com/", "hello@arisdot.com",
     "+90 222 226 000 1", "https://www.linkedin.com/company/arisdotdigital/", "",
     "13 yrs, UX/UI + e-com, multi-country. Small-mid team. hello@ + LinkedIn.",
     "Client", "Medium"),
    ("Adres Ajans", "Agency", "Gaziantep (Şehitkamil)", "https://www.adresajans.com/", "naci@adresajans.com",
     "+90 342 215 16 06", "", "Naci (owner)",
     "Gaziantep boutique. Founder email direct. Regional small-business clients.",
     "Client", "Medium"),
    ("Var Ajans", "Agency", "Gaziantep", "https://www.varajans.com/", "info@varajans.com",
     "+90 544 784 85 25", "", "",
     "Gaziantep SMB agency since 2015. Web/social/video.",
     "Client", "Low"),
    ("Fikir Sanat Medya", "Agency", "Gaziantep (Şehitkamil)", "https://www.fikirsanatmedya.com.tr/", "info@fikirsanatmedya.com.tr",
     "+90 850 466 0 376", "https://linkedin.com/in/fikir-sanat-creative-a58385318", "",
     "Gaziantep creative shop. Packaging + web + digital.",
     "Client", "Low"),
    ("F13 İstanbul", "Agency", "Istanbul (Başakşehir)", "https://f13istanbul.com/", "info@f13istanbul.com",
     "+90 212 843 40 72", "", "",
     "Young (est. 2023) indie agency, creative X tech. Works with premium brands. Good partnership fit.",
     "Partnership", "High"),
    ("MSH Creative", "Agency", "Istanbul (Bahçelievler)", "https://mshcreative.com/", "info@mshcreative.com",
     "+90 533 738 36 74", "", "",
     "10-yr, 27-person shop. 230+ brands, healthcare/real-estate/tourism. Full-service digital.",
     "Client", "Medium"),
    ("EFORWEB", "Agency", "Mersin (Yenişehir)", "https://www.eforweb.com/", "eforweb@eforweb.com",
     "+90 324 325 99 33", "https://www.linkedin.com/company/eforweb-internet-çözümleri/", "",
     "Since 1999. Google Partner. Regional Mersin SMB anchor. Stable shop.",
     "Client", "Medium"),
    ("Üç Yirmiiki (3.22)", "Agency", "Adana (Seyhan)", "https://www.ucyirmiiki.com/", "info@ucyirmiiki.com",
     "+90 322 457 83 22", "", "",
     "Adana boutique — graphic + web + social. Home-office team = small.",
     "Client", "Low"),
    ("Talya Tasarım", "Agency", "Antalya (Muratpaşa)", "https://talyatasarim.com/", "info@talyatasarim.com",
     "+90 543 499 55 55", "", "",
     "Antalya creative digital shop. Web + brand. Dual email for sales/support.",
     "Client", "Medium"),
    ("İMER İletişim", "Agency", "Antalya (Serik)", "https://imeriletisim.com.tr/", "info@imeriletisim.com.tr",
     "+90 545 636 16 48", "https://www.linkedin.com/in/imer-iletisim-ve-reklam-ajansi-38b864297/", "",
     "360° Antalya shop w/ LinkedIn presence. Web + mobile + brand consult.",
     "Client", "Medium"),
    ("Ovi Medya", "Agency", "Antalya + Istanbul (Levent) + Ankara (Çankaya)", "https://www.ovimedya.com/", "ovimedya@ajansovono.com",
     "+90 533 500 93 17", "", "",
     "10+ yr social-heavy shop, 3 cities. Owner under Ovono A.Ş. — responsive email.",
     "Client", "Medium"),
    ("WAX Ajans", "Agency", "Antalya (Muratpaşa)", "https://www.waxajans.com/", "info@waxajans.com",
     "+90 242 312 35 43", "https://www.linkedin.com/company/waxajans/", "",
     "Antalya web + hotel management systems. Hospitality niche.",
     "Client", "Medium"),
    ("Gözde Ajans", "Agency", "Aksaray", "https://www.gozdeajans.com.tr/", "info@gozdeajans.com.tr",
     "+90 382 213 33 45", "", "",
     "Aksaray full-service agency. Regional anchor — less competition.",
     "Client", "Low"),
    ("Stage Dijital", "Agency", "Konya", "https://stagedijital.com/", "merhaba@stagedijital.com",
     "", "https://www.linkedin.com/company/thestagedigital/", "",
     "Konya 15-yr shop. Modern brand, LinkedIn-active, e-com + mobile + web.",
     "Client", "Medium"),
    ("UGR Reklam Ajansı", "Agency", "Konya (Karatay)", "https://ugrajans.com.tr/", "bilgi@ugrajans.com",
     "+90 332 351 51 96", "https://tr.linkedin.com/company/ugrreklamajansi", "",
     "Konya design/digital boutique, LinkedIn-active. Regional.",
     "Client", "Low"),
    ("Bariz Medya", "Agency", "Kocaeli (İzmit)", "https://www.barizmedya.com/", "",
     "+90 507 564 09 32", "https://www.linkedin.com/company/barizmedya/", "",
     "Kocaeli SMB agency. LinkedIn-first — no public email; DM on LinkedIn/WhatsApp.",
     "Client", "Low"),
    ("Meri Creative", "Agency", "Istanbul (Taksim)", "https://mericreative.com/", "info@mericreative.com",
     "+90 212 251 00 18", "https://www.linkedin.com/company/mericreative/", "",
     "Boutique creative agency since 2015. Branding + web. LinkedIn-active.",
     "Client", "Medium"),
    ("Creodive", "Agency", "Istanbul", "https://www.creodive.com.tr/", "hi@creodive.com.tr",
     "+90 544 240 71 99", "https://tr.linkedin.com/company/creodive", "",
     "Award-winning Istanbul digital shop. hi@ email + LinkedIn presence.",
     "Client", "Medium"),
    ("MaunaUP", "Agency", "Samsun (Atakum)", "https://maunaup.com/", "info@maunaup.com",
     "+90 362 438 30 60", "https://www.linkedin.com/company/maunaupcom/", "",
     "Samsun digital media agency, LinkedIn-active. Karadeniz region anchor.",
     "Client", "Medium"),
    ("KNZ Ajans", "Agency", "Denizli", "https://www.knzajans.com.tr/", "info@knzajans.com.tr",
     "+90 507 631 73 82", "https://www.linkedin.com/company/82085328/", "",
     "Denizli boutique. Web + SEO + QR menu. LinkedIn company page.",
     "Client", "Low"),
    ("MAC ART", "Agency", "Denizli (ops Istanbul/Izmir/Muğla)", "https://macart.com.tr/", "bilgi@macart.com.tr",
     "+90 532 491 43 48", "", "",
     "30-yr brand design + digital. Denizli HQ, multi-city ops.",
     "Client", "Medium"),
    ("Wemotion İstanbul", "Agency", "Istanbul", "https://www.wemotionistanbul.com/", "hello@wemotionistanbul.com",
     "+90 507 422 91 65", "https://www.linkedin.com/company/wemotionistanbul/", "",
     "Creative agency — branding + motion + web. 80+ brands. hello@ + LinkedIn.",
     "Client", "Medium"),
    ("Esla Medya", "Agency", "Denizli + Antalya", "https://eslamedya.com/", "info@eslamedya.com",
     "+90 850 812 37 52", "", "",
     "Dual-city Aegean/Mediterranean shop. Web + social. 4 office phones.",
     "Client", "Low"),
    ("bw/a (better with agency)", "Agency", "Samsun", "https://bw.agency/", "",
     "", "", "",
     "Samsun sustainability-focused boutique. Local businesses only. LinkedIn DM via website contact form.",
     "Client", "Low"),
    ("Ekin Pekyiğit", "Freelancer", "Istanbul", "", "",
     "", "https://www.linkedin.com/in/ekinpekyigit/", "Ekin Pekyiğit",
     "Senior UI/UX designer — LinkedIn-first DM. Personal brand strong.",
     "Client", "Medium"),
    ("Can Ünal", "Freelancer", "Istanbul", "", "",
     "", "https://www.linkedin.com/in/canunal/", "Can Ünal",
     "Product designer — LinkedIn-first DM.",
     "Client", "Medium"),
    ("Ekrem Köse", "Freelancer", "Istanbul / Ankara", "", "",
     "", "https://www.linkedin.com/in/ekrem-kose/", "Ekrem Köse",
     "UI/UX freelancer — LinkedIn DM only.",
     "Client", "Low"),
    ("Tuğba Işık", "Freelancer", "Istanbul / Ankara", "", "",
     "", "https://www.linkedin.com/in/tugbaisik/", "Tuğba Işık",
     "UI/UX designer, LinkedIn-active. DM with Yappaflow positioning.",
     "Client", "Low"),
    ("Oğuzhan Öz", "Freelancer", "Istanbul", "https://www.behance.net/oguzhanozuag", "",
     "", "https://www.linkedin.com/in/oguzhan-oz/", "Oğuzhan Öz",
     "Freelance creative developer. Behance + LinkedIn. Strong solo dev + design.",
     "Client", "Low"),

    # ================== ROUND 4f — more Istanbul + Anatolia small agencies + freelancers ==================
    ("SIMPLEIST", "Agency", "Istanbul (Şişli)", "https://www.simple-ist.com/", "request@simple-ist.com",
     "+90 850 473 14 13", "", "",
     "International-focused marketing agency. 6 languages, real-estate/tourism niche. Unique email (request@).",
     "Client", "Medium"),
    ("Macfly Reklam Ajansı", "Agency", "Ankara (Çankaya)", "https://www.macfly.com.tr/", "info@macfly.com.tr",
     "+90 312 911 99 17", "", "",
     "Ankara digital media agency. Brand + social + SEO. Medium-small shop.",
     "Client", "Medium"),
    ("Ajans Bulut", "Agency", "Bursa", "https://www.ajansbulut.com/", "bilgi@ajansbulut.com",
     "+90 224 888 00 95", "https://www.linkedin.com/company/ajans-bulut/", "",
     "Bursa's self-styled leading digital shop since 2015. LinkedIn company presence.",
     "Client", "Medium"),
    ("Taha Erel", "Freelancer", "Istanbul", "https://tahaerel.com/", "",
     "", "https://www.linkedin.com/in/tahaerel/", "Taha Erel",
     "Freelance WordPress/full-stack web dev. Personal portfolio + LinkedIn (500+ conn).",
     "Client", "Medium"),
    ("Ahtapot Sosyal Medya", "Agency", "Istanbul (Şişli)", "https://www.ahtapotsosyalmedya.com/", "info@ahtapotsm.com",
     "+90 212 212 32 33", "https://www.linkedin.com/company/ahtapot-sosyal-medya/", "",
     "Mid-boutique (50 people), creative/social-heavy. 2014. Good LinkedIn presence.",
     "Client", "Medium"),
    ("Proji Digital", "Agency", "Izmir (Bayraklı) + Istanbul (Sarıyer)", "https://proji.com.tr/", "info@proji.com.tr",
     "+90 232 433 74 16", "https://www.linkedin.com/company/projicomtr/", "",
     "Dual-city (Izmir/Istanbul) digital agency since 2015. LinkedIn-active.",
     "Client", "Medium"),
    ("Zolpix", "Agency", "Istanbul (Kağıthane)", "https://kagithane.zolpix.com/", "info@zolpix.com",
     "+90 505 219 18 52", "", "",
     "District-focused Istanbul shop (39 districts). Transparent pricing (18K-85K TL).",
     "Client", "Low"),
    ("Rekclick", "Agency", "Istanbul (4.Levent)", "https://www.rekclick.com/", "merhaba@rekclick.com",
     "+90 212 269 95 96", "https://www.linkedin.com/company/rekclick/", "",
     "15-yr Levent-based 360° digital agency. merhaba@ — warm tone; Turkish-market-fit.",
     "Client", "Medium"),
    ("Ozge Keles", "Freelancer", "Ankara", "https://www.ozgekeles.com/", "",
     "", "https://www.linkedin.com/in/ozge-keles-webflow/", "Ozge Keles",
     "Certified Webflow Partner (since 2024). Solo freelancer, $5K+ projects. Strong Yappaflow fit (static-site alignment).",
     "Partnership", "High"),
    ("Berkay Çınar", "Freelancer", "Istanbul", "https://www.berkaycinar.com/", "",
     "", "https://linkedin.com/in/berkaycinar", "Berkay Çınar",
     "UI/UX/interaction designer w/ personal site. Behance + LinkedIn. LinkedIn-first DM.",
     "Client", "Medium"),
    ("Tolga Taşcı", "Freelancer", "Istanbul", "https://www.tolgatasci.co.uk/", "",
     "", "https://www.behance.net/tolgatasci", "Tolga Taşcı",
     "Senior UX/UI designer with international focus. Behance + personal .co.uk site.",
     "Client", "Low"),
    ("Gelişim Medya", "Agency", "Kayseri (Melikgazi)", "https://www.gelisimmedya.com/", "info@gelisimmedya.com",
     "", "", "",
     "Multi-city regional shop (Kayseri/Malatya/Elazığ). Eastern Anatolia coverage niche.",
     "Client", "Low"),
    ("OXO Creative", "Agency", "Kayseri (Kocasinan)", "https://oxocreative.com/", "info@oxocreative.com",
     "", "", "",
     "Kayseri boutique. Google Ads + web + brand. Small team. Regional.",
     "Client", "Low"),
]

# Header row
leads.append(HEADERS)
for col_idx in range(1, len(HEADERS) + 1):
    c = leads.cell(row=1, column=col_idx)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = CENTER
    c.border = BORDER
leads.row_dimensions[1].height = 32

# Data rows
for idx, row in enumerate(LEADS, start=1):
    company, kind, city, website, email, phone, linkedin, dm, fit, angle, priority = row
    r = idx + 1
    values = [idx, company, kind, city, website, email, phone, linkedin, dm, fit, angle, priority, ""]
    for col_idx, v in enumerate(values, start=1):
        c = leads.cell(row=r, column=col_idx, value=v)
        c.font = BODY_FONT
        c.alignment = WRAP
        c.border = BORDER
    # zebra on even rows
    if idx % 2 == 0:
        for col_idx in range(1, len(HEADERS) + 1):
            leads.cell(row=r, column=col_idx).fill = ZEBRA
    # priority color override on the priority cell
    pri_cell = leads.cell(row=r, column=12)
    if priority == "High":
        pri_cell.fill = PRIORITY_HIGH
        pri_cell.font = BODY_BOLD
        pri_cell.alignment = CENTER
    else:
        pri_cell.alignment = CENTER
    leads.cell(row=r, column=1).alignment = CENTER  # row number
    leads.cell(row=r, column=3).alignment = CENTER
    leads.row_dimensions[r].height = 54

# Column widths
widths = {
    "A": 5, "B": 32, "C": 18, "D": 26, "E": 38, "F": 32, "G": 20,
    "H": 44, "I": 24, "J": 55, "K": 14, "L": 12, "M": 14,
}
for col, w in widths.items():
    leads.column_dimensions[col].width = w

# Freeze header
leads.freeze_panes = "A2"

# Convert data range to a real Excel table for filter/sort UI
last_col = get_column_letter(len(HEADERS))
last_row = 1 + len(LEADS)
tbl_ref = f"A1:{last_col}{last_row}"
table = Table(displayName="Leads", ref=tbl_ref)
table.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
    showRowStripes=False, showColumnStripes=False,
)
# openpyxl requires unique style name, and the header style takes over — remove fills on headers so table style shows
for col_idx in range(1, len(HEADERS) + 1):
    h = leads.cell(row=1, column=col_idx)
    h.fill = PatternFill(fill_type=None)
    h.font = Font(name=FONT, bold=True, color="FFFFFF", size=11)
leads.add_table(table)

# Summary at the bottom using formulas (total, high priority count, platform partners)
summary_row = last_row + 3
leads.cell(row=summary_row, column=2, value="Total leads").font = BODY_BOLD
leads.cell(row=summary_row, column=3, value=f"=COUNTA(B2:B{last_row})").font = BODY_FONT
leads.cell(row=summary_row + 1, column=2, value="High-priority").font = BODY_BOLD
leads.cell(row=summary_row + 1, column=3, value=f'=COUNTIF(L2:L{last_row},"High")').font = BODY_FONT
leads.cell(row=summary_row + 2, column=2, value="Partnership-angle leads").font = BODY_BOLD
leads.cell(row=summary_row + 2, column=3, value=f'=COUNTIF(K2:K{last_row},"Partnership")').font = BODY_FONT
leads.cell(row=summary_row + 3, column=2, value="Agencies").font = BODY_BOLD
leads.cell(row=summary_row + 3, column=3, value=f'=COUNTIF(C2:C{last_row},"Agency")').font = BODY_FONT
leads.cell(row=summary_row + 4, column=2, value="Freelancers").font = BODY_BOLD
leads.cell(row=summary_row + 4, column=3, value=f'=COUNTIF(C2:C{last_row},"Freelancer")').font = BODY_FONT
leads.cell(row=summary_row + 5, column=2, value="Platform partners").font = BODY_BOLD
leads.cell(row=summary_row + 5, column=3, value=f'=COUNTIF(C2:C{last_row},"Platform Partner")').font = BODY_FONT

# =========================================================
# Sheet 3: Samsun (in-person visit list)
# =========================================================
# Columns tuned for walking into offices: Full Address + Phone are primary,
# email / LinkedIn are secondary because the user will be showing up in person.
samsun = wb.create_sheet("Samsun")
samsun.sheet_view.showGridLines = False

# Title band
samsun["A1"] = "Samsun — in-person visit list (küçük/orta web ajansları)"
samsun["A1"].font = TITLE_FONT
samsun.merge_cells("A1:M1")
samsun["A2"] = (
    "Address-first list of small/medium Samsun web/digital agencies. Full addresses + phones are "
    "primary because Yusuf Mirza plans to walk into offices; email / LinkedIn / Instagram are fallback."
)
samsun["A2"].font = Font(name=FONT, italic=True, size=10, color="666666")
samsun["A2"].alignment = WRAP
samsun.merge_cells("A2:M2")
samsun.row_dimensions[2].height = 30

SAMSUN_HEADERS = [
    "#", "Agency", "Type", "District (İlçe)", "Neighborhood (Mahalle)",
    "Full Address", "Phone", "Email", "Website", "Socials",
    "Visit Priority", "Notes", "Visit Status",
]

# (agency, type, district, neighborhood, full_address, phone, email, website, socials, priority, notes)
SAMSUN_LEADS = [
    ("MaunaUP", "Agency", "Atakum", "Cumhuriyet Mah.",
     "Cumhuriyet Mah. Atatürk Bulv. No:327, Kat:3 Daire:13, Yavuz Grup İş Merkezi, Atakum/Samsun",
     "+90 362 438 30 60", "info@maunaup.com", "https://maunaup.com/",
     "LinkedIn: maunaupcom · IG: @maunaupcom",
     "High",
     "Full digital-media agency, LinkedIn-active. Walking in as a fellow Samsun developer is a strong opener — ask for the founder."),
    ("Dijital10 Marka ve Strateji", "Agency", "Atakum", "Cumhuriyet Mah.",
     "Cumhuriyet Mah. Doğuş Cad. 38. Sk. No:2/8, Yavuz Tuna Center, Atakum/Samsun",
     "+90 539 771 34 54", "destek@dijital10.com", "https://www.dijital10.com/",
     "IG: @dijital10 · X: @Dijital10com",
     "High",
     "Brand + strategy agency inside Yavuz Tuna Center. Same building as Mustafa Sarıca (freelance) — you can double-hit in one trip."),
    ("Mustafa Sarıca (freelance)", "Freelancer", "Atakum", "Cumhuriyet Mah.",
     "Yavuz Tuna Center, Doğuş Cd. 38. Sk. No:2/8, Atakum/Samsun",
     "+90 543 535 15 13", "merhaba@mustafasarica.com", "https://mustafasarica.com/",
     "IG: @mustafasarica_com · LinkedIn: mustafa-sarica",
     "High",
     "Solo dev, SAME BUILDING as Dijital10. Freelancers are Yappaflow's easiest yes — one-person volume uplift is the pitch."),
    ("Deca Yazılım", "Agency", "Atakum", "Atakent Mah.",
     "Atakent Mah. 3307. Sk. No:4, 55200 Atakum/Samsun",
     "+90 545 270 13 46", "info@decayazilim.com", "https://decayazilim.com/",
     "Twitter: @vamtam",
     "High",
     "Yazılım-first shop (software house). Exact 'custom delivery' profile Yappaflow compresses from days → minutes."),
    ("Proweb Tasarım Samsun", "Agency", "İlkadım", "Kale Mah.",
     "Kale Mah. Kazımpaşa Cad. Meseret Sok. No:1 Kat:2/23, 55260 İlkadım/Samsun",
     "+90 545 368 53 69", "destek@prowebtasarim.net", "https://www.prowebtasarim.net/",
     "FB: prowebtasarim · IG: @prowebajans",
     "High",
     "Self-described #1 Samsun web-design firm. Central İlkadım address, easy walk-in from city center."),
    ("CodiaSoft", "Agency", "İlkadım", "Çiftlik Mah.",
     "Çiftlik Mah. İstiklal Cad. Başoğlu İş Merkezi Varol Apt. No:102/A Kat:3 Daire:2, İlkadım/Samsun",
     "+90 850 532 01 08", "info@codiasoft.com", "https://www.codiasoft.com/",
     "IG: @codiasoft · LinkedIn: codiasoft",
     "High",
     "Software + web shop, city-center (Çiftlik). LinkedIn-active — Yappaflow's AI-driven generator angle will resonate."),
    ("Webmooy Dijital Reklam", "Agency", "Atakum", "Yenimahalle",
     "Yenimahalle, İzmir Cd. Bina No:25 B.B.15, Atakum/Samsun",
     "+90 850 480 05 15", "info@webmooy.com", "https://webmooy.com/",
     "IG: @webmooy · LinkedIn: webmooy",
     "High",
     "Since 2020 — web + SEO + full-service digital marketing. Atakum address, easy visit."),
    ("İlgiHost / Samweb", "Agency", "İlkadım", "Kale Mah.",
     "Kale Mah. Gazi Cad. Bafra İşhanı No:52 Kat:2 Daire:6, İlkadım/Samsun",
     "+90 362 435 03 30", "bilgi@ilgihost.com", "https://samweb.com.tr/",
     "IG: @ilgihost",
     "Medium",
     "Hosting + web-design combo in city center (Bafra İşhanı). Walk-in friendly. Mention Yappaflow ZIP hand-off as a hosting upsell."),
    ("Fikir Kulübü", "Agency", "İlkadım", "Kılıçdede Mah.",
     "Kılıçdede Mah. İstiklal Cd. No:181/10, İlkadım/Samsun",
     "+90 362 233 28 49", "bilgi@fikirkulubu.com.tr", "https://www.fikirkulubu.com.tr/",
     "IG: @fikir.kulubu · FB: fikirkulup",
     "Medium",
     "Marka + iletişim + baskı + etkinlik ajansı. Partnership angle > client — they sub-contract website work."),
    ("Czia Kurumsal", "Agency", "Atakum", "Esenler Mah.",
     "Esenler Mah. 317. Sk. No:2 Kat:3, Atakum/Samsun",
     "+90 362 230 88 88", "info@czia.com", "https://czia.com/",
     "FB: CziaCom · X: @cziacom",
     "Medium",
     "Kurumsal hizmet evi (web + grafik + foto/video). Atakum office, easy walk-in from most of Samsun."),
    ("Sitebizden (Samsun Web Tasarımı)", "Agency", "İlkadım", "Kılıçdede Mah.",
     "Kılıçdede Mah. Ülkem Sk. No:8 Borkonut Niş İş Merkezi B Blok D:36, İlkadım/Samsun (2. ofis: OMÜ Teknopark, Atakum)",
     "+90 850 302 43 55", "", "https://www.samsunwebtasarimi.com/",
     "IG: @sitebizden · FB: sitebizden",
     "Medium",
     "14 yıllık firma, 600+ aktif müşteri. İki ofis (şehir + Teknopark). Kılıçdede ofisini seç — daha kalabalık."),
    ("Redia Dijital (redia.com.tr)", "Agency", "Atakum", "Yenimahalle",
     "Atakum Vatan Cd. 3169. Sk., Yenimahalle tramvay durağı karşısı, Atakum/Samsun",
     "+90 555 888 55 24", "info@redia.com.tr", "https://redia.com.tr/",
     "IG: @redia.ajans · LinkedIn: Redia Ajans",
     "Medium",
     "Atakum ofisi, tram stop landmark'ı kolay. 'Yenimahalle tramvay durağı karşısı' ile kolayca bulunur."),
    ("Redia Ajans (medya.red)", "Agency", "İlkadım", "Mevlana Mah.",
     "Mevlana Mah. Şehit Gökhan Çavuşoğlu Sk., İlkadım/Samsun",
     "+90 555 888 55 24", "bilgi@medya.red", "https://www.medya.red/",
     "IG: @medyared.samsun · LinkedIn: redia-ajans",
     "Medium",
     "Redia'nın İlkadım ayağı. 2006'dan beri 200+ website müşterisi. Partnership angle — Yappaflow = delivery accelerator."),
    ("Seven's Design", "Agency", "İlkadım", "Derebahçe Mah.",
     "Derebahçe Mah., İlkadım/Samsun (tam adres telefondan)",
     "+90 553 719 53 84", "info@sevens-design.com", "https://sevens-design.com/",
     "IG: @sevens7design",
     "Medium",
     "Web + grafik tasarım. Sokak/no web'de yok — önce arayıp tam adres al."),
    ("Annak Studio", "Agency", "Atakum", "Samsun (şehir içi)",
     "Samsun ofisi (spesifik adres gizli — Atakum semti; 2. ofis: Washington, USA)",
     "", "hi@studioannak.com", "https://studioannak.com/",
     "",
     "Medium",
     "UI/UX-forward stüdyo, US ofisi de var. Email-first: randevu iste, sonra Atakum ofisine git."),
    ("bw/a (Better With Agency)", "Agency", "Samsun", "—",
     "Samsun (contact form only — adres web'de yok)",
     "", "", "https://bw.agency/",
     "",
     "Medium",
     "Zaten ana Leads listesinde var. Modern Samsun ajansı — form üzerinden walk-in talebi gönder."),
    ("Hepar Web Tasarım", "Agency", "Samsun genel", "—",
     "Samsun merkezli (Atakum/İlkadım/Canik/Terme/Havza hizmet bölgesi). Adres yok — telefon/WhatsApp.",
     "+90 545 500 01 51", "", "https://hepar.org.tr/",
     "",
     "Low",
     "15-yıllık. Spesifik ofis yok gibi — önce WhatsApp'tan randevu iste."),
    ("İnnova Reklam Web Tasarım", "Agency", "Samsun", "—",
     "Samsun (tam adres web'de yok). Ticaret sicil no: 14040.",
     "+90 532 683 13 43", "", "https://www.innovaajans.com.tr/",
     "",
     "Low",
     "12+ yıl. Sadece telefon + sicil no. Önce ara, adresi tam al, sonra git."),
    ("Dijitasyon", "Agency", "Samsun", "—",
     "Samsun (spesifik sokak/kat web sitesinde yok)",
     "+90 530 642 69 55", "web@dijitasyon.com.tr", "http://dijitasyon.com.tr/",
     "IG: @dijitasyoncomtr",
     "Low",
     "Web + dijital pazarlama. Adres gizli — telefon/email ilk temas, sonra ziyaret."),
    ("Onsa Ajans", "Agency", "Samsun", "—",
     "Samsun (4-kişilik ekip, adres gizli)",
     "", "", "https://www.onsaajans.com/",
     "IG: @onsaajans",
     "Low",
     "2017'den beri — önce İstanbul, şimdi Samsun ekibi. Instagram DM veya site formu ilk temas."),

    # =========================================================
    # Round 2: extended Samsun research — more agencies + freelancers
    # =========================================================
    ("Webbeyaz", "Agency", "İlkadım", "Kılıçdede Mah.",
     "Kılıçdede Mah. Ülkem Sk. No:8 Borkonut Niş Plaza B Blok Kat:4 D:37, İlkadım/Samsun",
     "+90 545 832 39 29", "info@webbeyaz.com", "https://www.webbeyaz.com/",
     "IG: @webbeyaz · LinkedIn: webbeyaz",
     "High",
     "10+ yıllık. AYNI BİNA Sitebizden ile (Borkonut Niş Plaza B Blok — Webbeyaz D:37, Sitebizden D:36). Tek ziyaret iki pitch!"),
    ("Tasarım Dünyası", "Agency", "İlkadım", "Zafer Mah.",
     "Zafer Mah. Divitçioğlu Cad. No:34, İlkadım/Samsun",
     "+90 362 230 76 91", "info@tasarimdunyasi.com", "https://tasarimdunyasi.com/",
     "Twitter / Facebook / Vimeo",
     "Medium",
     "20+ yıllık hosting + tasarım firması. Zafer Mah şehir merkezi — kolay walk-in."),
    ("362 Tasarım ve Yazılım Ajansı", "Agency", "İlkadım", "Zafer Mah.",
     "Zafer Mah. Cumhuriyet Cd. Devecioğlu İş Merkezi Kat:6 No:23/102, İlkadım/Samsun",
     "+90 362 230 88 88", "info@362.com.tr", "https://www.362.com.tr/",
     "IG/FB/X: @362digital",
     "High",
     "14 yıllık — grafik + mobil uygulama + web yazılım. Belediye + kurumsal portföy. Zafer Mah merkez."),
    ("Adsmethod Dijital Reklam", "Agency", "İlkadım", "Adalet Mah.",
     "Adalet Mah. Şehit Polis Memuru Murat Çalışkan Sk. No:1/3, İlkadım/Samsun",
     "+90 536 577 91 01", "info@adsmethod.com", "https://adsmethod.com/",
     "",
     "Medium",
     "2024 kuruluşu — genç/büyüme odaklı. Web + grafik tasarım. Yeni ajanslar yeni araçlara daha açık."),
    ("Kayrasoft Yazılım", "Agency", "İlkadım", "Zafer Mah.",
     "Zafer Mah. Gazi Cad. Gazi Apt. Kat:1 No:150/1, İlkadım/Samsun",
     "+90 362 233 80 33", "info@kayrasoft.com.tr", "https://kayrasoft.com.tr/",
     "",
     "High",
     "2010'dan beri (kurucu: Ahmet Koç). Kurumsal yazılım + web. Zafer Mah merkez — CodiaSoft ve Tasarım Dünyası ile aynı bölge, tek güzergâh üç pitch."),
    ("Web Çizgisi", "Agency", "Canik", "Karşıyaka Mah.",
     "Karşıyaka Mah. Akıncılar Sk. No:24, Canik/Samsun",
     "+90 544 571 15 47", "", "",
     "Yandex Maps doğrulamalı",
     "Medium",
     "Canik yerel web tasarım dükkanı. Küçük + kişisel — Yappaflow'un 'bir kişi stüdyo hacmi' pitch'i tam onlara göre."),
    ("Vayes Digital", "Agency", "Atakum", "OMÜ Teknopark",
     "OMÜ Cad. No:165, OMÜ Teknopark Geliştirme Bölgesi, Atakum/Samsun",
     "", "", "https://www.vayes.com.tr/",
     "",
     "Medium",
     "Teknopark-resident yazılım + tasarım firması. Enterprise/ihracat odaklı — partnership angle (Yappaflow teslimat hızlandırıcı)."),
    ("Tekşen Bilişim", "Agency", "Atakum", "OMÜ Teknopark",
     "OMÜ Teknopark, Atakum/Samsun",
     "", "", "https://teksenbilisim.com/",
     "",
     "Low",
     "Teknopark bilişim firması — yazılım ihracatı odaklı. Sitebizden Teknopark ofisi de aynı çevre."),
    ("KODAKIL Web & Yazılım", "Agency", "Atakum", "Körfez Mah.",
     "Körfez Mah., Atakum/Samsun (sokak/kat gizli — telefon/email ile doğrulama)",
     "", "", "",
     "Bumaps Samsun listelemesi",
     "Low",
     "Bumaps'te listelenmiş küçük web/yazılım şirketi. Önce telefon al, sonra adres kesinleştir."),
    ("Dokuz Adworks", "Agency", "Samsun", "—",
     "Samsun (site iletişim formu ile — fiziksel adres web'de yok)",
     "", "", "https://dokuzadworks.com/",
     "",
     "Low",
     "Marka + kurumsal kimlik + reklam ajansı; full-service, web dahil. Form üzerinden randevu."),
    ("Must Creative", "Agency", "Canik/Bafra", "—",
     "Samsun (Canik/Bafra hizmet bölgesi — ofis adresi belirsiz)",
     "+90 541 896 70 89", "", "https://mustcreative.com/",
     "",
     "Low",
     "Canik + Bafra odaklı — telefon-only iletişim. Bafra'ya gitmek için Samsun dışı yolculuk (~70 km)."),

    # --- Freelancers (solo devs, LinkedIn/Behance/Dribbble-first) ---
    ("Can Turan (canturan.dev)", "Freelancer", "Samsun", "—",
     "Freelance — fiziksel ofis yok, görüşme talep üzerine",
     "", "", "https://www.canturan.dev/",
     "",
     "High",
     "15+ yıllık web tasarımcı, 500+ proje. SEO-uyumlu modern siteler — Yappaflow'un 'solo → stüdyo hacmi' pitch'i birebir fit."),
    ("Furkan Sağlam (Mono Interactive)", "Freelancer", "Atakum", "—",
     "Atakum/Samsun — freelance + Mono Interactive kurucusu (hibrit)",
     "", "", "https://www.furkansaglam.com/",
     "Upwork + YouTube + furkansaglam.com",
     "High",
     "Front-end + UI/UX + SEO uzmanı. Kendi ajansını kurmuş — partnership angle (Yappaflow = teslimat hızlandırıcı)."),
    ("Ayhan Karaman", "Freelancer", "Samsun", "—",
     "Samsun — freelance/remote (görüşme üzerine kafe buluşması)",
     "", "", "https://www.behance.net/yhnkaraman",
     "LinkedIn: ayhankaraman · Behance: yhnkaraman",
     "High",
     "10+ yıl IT + web + SEO + Digital Marketing Manager geçmişi. LinkedIn DM önce, sonra Samsun kafe buluşması."),
    ("Yasin Kökdemir", "Freelancer", "Atakum", "—",
     "Atakum/Samsun (Atakum Belediyesi bünyesinde frontend dev — yan projede freelance)",
     "", "", "",
     "LinkedIn",
     "Medium",
     "Frontend dev — Wegobi Studio, Workintech, ATF Studios geçmişi. LinkedIn-first; mesai dışı freelance ilgi alanı."),
    ("Alican Şahvelioğlu", "Freelancer", "Atakum", "Rasathane",
     "Atakum/Rasathane (Atakum Belediyesi bünyesinde — Behance üzerinden iletişim)",
     "", "", "https://www.behance.net/shvloglu",
     "Behance: shvloglu · LinkedIn",
     "Medium",
     "UI/UX + front-end + grafik tasarım. Belediyede tam zamanlı ama freelance work gösteriyor."),
    ("Mehmet Online", "Freelancer", "Samsun", "—",
     "Samsun — freelance (fiziksel adres yok, Dribbble DM)",
     "", "", "https://dribbble.com/mehmetonline",
     "Dribbble: mehmetonline",
     "Low",
     "Web + mobile design. Dribbble aktif — genç portföy."),
    ("Metehan Bahadır", "Freelancer", "Samsun", "—",
     "Samsun — freelance (Behance/LinkedIn DM)",
     "", "", "https://www.behance.net/metehanbahadir",
     "Behance + LinkedIn",
     "Low",
     "20 yıl brand/logo/visual design. Saf web odaklı değil — tasarım ortağı olarak cross-sell."),
    ("Tuğçe Deniz", "Freelancer", "Samsun", "—",
     "Samsun — freelance (OMÜ 2024 mezun)",
     "", "", "https://www.behance.net/ttugcednz",
     "Behance · IG: @tugcedenizz · LinkedIn",
     "Low",
     "Yeni mezun — early-career grafik tasarımcı. Uzun vadeli early-adopter olabilir."),
]

# Header row for Samsun (at row 4 to leave space for title band)
samsun_hdr_row = 4
for col_idx, h in enumerate(SAMSUN_HEADERS, start=1):
    c = samsun.cell(row=samsun_hdr_row, column=col_idx, value=h)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = CENTER
    c.border = BORDER
samsun.row_dimensions[samsun_hdr_row].height = 32

# Data rows
for idx, row in enumerate(SAMSUN_LEADS, start=1):
    agency, kind, district, neigh, full_addr, phone, email, website, socials, priority, notes = row
    r = samsun_hdr_row + idx
    values = [idx, agency, kind, district, neigh, full_addr, phone, email, website, socials, priority, notes, ""]
    for col_idx, v in enumerate(values, start=1):
        c = samsun.cell(row=r, column=col_idx, value=v)
        c.font = BODY_FONT
        c.alignment = WRAP
        c.border = BORDER
    # zebra on even rows
    if idx % 2 == 0:
        for col_idx in range(1, len(SAMSUN_HEADERS) + 1):
            samsun.cell(row=r, column=col_idx).fill = ZEBRA
    # priority color override
    pri_cell = samsun.cell(row=r, column=11)
    if priority == "High":
        pri_cell.fill = PRIORITY_HIGH
        pri_cell.font = BODY_BOLD
    pri_cell.alignment = CENTER
    # center the row #, type, district
    samsun.cell(row=r, column=1).alignment = CENTER
    samsun.cell(row=r, column=3).alignment = CENTER
    samsun.cell(row=r, column=4).alignment = CENTER
    samsun.row_dimensions[r].height = 72  # taller — full addresses wrap

# Column widths tuned for walk-in use (Full Address is the widest)
samsun_widths = {
    "A": 4,   # #
    "B": 30,  # Agency
    "C": 13,  # Type
    "D": 15,  # District
    "E": 20,  # Neighborhood
    "F": 55,  # Full Address ← primary
    "G": 20,  # Phone ← primary
    "H": 28,  # Email
    "I": 32,  # Website
    "J": 30,  # Socials
    "K": 13,  # Visit Priority
    "L": 50,  # Notes
    "M": 14,  # Visit Status
}
for col, w in samsun_widths.items():
    samsun.column_dimensions[col].width = w

# Freeze header + the agency name column so the address stays visible as you scroll right
samsun.freeze_panes = "C5"

# Convert to Excel Table for filter/sort UI
samsun_last_col = get_column_letter(len(SAMSUN_HEADERS))
samsun_last_row = samsun_hdr_row + len(SAMSUN_LEADS)
samsun_tbl_ref = f"A{samsun_hdr_row}:{samsun_last_col}{samsun_last_row}"
samsun_table = Table(displayName="SamsunLeads", ref=samsun_tbl_ref)
samsun_table.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium4", showFirstColumn=False, showLastColumn=False,
    showRowStripes=False, showColumnStripes=False,
)
# strip header fills so the table style shows
for col_idx in range(1, len(SAMSUN_HEADERS) + 1):
    h = samsun.cell(row=samsun_hdr_row, column=col_idx)
    h.fill = PatternFill(fill_type=None)
    h.font = Font(name=FONT, bold=True, color="FFFFFF", size=11)
samsun.add_table(samsun_table)

# Summary below the table
samsun_summary_row = samsun_last_row + 2
samsun.cell(row=samsun_summary_row, column=2, value="Samsun leads total").font = BODY_BOLD
samsun.cell(row=samsun_summary_row, column=3,
            value=f"=COUNTA(B{samsun_hdr_row + 1}:B{samsun_last_row})").font = BODY_FONT
samsun.cell(row=samsun_summary_row + 1, column=2, value="High visit-priority").font = BODY_BOLD
samsun.cell(row=samsun_summary_row + 1, column=3,
            value=f'=COUNTIF(K{samsun_hdr_row + 1}:K{samsun_last_row},"High")').font = BODY_FONT
samsun.cell(row=samsun_summary_row + 2, column=2, value="Atakum-based").font = BODY_BOLD
samsun.cell(row=samsun_summary_row + 2, column=3,
            value=f'=COUNTIF(D{samsun_hdr_row + 1}:D{samsun_last_row},"Atakum")').font = BODY_FONT
samsun.cell(row=samsun_summary_row + 3, column=2, value="İlkadım-based").font = BODY_BOLD
samsun.cell(row=samsun_summary_row + 3, column=3,
            value=f'=COUNTIF(D{samsun_hdr_row + 1}:D{samsun_last_row},"İlkadım")').font = BODY_FONT

# =========================================================
# Sheet 4: İzmir (in-person visit list — second travel city)
# =========================================================
# Same 13-column in-person layout as Samsun. İzmir is ~700 km south-west;
# trip planning matters — Folkart Towers (Bayraklı) and Alsancak have
# the densest agency clusters, so Visit Priority + District are the
# primary sort axes.
izmir = wb.create_sheet("İzmir")
izmir.sheet_view.showGridLines = False

izmir["A1"] = "İzmir — in-person visit list (küçük/orta web ajansları + freelancer'lar)"
izmir["A1"].font = TITLE_FONT
izmir.merge_cells("A1:M1")
izmir["A2"] = (
    "Address-first list of İzmir web/digital agencies + freelancers. Cluster hot-spots: Folkart Towers "
    "(Bayraklı/Adalet Mah.), Mansuroğlu (Bayraklı), Alsancak/Konak, Karşıyaka. Before travelling, call the "
    "'telefondan doğrulanmalı' entries to confirm kat/daire. Priority-High = clear SMB fit + verified address."
)
izmir["A2"].font = Font(name=FONT, italic=True, size=10, color="666666")
izmir["A2"].alignment = WRAP
izmir.merge_cells("A2:M2")
izmir.row_dimensions[2].height = 44

IZMIR_HEADERS = [
    "#", "Agency", "Type", "District (İlçe)", "Neighborhood (Mahalle)",
    "Full Address", "Phone", "Email", "Website", "Socials",
    "Visit Priority", "Notes", "Visit Status",
]

# (agency, type, district, neighborhood, full_address, phone, email, website, socials, priority, notes)
IZMIR_LEADS = [
    # =====================================================
    # Folkart Towers cluster (Bayraklı / Adalet Mah.) — 6 in one building
    # =====================================================
    ("Mudiweb", "Agency", "Bayraklı", "Adalet Mah.",
     "Adalet Mh. Manas Blv. No:48 Folkart Towers, 35530 Bayraklı/İzmir",
     "+90 532 770 46 23", "info@mudiweb.com", "https://www.mudiweb.com/",
     "LinkedIn: mudiweb",
     "High",
     "Folkart Towers CLUSTER — 6 ajans aynı binada (Mudiweb, OVARSA, Best4SEO, WebUsta, Çınar, Roof). Bayraklı metro çıkışı; tek ziyaret çoklu pitch."),
    ("OVARSA Yazılım", "Agency", "Bayraklı", "Adalet Mah.",
     "Adalet Mh. Manas Blv. No:39 Folkart Towers B Kule K:32 D:3201, Bayraklı/İzmir",
     "+90 232 332 37 76", "info@ovarsa.com", "https://www.ovarsa.com/",
     "LinkedIn: ovarsa-yazılım-teknolojileri",
     "High",
     "Aynı bina Mudiweb/Best4SEO/WebUsta ile — Folkart B Kule 32. kat. İzmir + Manisa ofis. LinkedIn DM sonra walk-in."),
    ("Best4SEO", "Agency", "Bayraklı", "Adalet Mah.",
     "Adalet Mh. Manas Blv. Folkart Towers B Blok D:3010, 35530 Bayraklı/İzmir",
     "+90 530 776 74 30", "info@best4seo.com", "https://best4seo.com.tr/",
     "LinkedIn: best4seo",
     "High",
     "Çok dilli SEO + web tasarım. Folkart B Blok. Cluster içinde 3. durak."),
    ("WebUsta", "Agency", "Bayraklı", "Adalet Mah.",
     "Adalet Mh. Manas Blv. No:47 Folkart Towers D:2896, 35530 Bayraklı/İzmir",
     "", "", "https://webusta.com.tr/",
     "",
     "Medium",
     "Folkart binasında web tasarım + destek + siber güvenlik. Telefon web'de yok — form ile randevu."),
    ("Çınar Ajans", "Agency", "Bayraklı", "Adalet Mah.",
     "Folkart Towers A Kule K:28 Ofis:2802, Bayraklı/İzmir (kat/daire telefondan doğrulanmalı)",
     "", "", "https://cinarajans.com.tr/",
     "",
     "Medium",
     "Folkart A Kule — dijital/medya/iletişim ajansı. Kat/daire doğrulanmalı."),
    ("Roof Digital", "Agency", "Bayraklı", "Adalet Mah.",
     "Folkart Towers A Kule K:26 Ofis:2601, Bayraklı/İzmir (kat/daire telefondan doğrulanmalı)",
     "+90 850 303 06 99", "", "https://roofdigital.com/",
     "LinkedIn: roof-digital",
     "Medium",
     "Google/Meta Ads + web tasarım. Folkart A Kule cluster'ı."),

    # =====================================================
    # Mansuroğlu / Bayraklı ekstra
    # =====================================================
    ("MEK Ajans", "Agency", "Bayraklı", "Mansuroğlu Mah.",
     "Mansuroğlu Mh. Ankara Cd. No:81 Bayraklı Tower K:9 Ofis:55, Bayraklı/İzmir",
     "+90 232 344 91 09", "", "https://mekajans.com/",
     "LinkedIn: mek-ajans",
     "High",
     "Bayraklı Tower — Folkart'ın yanında 2. yüksek bina. Kurumsal site + e-ticaret + Google Ads."),
    ("Proji Digital", "Agency", "Bayraklı", "Mansuroğlu Mah.",
     "Mansuroğlu Mh. 288/4. Sk. No:10 K:5 D:30, Bayraklı/İzmir",
     "+90 232 433 74 16", "info@proji.com.tr", "https://proji.com.tr/",
     "LinkedIn: projicomtr",
     "High",
     "Çift-şehir (İzmir + İstanbul Sarıyer) ajansı — 2015'ten beri. LinkedIn-aktif."),
    ("Veritas Dijital", "Agency", "Konak", "Halkapınar Mah.",
     "Halkapınar Mh. 1348 Sk. Keremoğlu İş Merkezi No:2 B Blok K:3 D:301-302, Konak/İzmir",
     "+90 850 259 35 67", "info@veritasdijital.com", "https://veritasdijital.com/",
     "LinkedIn: veritas-dijital",
     "High",
     "360° dijital pazarlama — 4 ofis (İzmir + İstanbul + Kocaeli + Londra). Halkapınar metro yakını."),
    ("FikirMedya", "Agency", "Bayraklı", "Adalet Mah.",
     "Adalet Mh. 1600 Sk. No:3A, Bayraklı/İzmir",
     "", "info@fikirmedya.com.tr", "https://fikirmedya.com.tr/",
     "LinkedIn: fikirmedya",
     "Medium",
     "2004'ten beri reklam + kreatif ajansı. Adalet Mah — Folkart bölgesi ile aynı semt."),
    ("Dijital Arı", "Agency", "Bayraklı", "Bayraklı",
     "Bayraklı/İzmir (tam adres telefondan doğrulanmalı)",
     "", "", "",
     "LinkedIn: dijital-ari",
     "Low",
     "Web + yazılım + sosyal medya + prodüksiyon. Adres doğrulanmadan gitme."),

    # =====================================================
    # Konak / Alsancak (şehir merkezi — yürüyerek cluster)
    # =====================================================
    ("Atölye15", "Agency", "Konak", "Akdeniz Mah.",
     "Akdeniz Mh. Cumhuriyet Blv. No:120, 35210 Konak/İzmir",
     "+90 232 464 04 15", "hello@atolye15.com", "https://atolye15.com/",
     "LinkedIn: atolye15",
     "High",
     "Full-service ürün geliştirme — web + mobil + UI/UX. Cumhuriyet Bulvarı merkezi, Alsancak'a yürüme mesafesi."),
    ("1007 Medya", "Agency", "Konak", "Akdeniz Mah.",
     "Akdeniz Mh. Şht. Fethi Bey Cd. Yüncüler İş Hanı No:77 D:11, Konak/İzmir 35210 "
     "(alt adres: 1347 Sk. No:8 Anba İş Hanı K:1 D:107, Pasaport-Alsancak)",
     "+90 533 260 51 39", "info@1007medya.com", "https://www.1007medya.com/",
     "LinkedIn: 1007medya",
     "High",
     "10+ yıl SMB web — Pasaport/Konak ofisi. Adres iki kaynakta farklı — önce telefon onayı."),
    ("Calisto Ajans", "Agency", "Konak", "Alsancak Mah.",
     "Kıbrıs Şehitleri Cd., Alsancak, 35220 Konak/İzmir (kat/daire telefondan)",
     "+90 232 421 01 70", "", "https://calistoajans.com.tr/",
     "LinkedIn: calisto-ajans",
     "High",
     "2015'ten beri kurumsal web + e-ticaret + kurumsal kimlik. Alsancak ana cadde."),
    ("Ervsoft", "Agency", "Konak", "Kemeraltı",
     "Mucibur Rahman Cd. Salihağa İş Hanı No:3 Ofis:517, Kemeraltı Çarşısı, 35250 Konak/İzmir",
     "+90 232 332 15 05", "info@ervsoft.com", "https://ervsoft.com/",
     "LinkedIn: ervsoft",
     "High",
     "Web + mobil + yazılım — 2012'den beri. Kemeraltı tarihî çarşı — walk-in mesai saatinde."),
    ("İnteso Yazılım", "Agency", "Konak", "Çınarlı Mah.",
     "Çınarlı Mh. 1572/1. Sk. No:33, Konak/İzmir",
     "+90 232 469 37 60", "info@intesoyazilim.com", "https://intesoyazilim.com/",
     "LinkedIn: inteso-yazilim",
     "High",
     "2009'dan beri lider web ajansı. Çınarlı ofis — Mansuroğlu'na yakın."),
    ("Vevona", "Agency", "Konak", "Önder",
     "1583/3 Sk. No:3 K:2 Büro:131, Önder, 35410 Konak/İzmir",
     "+90 506 843 94 96", "info@vevona.com", "https://vevona.com/",
     "LinkedIn: vevona",
     "High",
     "Google/Facebook Ads + sosyal medya + özel yazılım + mobil. Önder semt — Karabağlar'a yakın."),
    ("Localveri Yazılım", "Agency", "Konak", "Mimar Sinan Mah.",
     "Mimar Sinan Mh. Ali Çetinkaya Blv. No:50 D:7 Demet Apartmanı, Alsancak-Konak/İzmir",
     "+90 232 265 18 18", "info@localveri.com.tr", "https://www.localveri.com.tr/",
     "LinkedIn: localveri",
     "High",
     "22 yıl, 850+ müşteri, 4000+ proje. Alsancak ana omurga."),
    ("Ege Bilişim Teknolojileri", "Agency", "Konak", "Alsancak Mah.",
     "Alsancak Mh. Kıbrıs Şehitleri Cd. 1479. Sk. No:11/5, 35220 Konak/İzmir",
     "+90 232 445 66 15", "", "https://egebt.com/",
     "LinkedIn: ege-bilişim-teknolojileri",
     "Medium",
     "2004'ten beri BT + yazılım + ERP + e-ticaret. Alsancak merkez."),
    ("Family Ajans", "Agency", "Konak", "Alsancak Mah.",
     "Alsancak Mh. Kıbrıs Şehitleri Cd. 1443. Sk. No:185, Konak/İzmir",
     "+90 232 421 01 70", "", "https://familyajans.com/",
     "",
     "Medium",
     "Kurumsal tasarım ajansı — Alsancak 1443. Sk. Calisto ile aynı telefon gözüküyor: ikisini aynı gün görmek mantıklı."),
    ("Brainworks", "Agency", "Konak", "Akdeniz Mah.",
     "Akdeniz Mh. 1338 Sk. Business Centre No:2/8 K:2 D:22, 35220 Konak/İzmir",
     "+90 232 445 04 25", "", "https://brainworks.global/",
     "LinkedIn: brainworksglb",
     "High",
     "360° dijital reklam + yazılım. Akdeniz Mah. — Atölye15 ile aynı semt."),
    ("İnteraktif Medya Ajansı", "Agency", "Konak", "Mansuroğlu Mah.",
     "Mansuroğlu Mh. İnce Memed Plaza 1593/1 Sk. No:32 K:6 D:32, Konak/İzmir",
     "+90 232 445 04 25", "", "https://interaktifmedyaajansi.com/",
     "",
     "High",
     "Ödüllü Google Ads ajansı — 2010'dan beri. Brainworks ile aynı telefon — aynı grup olabilir, tek ziyaret iki kart."),
    ("Mantar Medya", "Agency", "Konak", "Çınarlı Mah.",
     "Çınarlı Mh. 1563 Sk. No:2 K:6 D:610 İnci Tower, Konak/İzmir",
     "", "", "https://mantarmedya.com/",
     "IG: @mantarmedya",
     "Medium",
     "Sosyal medya + video prodüksiyon + web tasarım + influencer. İnci Tower — Mansuroğlu'nun başka binası."),
    ("Office701", "Agency", "Konak", "İsmet Kaptan Mah.",
     "İsmet Kaptan Mh. Hürriyet Bul. No:5 D:401 Tezol İş Merkezi, Konak/İzmir",
     "+90 850 441 47 01", "destek@office701.com", "https://office701.com/",
     "LinkedIn: office701",
     "High",
     "10+ yıl, 587 proje. Hürriyet Bulvarı merkez — Konak metro yakını."),
    ("İzmir Reklam Ajansı", "Agency", "Konak", "Atilla Mah.",
     "Atilla Mh., 35270 Konak/İzmir (sokak/no telefondan)",
     "+90 536 608 55 57", "", "https://izmirreklamajans.com/",
     "",
     "Low",
     "Sosyal medya + web + SEO. Atilla Mah — Konak'ın güneyi. Önce telefon."),
    ("Egegen", "Agency", "Konak", "Alsancak Mah.",
     "Alsancak/Konak-İzmir (tam adres telefondan doğrulanmalı)",
     "", "", "https://egegen.com/",
     "LinkedIn: egegen",
     "High",
     "2000'den beri ödüllü 360° entegre iletişim ajansı. İzmir veteran — anchor account."),
    ("Astro Dijital", "Agency", "Konak", "Konak",
     "Konak/İzmir (tam adres telefondan doğrulanmalı)",
     "", "", "https://astrodijital.com/",
     "LinkedIn: astrodijital",
     "Medium",
     "2019 kuruluşu — ürün geliştirme + web + mobil + SEO. Genç ajans — yeni araca açık."),
    ("W-2GO Creative", "Agency", "Konak", "—",
     "Konak/İzmir (tam adres telefondan doğrulanmalı)",
     "", "info@w-2go.com", "https://w-2go.com/",
     "LinkedIn: w-2go-creative",
     "Medium",
     "Grafik + web + sosyal medya + özel yazılım — 2010'dan beri."),
    ("Creatag Medya", "Agency", "Konak", "Alsancak Mah.",
     "Alsancak/Konak-İzmir (tam adres telefondan doğrulanmalı)",
     "", "info@creatag.com.tr", "https://creatag.com.tr/",
     "LinkedIn: creatag-medya",
     "Medium",
     "Dijital branding + sosyal medya + web + foto/video prodüksiyon."),
    ("Venüs Ajans", "Agency", "Konak", "—",
     "Konak/İzmir (tam adres telefondan)",
     "", "info@venusajans.com", "https://venusajans.com/",
     "LinkedIn: venus-ajans",
     "Low",
     "Dijital reklam + sosyal medya + web + SEO."),
    ("Aren Digital", "Agency", "Konak", "—",
     "Konak/İzmir (tam adres telefondan)",
     "", "", "https://arendigital.net/",
     "LinkedIn: aren-digital",
     "Low",
     "Web tasarım odaklı. Website bilgisi az — önce telefon."),

    # =====================================================
    # Karşıyaka (metro bölgesi, şehrin kuzey yakası)
    # =====================================================
    ("Digital Karınca", "Agency", "Karşıyaka", "Park Yaşam Ticaret Merkezi",
     "6523 Sk. No:32/CB Park Yaşam Ticaret Merkezi AVM Girişi, Karşıyaka/İzmir",
     "+90 553 315 62 10", "bilgi@digitalkarinca.com", "https://digitalkarinca.com/",
     "LinkedIn: digital-karinca",
     "High",
     "11+ yıl kurumsal web + özel yazılım + mobil. Park Yaşam AVM — transit friendly."),
    ("İmgesel Tasarım", "Agency", "Karşıyaka", "—",
     "Ordu Bulvarı No:63/B, Karşıyaka/İzmir (kat/daire belirsiz)",
     "+90 537 381 73 57", "satis@imgeseltasarim.com", "https://www.imgeseltasarim.com/",
     "LinkedIn: imgesel-tasarim",
     "Medium",
     "Marka ajansı — 12 yıl. Ordu Blv. ana cadde; dükkân tipi ofis."),
    ("Sibersonik Web Tasarım", "Agency", "Karşıyaka", "—",
     "Karşıyaka/İzmir (tam sokak/no telefondan)",
     "+90 232 366 99 44", "", "https://sibersonik.com/",
     "IG: @sibersonik",
     "High",
     "Web tasarım + SEO + kurumsal. Karşıyaka yerel shop — walk-in friendly telefon sonrası."),
    ("Murat Sümbül SEO Ajansı", "Freelancer", "Karşıyaka", "—",
     "Erdoğan Akkaya Cd. No:2, Karşıyaka/İzmir",
     "+90 532 328 41 11", "info@muratsumbul.com.tr", "https://muratsumbul.com.tr/",
     "",
     "Medium",
     "SEO uzmanı, 1-kişilik operasyon. Yappaflow 'solo stüdyo hacmi' pitch'i birebir."),
    ("Swonie Creative", "Platform Partner", "Karşıyaka", "—",
     "Karşıyaka/İzmir (tam adres telefondan)",
     "", "hello@swonie.com", "https://swonie.com/",
     "LinkedIn: swonie",
     "High",
     "Shopify + branding butik. Platform-partner → Yappaflow e-ticaret hand-off fit."),
    ("GO Dijital", "Agency", "Karşıyaka", "—",
     "Karşıyaka/İzmir (tam adres telefondan)",
     "", "", "https://godijital.net/",
     "LinkedIn: go-dijital",
     "Medium",
     "Web + SEO + Google/Facebook Ads — 10+ yıl."),
    ("Kreamotif", "Agency", "Karşıyaka", "—",
     "Karşıyaka/İzmir (tam adres telefondan)",
     "", "", "https://kreamotif.com/",
     "LinkedIn: kreamotif",
     "Medium",
     "2017 kuruluşu — dijital tasarım + yazılım + mobil + SEO."),
    ("Ruber Media", "Agency", "Karşıyaka", "—",
     "Karşıyaka/İzmir (tam adres telefondan)",
     "", "", "https://rubermedia.com/",
     "LinkedIn: ruber-media",
     "Medium",
     "SEO + web + grafik + performans reklam."),

    # =====================================================
    # Bornova (OMÜ/Ege ekosistemi)
    # =====================================================
    ("YCF Digital", "Platform Partner", "Bornova", "Yeşilova Mah.",
     "Yeşilova Mh. 4174/5 Sk. No:1, Bornova/İzmir",
     "+90 531 400 39 39", "sales@ycfdigital.com", "https://ycfdigital.com/",
     "LinkedIn: ycfdigital",
     "High",
     "ikas + Shopify + Amazon partner. Bornova + USA ofis. Mansur Sarı kurucu."),
    ("Emre Alkaç", "Freelancer", "Bornova", "Erzene Mah.",
     "Erzene Mh. Gençlik Cd. No:11, Bornova/İzmir",
     "+90 505 257 59 25", "ivan@emrealkac.com", "https://www.emrealkac.com/",
     "LinkedIn: emre-alkac",
     "High",
     "İzmir + Ankara çift ofis freelance. Bornova üniversite bölgesi — kafe randevusu kolay."),
    ("Goiz Dijital", "Agency", "Bornova", "Kazımdirik Mah.",
     "Kazımdirik Mh. 229/2. Sk. No:15 İç Kapı:3, Bornova/İzmir",
     "+90 535 363 35 00", "info@goizdijital.com", "https://www.goizdijital.com/",
     "LinkedIn: goiz-dijital",
     "Medium",
     "Web + dijital pazarlama. Kazımdirik = Ege Üniversitesi yakını; öğrenci/esnaf karışımı."),
    ("Cool Digital Solutions", "Agency", "Bornova", "—",
     "Ankara Cd. No:172/67, 35530 Bornova/İzmir",
     "", "hello@cooldijitalcozumler.com", "https://cooldijitalcozumler.com/",
     "LinkedIn: cool-digital-solutions",
     "High",
     "Tailor-made web + proje yönetimi. Ankara Cd — Bornova'nın ana arteri."),
    ("Kristal Fikirler", "Agency", "Bornova", "Çınar Mah.",
     "Çınar Mh. 5003. Sk. No:2/2 Ofis:404, 35100 Bornova/İzmir",
     "+90 533 727 94 69", "", "https://kristalfikirler.com/",
     "LinkedIn: kristal-fikirler",
     "High",
     "Dijital reklam + web + sosyal medya + Google/Instagram Ads."),
    ("Serdar Öztürk", "Freelancer", "Bornova", "—",
     "Bornova/İzmir — freelance (telefon/LinkedIn üzerinden randevu)",
     "+90 553 298 47 48", "", "https://ozturkserdar.com/",
     "LinkedIn",
     "High",
     "WordPress + e-ticaret + web tasarım. Bornova ikamet — üniversite çevresi."),

    # =====================================================
    # Çiğli
    # =====================================================
    ("Deniz Web Ajans", "Agency", "Çiğli", "Ataşehir Mah.",
     "Ataşehir Mh. 8001/3 Sk. Kaşarcı Sitesi A Blok No:7 K:4 D:20, Çiğli/İzmir",
     "+90 507 868 27 75", "info@denizweb.com.tr", "https://deniz-web.com/",
     "LinkedIn: deniz-web-ajans",
     "High",
     "360° ajans — 2018'den beri. Çiğli Ataşehir — Bayraklı-Karşıyaka hattının sonu; tek güne sığar."),
    ("3adım Dijital Ajans", "Agency", "Çiğli", "Küçük Çiğli",
     "Küçük Çiğli 8758. Sk. No:18 D:1, 35620 Çiğli/İzmir",
     "", "", "https://3adimajans.com/",
     "",
     "Medium",
     "Tabela + reklam + web + sosyal medya. Deniz Web ile aynı Çiğli güzergâhı."),

    # =====================================================
    # Balçova (termal + havalimanı aksı)
    # =====================================================
    ("İzmir Web Ajans", "Agency", "Balçova", "Onur Mah.",
     "Onur Mh. Limon Sk. No:1/73 Duru Plaza İş Merkezi, Balçova/İzmir",
     "+90 554 363 14 91", "info@izmirwebajans.com", "https://www.izmirwebajans.com/",
     "LinkedIn: izmir-web-ajans",
     "High",
     "Local İzmir shop — warm opener: 'İzmir'li bir tool olarak...'. Balçova Onur Mah. metro yakını."),
    ("İZ AJANS", "Agency", "Balçova", "Onur Mah.",
     "Onur Mh. Limon Sk. No:1/73, Balçova/İzmir (aynı bina İzmir Web Ajans ile?)",
     "+90 850 203 10 17", "info@izajans.com", "https://calisma.izajans.com/",
     "",
     "Medium",
     "Full-service web + e-ticaret + mobil + video. Adres İzmir Web Ajans ile benzer — aynı binada çift pitch."),

    # =====================================================
    # Karabağlar
    # =====================================================
    ("Onex Yazılım", "Agency", "Karabağlar", "Vatan Mah.",
     "Vatan Mh., Karabağlar/İzmir (sokak/no telefondan)",
     "+90 232 339 39 00", "", "https://onexyazilim.com/",
     "LinkedIn: onex-yazilim",
     "High",
     "Özel yazılım + web/mobil + blockchain + data science. Karabağlar — güney İzmir; tek rota İZ AJANS ile."),

    # =====================================================
    # Kemalpaşa / regional
    # =====================================================
    ("Grande Ajans", "Agency", "Kemalpaşa", "—",
     "Kemalpaşa + Bornova + Bayraklı + Torbalı (hizmet bölgesi — ana ofis Kemalpaşa)",
     "", "", "https://grandeajans.com/",
     "",
     "Low",
     "Mobile-uyumlu + SEO-friendly website. Ofis Kemalpaşa — İzmir merkez dışı (~30 km)."),

    # =====================================================
    # Freelancers (portföy-lider solo web devs)
    # =====================================================
    ("Tezel Zenginoğlu", "Freelancer", "İzmir", "—",
     "İzmir — freelance (remote + in-person randevu)",
     "+90 553 558 90 41", "tezelzenginoglu@gmail.com", "https://www.tezelzenginoglu.com/",
     "LinkedIn: tezel-zenginoğlu",
     "High",
     "Portfolio-led freelance — güçlü web işi. Yappaflow solo-pitch birebir fit."),
    ("Uğur Topuz", "Freelancer", "İzmir", "—",
     "İzmir — freelance (100+ site portföyü)",
     "", "", "https://www.ugurtopuz.com/",
     "",
     "High",
     "100+ website — solid volume. İletişim bilgisi gizli; IG/portföy formu ile."),
    ("Ali Gökalp (A&G)", "Freelancer", "Konya / İzmir", "—",
     "Konya merkezli + İzmir ikinci pazar — freelance",
     "+90 507 736 26 88", "", "https://aligokalp.com.tr/",
     "LinkedIn: ali-gökalp",
     "High",
     "Konya Web Tasarım markası. İzmir pazarlıyor — İzmir ziyaretinde kahve randevusu mümkün."),
    ("Batuhan Özyavru", "Freelancer", "İzmir", "—",
     "İzmir — freelance (WordPress/WooCommerce + SEO)",
     "+90 554 167 16 15", "", "https://batuhanozyavru.com.tr/",
     "",
     "High",
     "WordPress + WooCommerce + web tasarım + SEO. Aktif telefon — WhatsApp randevu."),
    ("Gökmen Bekar", "Freelancer", "İzmir", "—",
     "İzmir — freelance (React/Vue full-stack + tasarım)",
     "", "", "https://gokmenbekar.com/",
     "LinkedIn: gokmenbekar",
     "High",
     "Full-stack web — React, Vue. Modern tech; Yappaflow code-export pitch ona uygun."),
    ("Cemal Karapınar", "Freelancer", "İzmir", "—",
     "İzmir — freelance (WordPress + web tasarım)",
     "", "", "https://cemalkarapinar.com.tr/",
     "LinkedIn: cemalkarapinar97",
     "High",
     "WordPress uzmanı — genç freelancer, LinkedIn aktif."),
    ("Oğuz Atılan", "Freelancer", "İzmir", "—",
     "İzmir — freelance (UI/UX + ürün + web)",
     "", "", "https://oguzatilan.com/",
     "Behance: OguzAtilan",
     "Medium",
     "UI/UX + ürün tasarım + web. Design-first freelancer."),
    ("Salim Dabanca", "Freelancer", "İzmir", "—",
     "İzmir — freelance (13+ yıl UI/UX + web)",
     "", "", "https://behance.net/ux-mars",
     "LinkedIn: salim-dabanca · Dribbble",
     "Medium",
     "13+ yıl UI/UX + web. Kıdemli — consultancy fit, hızlı müşteri deneyimi."),
    ("Ayşe Akar", "Freelancer", "Ankara / İzmir", "—",
     "Ankara + İzmir — freelance (LinkedIn first)",
     "", "", "",
     "LinkedIn: ayse-akar",
     "Medium",
     "Front-end + WordPress. Çift-şehir — İzmir ziyareti sırasında önceden randevu."),
    ("Özlem Design", "Freelancer", "İzmir", "—",
     "İzmir — freelance (web + grafik + içerik)",
     "", "", "https://www.behance.net/ozlemdesign",
     "Behance: ozlemdesign",
     "Medium",
     "Web + grafik tasarım + içerik. Behance portföy odaklı."),
    ("Buse Eker", "Freelancer", "İzmir", "—",
     "İzmir — freelance (product/UI + web)",
     "", "", "https://www.behance.net/buseeker",
     "LinkedIn · X: @buuseker",
     "Low",
     "Product/UI designer — web'e yan odak. Genç portföy."),
    ("Berktuğ Mutlu", "Freelancer", "İzmir", "—",
     "İzmir — freelance (product + UI + web)",
     "", "", "https://www.behance.net/berktugmutlu",
     "LinkedIn · Dribbble",
     "Low",
     "Product designer — UI/UX odaklı. Web ikincil."),

    # =====================================================
    # Platform Partners (Shopify / ikas / Webflow)
    # =====================================================
    ("Cartel (ikas Partner)", "Platform Partner", "İzmir", "—",
     "İzmir (tam adres telefondan — ikas Partner dizini)",
     "", "", "https://cartel.com.tr/",
     "",
     "High",
     "Resmi ikas Partner — e-ticaret kurulum/yönetim. Platform-partner tipik Yappaflow sınıfı."),
    ("Protan (ikas + ideasoft Partner)", "Platform Partner", "İzmir", "—",
     "İzmir + diğer şehirler (tam adres telefondan)",
     "", "", "https://protan.com.tr/",
     "",
     "High",
     "ikas + ideasoft çift-partner. Yappaflow ZIP export ↔ platform import akışı."),
    ("Webflower (Webflow Expert)", "Platform Partner", "İzmir", "—",
     "İzmir (tam adres telefondan)",
     "", "", "https://webflower.co/",
     "IG: @webflowerco",
     "High",
     "Webflow Expert Partner — Türkiye'de nadir. Yappaflow = Webflow alternatifi olarak pozisyon."),

    # =====================================================
    # Diğer (verified company + unverified address)
    # =====================================================
    ("KordonSoft", "Agency", "İzmir", "—",
     "İzmir (tam adres telefondan)",
     "", "info@kordonsoft.com", "https://kordonsoft.com/",
     "LinkedIn: kordonsoft",
     "Medium",
     "2005'ten beri web + yazılım + SEO + e-ticaret. 20 yıl İzmir veteran."),
    ("CEO Yazılım", "Agency", "İzmir", "—",
     "İzmir (tam adres telefondan)",
     "", "", "https://ceoyazilim.com/",
     "LinkedIn: ceo-yazilim",
     "Medium",
     "2007 kuruluşu — web + SEO + özel yazılım + mobil. 18 yıl pazarda."),
    ("Smartmetrics", "Agency", "İzmir", "—",
     "İzmir (tam adres telefondan — 30-kişilik ekip)",
     "", "", "https://smartmetrics.com.tr/",
     "LinkedIn: smartmetrics",
     "Medium",
     "15 yıl, 30 kişi — web/mobil yazılım + dijital reklam. Orta-enterprise profili."),
    ("İzmir Web Tasarım (IWT)", "Agency", "İzmir", "—",
     "İzmir (tam adres telefondan)",
     "", "", "https://izmirwebtasarim.com/",
     "LinkedIn: iwt-dijital-medya",
     "Medium",
     "Kurumsal web + SEO + kurumsal kimlik. Generik isim — doğrulama önce."),
    ("Olay Yeri Ajans", "Agency", "İzmir", "—",
     "İzmir (tam adres telefondan)",
     "", "", "https://olayyeriajans.com/",
     "LinkedIn: olay-yeri-ajans",
     "Medium",
     "Dijital pazarlama + web yazılım — 15 yıl."),
    ("Kaktüs Yazılım", "Agency", "İzmir", "—",
     "İzmir (tam adres telefondan)",
     "", "info@kaktusyazilim.com", "https://kaktusyazilim.com/",
     "LinkedIn: kaktus-yazilim",
     "Medium",
     "UX/UI + mobile-responsive + SEO-friendly + özel yazılım."),
    ("More Dijital Ajans", "Agency", "İzmir", "—",
     "İzmir (tam adres telefondan)",
     "", "", "https://moredijital.com/",
     "LinkedIn: moredijitalajans",
     "Medium",
     "Yazılım + tasarım + pazarlama + analitik birleşimi."),
    ("Gidea Medya Teknoloji", "Agency", "İzmir", "—",
     "İzmir (tam adres telefondan)",
     "", "", "https://gidea.tr/",
     "LinkedIn: gidea-medya",
     "Medium",
     "Web + e-ticaret + grafik + kurumsal kimlik + SEO."),
    ("Ege Ad Works", "Agency", "İzmir", "—",
     "İzmir (tam adres telefondan)",
     "", "", "https://egeadworks.com/",
     "",
     "Medium",
     "Web + SEO + Google Ads + sosyal medya + performans pazarlama."),
    ("Jeton Dijital", "Agency", "İzmir", "—",
     "İzmir (tam adres telefondan)",
     "", "", "https://jetondijital.com/",
     "",
     "Medium",
     "Web + yazılım + dijital pazarlama. Global remote profili."),
    ("İzmir Medya A.Ş.", "Agency", "İzmir", "—",
     "İzmir (tam adres telefondan — multi-district hizmet)",
     "", "", "https://izmirmedya.com.tr/",
     "LinkedIn: izmir-medya",
     "Low",
     "Web + SEO + sosyal medya. Tüm İzmir ilçelerinde hizmet."),
    ("Ege Dijital", "Agency", "İzmir", "—",
     "İzmir (tam adres telefondan)",
     "", "", "https://egedijital.com/",
     "LinkedIn: ege-dijital",
     "Low",
     "Kurumsal web + yazılım danışmanlığı."),
    ("Fok Dijital", "Agency", "İzmir", "—",
     "İzmir (tam adres telefondan)",
     "", "", "https://fok.com.tr/",
     "LinkedIn: fok-dijital",
     "Low",
     "Sosyal medya içerik + web + video + katalog."),
    ("RE Media", "Agency", "İzmir", "—",
     "İzmir (tam adres telefondan)",
     "", "", "https://remediaworks.com/",
     "LinkedIn: re-media-ajansi",
     "Low",
     "Dijital reklam + sosyal medya + grafik."),
    ("Zero Ajans", "Agency", "İzmir", "—",
     "İzmir (tam adres telefondan)",
     "", "", "https://zeroajans.com/",
     "LinkedIn: zero-ajans",
     "Low",
     "Kreatif + dijital reklam ajansı."),
    ("Ajans Bee", "Agency", "İzmir", "—",
     "İzmir (tam adres telefondan)",
     "", "", "https://ajansbee.com/",
     "LinkedIn: ajans-bee",
     "Low",
     "360° pazarlama + web."),
    ("Venio", "Agency", "İzmir", "—",
     "İzmir (tam adres telefondan)",
     "", "", "https://venio.com.tr/",
     "LinkedIn: venio-dijital",
     "Low",
     "360° dijital reklam + sosyal medya + video."),
    ("Nokta Art", "Agency", "İzmir", "Piri Reis Mah.",
     "Piri Reis Mh., İzmir (sokak telefondan)",
     "+90 537 971 70 03", "", "https://noktart.com.tr/",
     "",
     "Low",
     "Web tasarım + web uygulamaları + dijital pazarlama."),
    ("CEK Medya", "Agency", "İzmir", "—",
     "İzmir (tam adres telefondan)",
     "", "", "https://cekmedya.com/",
     "LinkedIn: cek-medya",
     "Low",
     "SEO danışmanlığı + web tasarım + dijital reklam."),
    ("Master Tasarım", "Agency", "İzmir", "—",
     "İzmir (tam adres telefondan)",
     "", "", "https://mastertasarim.com/",
     "",
     "Low",
     "Web tasarım + domain + hosting + sunucu + yazılım."),
    ("M'Glay Creative", "Agency", "İzmir", "—",
     "İzmir (tam adres telefondan)",
     "", "", "https://mglaycreative.tr/",
     "",
     "Low",
     "Web + e-ticaret + sosyal medya + Google/Meta Ads."),
    ("SRN Dijital Ajans", "Agency", "İzmir", "—",
     "İzmir (tam adres telefondan)",
     "", "", "https://srndijital.com/",
     "LinkedIn: srn-dijital",
     "Low",
     "Web design/yazılım + reklam + mobil + özel yazılım."),
    ("SistemSensin", "Agency", "İzmir", "—",
     "İzmir (tam adres telefondan — multi-district)",
     "", "", "https://sistemsensin.com/",
     "LinkedIn: sistemsensin",
     "Low",
     "Web tasarım + dijital çözümler. Alsancak'tan Bornova'ya hizmet."),
    ("Five Brand (5Brand)", "Agency", "İzmir", "—",
     "İzmir (tam adres telefondan)",
     "", "", "https://5brand.co/",
     "",
     "Low",
     "Marka kimliği + web tasarım. Küçük boutique."),
    ("Four Dijital", "Agency", "İzmir", "—",
     "İzmir (tam adres telefondan)",
     "", "", "https://fourdijital.com/",
     "",
     "Low",
     "Web + dijital pazarlama + grafik."),
    ("İzmir Creative", "Agency", "İzmir", "—",
     "İzmir (tam adres telefondan)",
     "", "", "https://izmircreative.com/",
     "",
     "Low",
     "Grafik + marka + web tasarım."),
    ("Creas Creative", "Agency", "İzmir", "—",
     "İzmir (tam adres telefondan)",
     "", "", "https://creascreative.com/",
     "",
     "Low",
     "Web + marka + kreatif hizmetler."),
    ("Ajans Esperto", "Agency", "İzmir", "—",
     "İzmir (tam adres telefondan)",
     "", "", "https://ajansesperto.com/",
     "",
     "Low",
     "Web tabanlı projeler + foto + video."),
    ("SerNis Dijital Ajans", "Agency", "İzmir", "—",
     "İzmir (tam adres telefondan)",
     "", "", "https://sernisdijitalajans.com/",
     "",
     "Low",
     "SEO + sosyal medya stratejisi."),
    ("Websight", "Agency", "İzmir", "—",
     "İzmir (tam adres telefondan)",
     "", "", "https://websight.com.tr/",
     "",
     "Low",
     "Web tasarım + dijital pazarlama."),
    ("Realight Ajans", "Agency", "İzmir", "—",
     "İzmir (tam adres telefondan)",
     "", "", "https://realightajans.com/",
     "",
     "Low",
     "SEO + web + sosyal medya + dijital pazarlama."),
    ("Kent Ajans", "Agency", "İzmir", "—",
     "İzmir (tam adres telefondan)",
     "", "", "https://kentajans.com.tr/",
     "",
     "Low",
     "SEO + web + Google Ads + dijital pazarlama."),
    ("Formix Media", "Agency", "İzmir", "—",
     "İzmir (tam adres telefondan)",
     "+90 535 374 48 01", "", "https://formixmedia.com/",
     "",
     "Low",
     "SEO + dinamik tasarım ekibi."),
    ("Portakal Dijital Çözümler", "Agency", "İzmir", "—",
     "İzmir (tam adres telefondan)",
     "", "", "",
     "LinkedIn: web-tasarim-izmir",
     "Low",
     "2015 kuruluşu — 'web tasarım İzmir' jenerik markası."),
    ("İzmir Web Tasarım Ofisi", "Agency", "İzmir", "—",
     "İzmir (tam adres telefondan)",
     "", "", "https://izmirwebtasarimofisi.com/",
     "",
     "Low",
     "Kurumsal web + e-ticaret + mobil."),
    ("İzmir Agency", "Agency", "İzmir", "—",
     "İzmir (tam adres telefondan)",
     "", "", "https://izmiragency.com/",
     "",
     "Low",
     "Web + grafik + sosyal medya + SEO + Google Ads."),
    ("Grafiva", "Agency", "İzmir", "—",
     "İzmir (tam adres telefondan)",
     "", "", "https://grafiva.com/",
     "",
     "Low",
     "Web + grafik + marka."),
    ("EF Tasarım", "Agency", "İzmir", "—",
     "İzmir (tam adres telefondan)",
     "", "", "https://eftasarim.com/",
     "",
     "Low",
     "Web + kurumsal kimlik + mobil."),
]

# Header row
izmir_hdr_row = 4
for col_idx, h in enumerate(IZMIR_HEADERS, start=1):
    c = izmir.cell(row=izmir_hdr_row, column=col_idx, value=h)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = CENTER
    c.border = BORDER
izmir.row_dimensions[izmir_hdr_row].height = 32

# Data rows
for idx, row in enumerate(IZMIR_LEADS, start=1):
    agency, kind, district, neigh, full_addr, phone, email, website, socials, priority, notes = row
    r = izmir_hdr_row + idx
    values = [idx, agency, kind, district, neigh, full_addr, phone, email, website, socials, priority, notes, ""]
    for col_idx, v in enumerate(values, start=1):
        c = izmir.cell(row=r, column=col_idx, value=v)
        c.font = BODY_FONT
        c.alignment = WRAP
        c.border = BORDER
    if idx % 2 == 0:
        for col_idx in range(1, len(IZMIR_HEADERS) + 1):
            izmir.cell(row=r, column=col_idx).fill = ZEBRA
    pri_cell = izmir.cell(row=r, column=11)
    if priority == "High":
        pri_cell.fill = PRIORITY_HIGH
        pri_cell.font = BODY_BOLD
    pri_cell.alignment = CENTER
    izmir.cell(row=r, column=1).alignment = CENTER
    izmir.cell(row=r, column=3).alignment = CENTER
    izmir.cell(row=r, column=4).alignment = CENTER
    izmir.row_dimensions[r].height = 72

# Column widths (same as Samsun)
izmir_widths = {
    "A": 4, "B": 32, "C": 14, "D": 14, "E": 22,
    "F": 58, "G": 20, "H": 28, "I": 32, "J": 30,
    "K": 13, "L": 52, "M": 14,
}
for col, w in izmir_widths.items():
    izmir.column_dimensions[col].width = w

izmir.freeze_panes = "C5"

# Excel Table UI
izmir_last_col = get_column_letter(len(IZMIR_HEADERS))
izmir_last_row = izmir_hdr_row + len(IZMIR_LEADS)
izmir_tbl_ref = f"A{izmir_hdr_row}:{izmir_last_col}{izmir_last_row}"
izmir_table = Table(displayName="IzmirLeads", ref=izmir_tbl_ref)
izmir_table.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium4", showFirstColumn=False, showLastColumn=False,
    showRowStripes=False, showColumnStripes=False,
)
for col_idx in range(1, len(IZMIR_HEADERS) + 1):
    h = izmir.cell(row=izmir_hdr_row, column=col_idx)
    h.fill = PatternFill(fill_type=None)
    h.font = Font(name=FONT, bold=True, color="FFFFFF", size=11)
izmir.add_table(izmir_table)

# Summary
izmir_summary_row = izmir_last_row + 2
izmir.cell(row=izmir_summary_row, column=2, value="İzmir leads total").font = BODY_BOLD
izmir.cell(row=izmir_summary_row, column=3,
           value=f"=COUNTA(B{izmir_hdr_row + 1}:B{izmir_last_row})").font = BODY_FONT
izmir.cell(row=izmir_summary_row + 1, column=2, value="High visit-priority").font = BODY_BOLD
izmir.cell(row=izmir_summary_row + 1, column=3,
           value=f'=COUNTIF(K{izmir_hdr_row + 1}:K{izmir_last_row},"High")').font = BODY_FONT
izmir.cell(row=izmir_summary_row + 2, column=2, value="Bayraklı-based").font = BODY_BOLD
izmir.cell(row=izmir_summary_row + 2, column=3,
           value=f'=COUNTIF(D{izmir_hdr_row + 1}:D{izmir_last_row},"Bayraklı")').font = BODY_FONT
izmir.cell(row=izmir_summary_row + 3, column=2, value="Konak-based").font = BODY_BOLD
izmir.cell(row=izmir_summary_row + 3, column=3,
           value=f'=COUNTIF(D{izmir_hdr_row + 1}:D{izmir_last_row},"Konak")').font = BODY_FONT
izmir.cell(row=izmir_summary_row + 4, column=2, value="Karşıyaka-based").font = BODY_BOLD
izmir.cell(row=izmir_summary_row + 4, column=3,
           value=f'=COUNTIF(D{izmir_hdr_row + 1}:D{izmir_last_row},"Karşıyaka")').font = BODY_FONT
izmir.cell(row=izmir_summary_row + 5, column=2, value="Bornova-based").font = BODY_BOLD
izmir.cell(row=izmir_summary_row + 5, column=3,
           value=f'=COUNTIF(D{izmir_hdr_row + 1}:D{izmir_last_row},"Bornova")').font = BODY_FONT

wb.save(OUT)
print(f"Wrote {OUT}")
print(f"Leads: {len(LEADS)}")
print(f"Samsun leads: {len(SAMSUN_LEADS)}")
print(f"İzmir leads: {len(IZMIR_LEADS)}")
