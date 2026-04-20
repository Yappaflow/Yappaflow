"""
Yappaflow — Marketing Strategy workbook builder

Produces Yappaflow_Marketing_Strategy.xlsx with 7 tabs:
  1. Strategy       — Positioning, ICP, pillars, 12-mo outcomes
  2. Playbook       — Per-platform cadence/format/voice/CTA/KPI
  3. Calendar       — 90-day rolling content calendar (weeks 1-12)
  4. Hooks          — Content hook & idea bank (Türkçe)
  5. Influencers    — Curated TR creator list (YouTube / LinkedIn / IG / TikTok / Newsletter)
  6. Scripts        — TR DM / email templates for influencer & sponsor outreach
  7. Budget         — Monthly budget allocation + KPI targets

Styling matches Yappaflow_TR_Outreach_Leads.xlsx so the two workbooks feel like a pair.
Run:
    python build_marketing.py
    soffice --headless --calc --convert-to xlsx Yappaflow_Marketing_Strategy.xlsx
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


OUT = Path(__file__).parent / "Yappaflow_Marketing_Strategy.xlsx"

# ------------------------------------------------------------------ styling
FONT = "Arial"
TITLE_FONT = Font(name=FONT, size=16, bold=True, color="1F3A5F")
H2_FONT = Font(name=FONT, size=12, bold=True, color="1F3A5F")
HDR_FONT = Font(name=FONT, size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name=FONT, size=10)
BODY_BOLD = Font(name=FONT, size=10, bold=True)
NOTE_FONT = Font(name=FONT, size=9, italic=True, color="555555")

HDR_FILL = PatternFill("solid", start_color="1F3A5F")
SUB_FILL = PatternFill("solid", start_color="DCE3EE")
ZEBRA = PatternFill("solid", start_color="F2F5FA")
PRIORITY_HIGH = PatternFill("solid", start_color="FFE1A8")
PRIORITY_MED = PatternFill("solid", start_color="FFF3CC")
PILLAR_FILL = PatternFill("solid", start_color="E6EEF8")
GOOD_FILL = PatternFill("solid", start_color="D5EAD8")

THIN = Side(border_style="thin", color="B5BECB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_header_row(ws, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER
        cell.border = BORDER


def zebra_body(ws, first_row: int, last_row: int, cols: int) -> None:
    for r in range(first_row, last_row + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.alignment = WRAP
            cell.border = BORDER
            if (r - first_row) % 2 == 1:
                cell.fill = ZEBRA


def set_widths(ws, widths: dict) -> None:
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w


# ====================================================================
wb = Workbook()

# ============================================================= 1. Strategy
strat = wb.active
strat.title = "Strategy"
strat.sheet_view.showGridLines = False

strat["A1"] = "Yappaflow — pazarlama stratejisi (12 aylık)"
strat["A1"].font = TITLE_FONT
strat.merge_cells("A1:F1")

strat["A2"] = (
    "Signal'den ana sayfaya: kullanıcı kahvesini içerken sohbet ediyor, "
    "bitiminde elinde çalışan bir static site ZIP'i oluyor. Bu belge Yappaflow'un "
    "organik sosyal + influencer + topluluk büyüme planıdır. İlk 90 gün content-led "
    "launch, sonraki 9 ay content + partnership flywheel."
)
strat["A2"].font = NOTE_FONT
strat.merge_cells("A2:F2")
strat.row_dimensions[2].height = 48

# Positioning block
strat["A4"] = "Konumlandırma (Positioning)"
strat["A4"].font = H2_FONT
strat.merge_cells("A4:F4")

positioning = [
    ("One-liner (TR)", "Signal'de anlatıyorsun, Yappaflow ana sayfanı kuruyor — kahven soğumadan deploy-ready ZIP senin."),
    ("One-liner (EN)", "Tell your business story in Signal. Yappaflow ships a deploy-ready static website before your coffee cools."),
    ("Elevator pitch", "Yappaflow, freelancer ve küçük ajansların en büyük zaman kaybını (brief alma + ilk sayfa kurma) Signal sohbetine indiriyor. Müşteri yazıyor, AI kimliği çıkarıyor, Yappaflow statik site üretip ZIP olarak veriyor. İlk draft dakikalar içinde, hosting opsiyonel."),
    ("Ana rakipler", "Wix / Squarespace / Hostinger AI Builder / Durable. Farkımız: Türkçe-first, sohbet kaynaklı brief, ZIP ownership (lock-in yok)."),
    ("Neden şimdi?", "Solo freelancer + 2-8 kişilik ajans pazarı Türkiye'de patlıyor; Teknopark İzmir + Samsun OSB hızla büyüyor. SMB'ler 'Instagram + statik site' kombinasyonuyla yaşıyor — bu statik site kısmını otomatikleştirmek büyük değer."),
    ("Brand voice", "Sıcak, Türkçe-first, abartısız, teknik ama jargon-az. Hype yok; 'çalışan site' göster. build-in-public enerjisi."),
]
for i, (k, v) in enumerate(positioning):
    r = 5 + i
    strat.cell(row=r, column=1, value=k).font = BODY_BOLD
    strat.cell(row=r, column=1).fill = SUB_FILL
    strat.cell(row=r, column=1).alignment = WRAP
    strat.cell(row=r, column=1).border = BORDER
    strat.cell(row=r, column=2, value=v).font = BODY_FONT
    strat.cell(row=r, column=2).alignment = WRAP
    strat.cell(row=r, column=2).border = BORDER
    strat.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    strat.row_dimensions[r].height = 42

# ICP block
icp_row = 5 + len(positioning) + 2
strat.cell(row=icp_row, column=1, value="Ideal Customer Profile (ICP)").font = H2_FONT
strat.merge_cells(start_row=icp_row, start_column=1, end_row=icp_row, end_column=6)

icp_hdr = ["Segment", "Kim", "Acı nokta", "Kanal (nerede bulurum)", "Mesaj açısı", "İlk 90 günde hedef"]
for i, h in enumerate(icp_hdr, start=1):
    strat.cell(row=icp_row + 1, column=i, value=h)
style_header_row(strat, icp_row + 1, len(icp_hdr))

icp_rows = [
    (
        "S1 Solo freelancer (Tier-2 şehirler)",
        "22-35 yaş, WordPress/HTML ile site yapıp 8-25k ₺ kesen tek kişi. Samsun, Kayseri, Trabzon, Sakarya, Eskişehir.",
        "Her brief 2-3 gün sürüyor; müşteri değişikliği için tekrar zaman kaybediyor. Fiyat baskısı.",
        "LinkedIn, YouTube (Murat Yücedağ / Adem İlter / Selman Kâhya), bionluk, Behance, Discord (patika.dev), Telegram gruplari",
        "'Brief'i Signal'de al, sabah uyanmadan ilk site ZIP'i hazır olsun — müşteri sana para saysın, sen yönet.'",
        "500 freelancer Yappaflow'u denedi, 50 aktif kullanıcı.",
    ),
    (
        "S2 2-8 kişilik ajans",
        "Sahip + 2-3 developer + 1-2 tasarımcı. Aylık 10-30 proje. İstanbul + İzmir + Ankara yoğunluklu; Tier-2 şehirlerde de büyüyor.",
        "Briefler tutarsız, junior'lar ilk sayfayı 2 günde çıkarıyor, sahip sürekli revize ediyor.",
        "LinkedIn (Yiğit Konur, Baran Cezayirli), Twitter/X TR tech circle, Kommunity etkinlikleri, Teknopark etkinlikleri",
        "'Yappaflow, junior'larının ilk sayfayı 30 dakikaya indirmesini sağlar. Agency margin'i artar.'",
        "20 ajans demo izledi, 5 ödeme yapan ajans.",
    ),
    (
        "S3 SMB sahibi / küçük işletme",
        "Kafe, butik, mimar, klinik, eğitim kursu sahibi. Instagram'da var ama kendi siteleri yok.",
        "'Web sitem olsa iyi olur ama hangi freelancer'a güveneyim? 15k ₺ de vermek istemiyorum.'",
        "Instagram, TikTok (küçük işletme hesapları), yerel Facebook grupları, Google 'X şehri web sitesi' araması, esnaf odası",
        "'Telefonundaki Signal'de sohbet et. Yappaflow işletmenin sitesini yazıyor — 0 ₺ ile başla, beğenirsen host et.'",
        "100 SMB self-serve aktif hesap (freemium → paid conv).",
    ),
]
for i, row in enumerate(icp_rows):
    r = icp_row + 2 + i
    for j, val in enumerate(row, start=1):
        strat.cell(row=r, column=j, value=val).alignment = WRAP
    strat.row_dimensions[r].height = 98
zebra_body(strat, icp_row + 2, icp_row + 1 + len(icp_rows), len(icp_hdr))

# Pillars block
pill_row = icp_row + 2 + len(icp_rows) + 2
strat.cell(row=pill_row, column=1, value="3 içerik sütunu (content pillars)").font = H2_FONT
strat.merge_cells(start_row=pill_row, start_column=1, end_row=pill_row, end_column=6)

pill_hdr = ["Sütun", "Ne", "Neden işe yarar", "Örnek başlıklar", "Ağırlık (%)"]
for i, h in enumerate(pill_hdr, start=1):
    strat.cell(row=pill_row + 1, column=i, value=h)
style_header_row(strat, pill_row + 1, len(pill_hdr))

pill_rows = [
    (
        "P1 Build-in-public",
        "Haftalık 'bu hafta Yappaflow'da neyi ekledik, neyi kırdık' postları + kısa video klipler.",
        "Freelancer'lar kurucuyla ilişki kurmaya bayılır. Ürün-güveni erkenden kurar, launch momentum verir.",
        "'Bugün Signal webhook'u 3. kez kırdım, şu şekilde çözdüm'; 'Haftalık MRR grafiği'; 'Ajans X ilk siteyi deploy etti'",
        "35",
    ),
    (
        "P2 Mikro eğitimler (how-to)",
        "Belirli bir sorun için 60-90 saniyelik TR video / LinkedIn post: 'Freelancer olarak bu akşam 3 site nasıl teslim edilir?', 'Brief'i Signal'de nasıl yazdırırım?'",
        "SEO/algoritma dostu, yeni kullanıcıyı doğrudan ürüne taşır. Kaydedilebilir içerik = uzun kuyruk.",
        "'Signal bot ile 15 dakikada landing page'; 'SMB için 0 kod statik site'; 'ZIP'i Netlify'a 90 saniyede deploy'",
        "40",
    ),
    (
        "P3 Topluluk + partnership",
        "Patika/Kommunity AMA'lar, Samsun+İzmir meetup'lar, YouTuber entegrasyon videoları, newsletter swap.",
        "Uzun vadeli organik huni. Bir influencer 25k kişiye ulaşırken biz 3 farklı yerde görünürüz.",
        "'Murat Yücedağ x Yappaflow WordPress'ten ZIP'e 1-tık'; 'Teknopark İzmir meetup'; 'Turks in Tech newsletter mention'",
        "25",
    ),
]
for i, row in enumerate(pill_rows):
    r = pill_row + 2 + i
    for j, val in enumerate(row, start=1):
        c = strat.cell(row=r, column=j, value=val)
        c.alignment = WRAP
        if j == 1:
            c.fill = PILLAR_FILL
            c.font = BODY_BOLD
    strat.row_dimensions[r].height = 86
zebra_body(strat, pill_row + 2, pill_row + 1 + len(pill_rows), len(pill_hdr))
# sum row
sum_row = pill_row + 2 + len(pill_rows)
strat.cell(row=sum_row, column=4, value="Toplam").font = BODY_BOLD
strat.cell(row=sum_row, column=4).alignment = Alignment(horizontal="right")
strat.cell(row=sum_row, column=5, value=f"=SUM(E{pill_row + 2}:E{pill_row + 1 + len(pill_rows)})").font = BODY_BOLD

# 12-month outcomes
out_row = sum_row + 2
strat.cell(row=out_row, column=1, value="12 aylık hedef çıktılar").font = H2_FONT
strat.merge_cells(start_row=out_row, start_column=1, end_row=out_row, end_column=6)

out_hdr = ["Metrik", "Q1 (Ay 1-3)", "Q2 (Ay 4-6)", "Q3 (Ay 7-9)", "Q4 (Ay 10-12)", "Not"]
for i, h in enumerate(out_hdr, start=1):
    strat.cell(row=out_row + 1, column=i, value=h)
style_header_row(strat, out_row + 1, len(out_hdr))

outs = [
    ("LinkedIn takipçi", 500, 2000, 5000, 10000, "Yusuf + Yappaflow company page birlikte"),
    ("Twitter/X takipçi", 300, 1500, 4000, 8000, "Türk dev/founder çevresi"),
    ("Instagram takipçi", 200, 1000, 3000, 5000, "Reels odaklı; SMB segmenti"),
    ("TikTok takipçi", 100, 800, 2500, 5000, "Quick how-to, build-in-public"),
    ("YouTube subs (opsiyonel)", 50, 300, 1000, 2500, "Long-form tutorial; haftalık 1 video"),
    ("Newsletter abonesi", 100, 500, 1500, 3000, "Substack'te Türkçe + EN"),
    ("Demo izleyen kişi", 50, 300, 800, 2000, "Landing page video + webinar"),
    ("Aktif ücretli müşteri", 5, 30, 100, 250, "Freelancer + ajans + SMB hepsi dahil"),
    ("Case study yayınlanan", 1, 5, 15, 25, "Samsun + İzmir ilk 5 case; video version'lar"),
    ("Influencer kolabı", 0, 3, 8, 15, "YouTube + LinkedIn karışık"),
]
for i, row in enumerate(outs):
    r = out_row + 2 + i
    for j, val in enumerate(row, start=1):
        c = strat.cell(row=r, column=j, value=val)
        c.alignment = WRAP if j in (1, 6) else CENTER
        if j in (2, 3, 4, 5):
            c.number_format = "#,##0"
    strat.row_dimensions[r].height = 22
zebra_body(strat, out_row + 2, out_row + 1 + len(outs), len(out_hdr))

set_widths(strat, {"A": 28, "B": 24, "C": 26, "D": 26, "E": 26, "F": 36})
strat.freeze_panes = "A4"

# ============================================================= 2. Playbook
play = wb.create_sheet("Playbook")
play.sheet_view.showGridLines = False

play["A1"] = "Platform Playbook — kanal bazlı oyun kitabı"
play["A1"].font = TITLE_FONT
play.merge_cells("A1:I1")

play["A2"] = (
    "Her platformun kendi dili, cadence'i ve 'kazanma' tanımı var. Bu tablo Yusuf'un ve Yappaflow "
    "hesabının haftalık olarak ne yapacağına baktığı tek panodur. Kopya-yapıştır içerik yok — her kanalda "
    "nitelik hedefi ve KPI yanında duruyor."
)
play["A2"].font = NOTE_FONT
play.merge_cells("A2:I2")
play.row_dimensions[2].height = 48

play_hdr = [
    "Platform",
    "Hedef kitle",
    "Cadence (haftalık)",
    "Format",
    "Ton / dil",
    "Pillar ağırlığı",
    "CTA",
    "Birincil KPI",
    "Red flags",
]
for i, h in enumerate(play_hdr, start=1):
    play.cell(row=4, column=i, value=h)
style_header_row(play, 4, len(play_hdr))

play_rows = [
    (
        "LinkedIn (Yusuf kişisel)",
        "2-8 kişilik ajans sahipleri + senior freelancer + Türk tech yatırımcı",
        "3 post + 2 yorum seansı/gün + 1 carousel",
        "Metin post (>900 karakter), carousel PDF, kısa native video",
        "Kişisel, build-in-public, Türkçe+İngilizce dengesi (%70 TR)",
        "P1 %50 / P2 %35 / P3 %15",
        "DM → 15 dk demo çağrısı",
        "Post impression > 3000, DM'den demo > 2/hafta",
        "Ajanda-olmayan motivational post, 'LinkedIn post yazar AI' tuzağı",
    ),
    (
        "LinkedIn (Yappaflow company page)",
        "Ajans + SMB karar vericileri",
        "3 post/hafta, 1 canlı yayın (aylık)",
        "Case study, müşteri sözü, ürün güncelleme, ekran kaydı",
        "Kurumsal-sıcak, case-first. Hep Türkçe.",
        "P1 %30 / P2 %40 / P3 %30",
        "Demo → ücretsiz hesap",
        "Page takipçi büyüme %/ay, landing CTR",
        "Kurucu hesabıyla aynı içeriği tekrarlamak",
    ),
    (
        "Twitter / X",
        "Türk dev/founder çevresi + Avrupa indie-hacker kitlesi",
        "15-25 tweet/hafta (thread dahil)",
        "Tek-tweet build-in-public, haftalık thread, ekran kaydı GIF",
        "Samimi, teknik, kısa. Türkçe + İngilizce karışık.",
        "P1 %55 / P2 %30 / P3 %15",
        "Profildeki link → Signal bot dene",
        "Thread impression > 20k, waitlist signup > 30/hafta",
        "TR tech drama'sına dalmak, spam quote-tweet",
    ),
    (
        "Instagram",
        "SMB sahipleri (kafe, butik, klinik), junior tasarımcılar",
        "5 reels + 3 story + 1 carousel/hafta",
        "30-60 sn reels, before/after statik site, 'müşteri konuşturdum, site geldi' klipleri",
        "Görsel-ağırlıklı, tamamen Türkçe, duygusal aç",
        "P1 %20 / P2 %55 / P3 %25",
        "Bio'da link → WhatsApp/Signal deneme",
        "Reels saves+share oranı > %3, profil ziyaret > 2k/hafta",
        "Sadece text slide post; IG algoritması cezalandırıyor",
    ),
    (
        "TikTok",
        "22-30 yaş freelancer, junior dev, SMB meraklısı",
        "7-10 klip/hafta (küçük çapta sıklık KRİTİK)",
        "Pure screen recording + subtitle, 'brief'ten siteye 15 saniye', canlı hata",
        "Hype-az, eğitsel, subtitle ile Türkçe",
        "P1 %30 / P2 %60 / P3 %10",
        "Bio → landing + sabit pinned video",
        "Ortalama view > 5k, kaydedilme oranı > %2",
        "Trend sound baglantısızken sırf trend olsun diye kullanmak",
    ),
    (
        "YouTube (long-form)",
        "Freelancer + öğrenci dev + ajans junior'ları",
        "1 video/hafta (10-18 dk), ayda 1 canlı yayın",
        "'Canlı brief'ten ZIP'e' tutorial'ları, karşılaştırma videoları (vs Wix / Durable)",
        "Ekran+kamera, Türkçe, eğitsel",
        "P1 %15 / P2 %70 / P3 %15",
        "Açıklama → Signal bot + ücretsiz deneme",
        "Izlenme süresi > %50, abone büyüme > 30/hafta",
        "Clickbait başlık + ilk 30 sn'de değer vermemek",
    ),
    (
        "Newsletter (Substack)",
        "İleri freelancer + ajans sahibi + TR tech meraklısı",
        "Haftalık Perşembe 20:00, ayda 1 'deep-dive'",
        "Build journal + müşteri hikayesi + araç ipucu",
        "Kişisel, daha uzun (800-1500 kelime), Türkçe+EN versiyon",
        "P1 %50 / P2 %20 / P3 %30",
        "Yazı sonu → demo video + referans hediye",
        "Açılma oranı > %40, demo tıklaması > 25/issue",
        "Şablon e-posta tonu, güncelleme logu kuruluğu",
    ),
    (
        "Topluluklar (Patika, Kommunity, Telegram, Discord)",
        "Junior dev + freelance topluluğu + Türk tech meetup'ı",
        "Haftada 2 aktif AMA/yardım + aylık 1 etkinlik",
        "Canlı AMA, slack/discord thread, meetup konuşması",
        "Yardımsever, satış-yok, sorulara cevap-önce",
        "P1 %20 / P2 %30 / P3 %50",
        "Direkt link yerine kullanıcı adı + profil",
        "Ayda demo çağrısı > 5, topluluk ismi duyulan > 3",
        "Topluluğa link yapıştırıp kaçmak = sürekli ban",
    ),
]
for i, row in enumerate(play_rows):
    r = 5 + i
    for j, val in enumerate(row, start=1):
        c = play.cell(row=r, column=j, value=val)
        c.alignment = WRAP
        if j == 1:
            c.font = BODY_BOLD
    play.row_dimensions[r].height = 92
zebra_body(play, 5, 4 + len(play_rows), len(play_hdr))

# Excel Table
play_tbl_ref = f"A4:{get_column_letter(len(play_hdr))}{4 + len(play_rows)}"
play_tbl = Table(displayName="PlaybookTbl", ref=play_tbl_ref)
play_tbl.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium4", showFirstColumn=False, showLastColumn=False,
    showRowStripes=True, showColumnStripes=False,
)
play.add_table(play_tbl)

set_widths(play, {"A": 26, "B": 28, "C": 22, "D": 26, "E": 26, "F": 18, "G": 22, "H": 26, "I": 26})
play.freeze_panes = "B5"

# ============================================================= 3. Calendar
cal = wb.create_sheet("Calendar")
cal.sheet_view.showGridLines = False

cal["A1"] = "90 günlük içerik takvimi — haftalık tematik plan"
cal["A1"].font = TITLE_FONT
cal.merge_cells("A1:H1")

cal["A2"] = (
    "Her hafta bir ana tema, her platforma o temadan türeyen 1-2 yayın. İlk 4 hafta 'launch' rüzgarı, "
    "5-8 'partnership wave', 9-12 'case study + ürün lansmanı'. 'Asset' kolonu o hafta için üretilecek "
    "reusable parça (video / carousel / blog)."
)
cal["A2"].font = NOTE_FONT
cal.merge_cells("A2:H2")
cal.row_dimensions[2].height = 48

cal_hdr = ["Hafta", "Tarih aralığı", "Tema", "LinkedIn odak", "Twitter/X odak", "IG + TikTok odak", "Newsletter / YouTube", "Ana üretilecek asset"]
for i, h in enumerate(cal_hdr, start=1):
    cal.cell(row=4, column=i, value=h)
style_header_row(cal, 4, len(cal_hdr))

cal_rows = [
    ("W1", "27 Nis - 3 May 2026", "LAUNCH — 'Signal'den ZIP'e'",
     "Yusuf kişisel launch post + 10 özel kişiye öncelikli DM demo",
     "Launch thread (7 adet tweet) + GIF",
     "Reels: 'Brief'i Signal'de yazdım, ZIP geldi' (60 sn)",
     "NL #1: 'Neden Yappaflow? (Türkçe)'; YT ilk tutorial",
     "Launch thread + 60 sn hero reels + NL #1 (×2 dil)"),
    ("W2", "4 May - 10 May 2026", "Soru & cevap — itirazlar",
     "Carousel: '5 itiraz: Yappaflow güvenli mi, hosting?, SEO?, domain?, özelleştirilebilirlik?'",
     "Thread: 'En çok sorulan 10 soru'",
     "Reels: 3 farklı itiraza 10 sn cevap (serisi)",
     "NL #2: SEO + hosting akışı açıklaması",
     "10-soru FAQ belgesi + 3 Reels"),
    ("W3", "11 May - 17 May 2026", "Freelancer use-case",
     "Carousel + post: 'Freelancer için akşam 3 brief, sabah 3 site'",
     "Build-in-public günlüğü (her gün 1 kısa tweet)",
     "TikTok: 'Freelance 1 gün' vlog + Reels before/after",
     "YT: 'Freelancer'a Yappaflow 14 dk' derin eğitim",
     "14 dk YT video + 5 TikTok klibi"),
    ("W4", "18 May - 24 May 2026", "İlk case study — Samsun freelancer",
     "Case study post + carousel 'before / after'",
     "Thread: 'Müşteriyi nasıl ikna ettim' (freelancer anlatımı)",
     "IG hikaye-serisi: ekran video + kısa alıntılar",
     "NL #3 case study + YT röportaj",
     "Case study PDF + 10 dk röportaj videosu"),
    ("W5", "25 May - 31 May 2026", "Ajans use-case giriş",
     "Post: 'Neden 2-8 kişilik ajanslara Yappaflow gerekir'",
     "Thread: 'Junior'ların ilk sayfa süresini nasıl indirdik' (çizgi)",
     "Reels: 'Ajans sahibine 3 dakika' (kurucu konuşması)",
     "NL #4 + YT demo: ajans akışı",
     "3 dk 'ajans için' açıklama videosu"),
    ("W6", "1 Haz - 7 Haz 2026", "Influencer kolaborasyon #1",
     "Co-marketing post (Yiğit Konur veya Serkan Ünsal newsletter)",
     "Quote-retweet + konuk thread",
     "IG collab reel (konuk YouTuber)",
     "NL #5: 'Guest yazar' formatı",
     "Guest yazısı + cross-post video"),
    ("W7", "8 Haz - 14 Haz 2026", "Ürün derinliği — AI pipeline",
     "Build post: 'DeepSeek + OpenRouter fallback: canlı mimari'",
     "Thread: '3 yanlış yaptık, 3 doğru kurduk'",
     "TikTok: 'Kod ekran kaydı + açıklama' (subtitle Türkçe)",
     "YT deep dive: 12 dk 'AI pipeline'",
     "12 dk YT + mimari carousel"),
    ("W8", "15 Haz - 21 Haz 2026", "İzmir meetup / ön hazırlık",
     "Post: 'İzmir etkinliğine kim gelir?' + duyuru",
     "Twitter space (1 saat, Türkçe)",
     "Reels: İzmir yol videosu + 'gelecek ajansları arıyorum'",
     "NL #6: 'Neden meetup?' + kayıt link",
     "Meetup slayt + kayıt formu"),
    ("W9", "22 Haz - 28 Haz 2026", "İzmir meetup — canlı",
     "Canlı LinkedIn yayın",
     "Live-tweet serisi",
     "IG canlı + Reels highlight",
     "YT: meetup record (30 dk özel)",
     "Meetup highlight reel + YT kayıt"),
    ("W10", "29 Haz - 5 Tem 2026", "SMB use-case — kafe / butik",
     "Post: 'Kafe sahibi Signal'de ne sordu?'",
     "Thread: '5 SMB müşteriden 5 ders'",
     "Reels: 'Butik sahibi için canlı site kurdum' (70 sn)",
     "NL #7: SMB segmenti derin dalış",
     "SMB kılavuzu PDF + Reels serisi"),
    ("W11", "6 Tem - 12 Tem 2026", "Topluluk dalgası — Patika AMA",
     "Duyuru + post-AMA özet carousel",
     "AMA soruları thread",
     "IG live clip + TikTok'a aktar",
     "NL #8: AMA'dan öğrendiğim 3 şey",
     "AMA kayıt + 10 soru-özet carousel"),
    ("W12", "13 Tem - 19 Tem 2026", "Ürün v2 lansmanı + retro",
     "Yusuf kişisel retro post: '90 gün sonuçları, ne öğrendik'",
     "Metrics thread + grafikler",
     "Reels: '90 günde 250 aktif user' kutlama",
     "NL #9: 90-gün retro (uzun yazı)",
     "90-gün retro deck + public KPI dashboard"),
]
for i, row in enumerate(cal_rows):
    r = 5 + i
    for j, val in enumerate(row, start=1):
        c = cal.cell(row=r, column=j, value=val)
        c.alignment = WRAP
        if j == 3:
            c.font = BODY_BOLD
            c.fill = PILLAR_FILL
    cal.row_dimensions[r].height = 70
zebra_body(cal, 5, 4 + len(cal_rows), len(cal_hdr))

cal_tbl = Table(displayName="CalendarTbl", ref=f"A4:{get_column_letter(len(cal_hdr))}{4 + len(cal_rows)}")
cal_tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
cal.add_table(cal_tbl)

set_widths(cal, {"A": 7, "B": 20, "C": 26, "D": 34, "E": 28, "F": 30, "G": 30, "H": 32})
cal.freeze_panes = "C5"

# ============================================================= 4. Hooks
hk = wb.create_sheet("Hooks")
hk.sheet_view.showGridLines = False

hk["A1"] = "İçerik hook bankası — 60+ hazır açılış"
hk["A1"].font = TITLE_FONT
hk.merge_cells("A1:F1")

hk["A2"] = (
    "Her hook doğrudan bir posta çevrilecek şekilde yazıldı. 'Platform' kolonu öncelikli kanalı gösterir "
    "ama çoğu cross-post edilebilir. 'Pillar' kolonu stratejiyle hizalar. Yusuf her hafta min. 8 hook seçip "
    "üretime sokmalı; sonra bu hook 'Used' olarak güncellenir."
)
hk["A2"].font = NOTE_FONT
hk.merge_cells("A2:F2")
hk.row_dimensions[2].height = 48

hk_hdr = ["#", "Platform", "Pillar", "Açılış cümlesi (Türkçe)", "Gövde (ana mesaj)", "CTA"]
for i, h in enumerate(hk_hdr, start=1):
    hk.cell(row=4, column=i, value=h)
style_header_row(hk, 4, len(hk_hdr))

hooks = [
    ("LinkedIn", "P1", "Bugün Yappaflow'da bir şeyi kırdım. Sizinle paylaşmak istedim.",
     "Signal webhook'unda race-condition vardı; 3 farklı log + 1 Slack alarmıyla buldum. Çözüm: 12 satırlık kilitleme. Öğretici yanı: küçük ekip, minik altyapı, büyük ders.", "Aynı problemi yaşayan var mı? Yorumlarda tartışalım."),
    ("LinkedIn", "P2", "2-8 kişilik ajans sahipleri: bu pazarlama postunu 30 sn okuyun.",
     "Junior bir developer'a ilk sayfayı 2 gün kurdurmak ajans marjinizi eritiyor. Yappaflow ile brief Signal'de geliyor, sabah ZIP hazır. 3 İzmir ajansı test etti; ortalama 28 dk.", "DM'den ücretsiz pilot demo rezerve edin."),
    ("LinkedIn", "P3", "İzmir'deyim. 24 Haziran Çarşamba küçük bir ajans meetup'ı düzenliyorum.",
     "Alsancak'ta 2 saat, pizzalar bende. Konular: (1) brief otomatikleştirmek; (2) müşteriye fiyat anlatmak; (3) bir ajans nasıl Yappaflow'u iç pipeline yaptı.", "Gelmek istiyorsanız yoruma 'İzmir' yazın, 20 kişi ile sınırlı."),
    ("Twitter/X", "P1", "Yappaflow MRR'i bu hafta $340 → $420.",
     "Kaç freelancer kayıt oldu, kaç hesap aktifleşti, ne kırdım. Küçük ama doğru yönde.", "Thread içinde tüm grafikler ↓"),
    ("Twitter/X", "P2", "Brief'i Signal'de yazdırmanın 3 kuralı:",
     "1) müşteri sevdiği 3 rakibi söylesin; 2) 'site açıldığında ne hissetsin' sorusu; 3) sitede olmayacakları da listele. Geri kalanı AI hallediyor.", "Detaylı thread ↓"),
    ("Twitter/X", "P3", "Turks in Tech newsletter'ın bu haftaki issue'sunu yazarken fark ettim:",
     "Türk indie-hacker sahnesi 2023'e göre 2 kat büyümüş. Bu, Yappaflow gibi dikey ürünler için harika haber.", "Turks in Tech'e abone olun (link bio'da)."),
    ("Instagram", "P1", "Bir kurucunun 1 dakikası: bu sabah Signal botuna ilk SMB müşterisi yazdı.",
     "30 yaşında bir butik sahibi, telefonundan sohbet etti; sonunda elinde .zip bir site. Hiç 'kodla' demedi.", "Bio'daki linkten Signal'e yaz; sen de dene."),
    ("Instagram", "P2", "Kafeci dostun sordu: sitem olsa faydası olur mu? 60 sn'de gösteriyorum.",
     "Google aramada gözükmezsen komşu kafeyi seçer. Menüyü .zip siteye koyarız, Netlify'a yüklersin, haftaya aramalarda gözükürsün.", "Kaydet + kafeciye gönder."),
    ("Instagram", "P3", "Esnaf odası etkinliğindeydim. 10 SMB'ye Yappaflow'u anlattım. 8'i 'bunu bana lazım' dedi.",
     "Türk SMB'nin en büyük engeli güven: 'parayı verip siteyi bulmazsam' korkusu. Freemium modeli bunun cevabı.", "Aynı etkinlik yapmak isteyen esnaf odası varsa DM'den yazın."),
    ("TikTok", "P1", "Canlı yayın: 5 dakikada bir SaaS ürünü daha iyi nasıl yapılır.",
     "Bugün: Signal bot'un hata mesajlarını yeniden yazdım. Sebep: müşteri 'ne demek bu?' diye sordu.", "Takip et, her gün bir ilerleme paylaşıyorum."),
    ("TikTok", "P2", "15 saniyede: brief'ten siteye.",
     "Signal'e yaz → AI kimlik çıkar → şablon seç → .zip indir. Tam ekran kaydı.", "Detaylı adım adım için LinkedIn'e bak (link bio'da)."),
    ("TikTok", "P3", "Bu hafta Patika'da AMA yaptım. 120 kişi geldi. En çok sorulan soru:",
     "'Ürünü 0'dan tek başına nasıl kurdun?' Cevap: ekipten değil, topluluktan destek aldım.", "Patika'nın Discord'una sen de gel."),
    ("YouTube", "P1", "90 gün sonra: Yappaflow ne öğretti?",
     "Tüm metrikler, hataları, 3 doğru kararı, 2 büyük yanlışı. 18 dk dürüst retro.", "Detay için açıklama + abone ol."),
    ("YouTube", "P2", "Freelancer için Yappaflow tam eğitimi (14 dk).",
     "Brief → AI kimlik → şablon → ZIP → Netlify deploy. Müşteri faturası bile dahil.", "Açıklamadaki Signal linkinden dene."),
    ("YouTube", "P3", "Murat Yücedağ + Yusuf: WordPress'ten ZIP'e geçiş (ortak video).",
     "İki farklı pencere: biri klasik WP freelancer, diğeri static-site founder. Müşteri tercihleri karşılaştırılıyor.", "Murat'ın kanalını da izle + abone."),
    ("Newsletter", "P1", "Yappaflow'ta bu hafta: Signal webhook kırıldı, 3 kişi yardım etti.",
     "1800 kelimelik 'bu hafta ne oldu, ne öğrendim, gelecek hafta ne?' içerik. Açıklık = güven.", "Geri dönüş bekliyorum — cevap gönder."),
    ("Newsletter", "P2", "Bir ajans owner için: Yappaflow pipeline'ını iç araç yapmanın 5 yolu.",
     "Case: İzmir'deki 4 kişilik dijital ajans; iç brief akışlarına Yappaflow ekledi. Junior verimi %40 arttı.", "Demo için yanıtla + Zoom ayarlayalım."),
    ("Newsletter", "P3", "Turks in Tech + Yappaflow: birlikte bir deneme.",
     "Co-issue: kurucu hikayesi + Türk early-stage ürün listesi. Newsletter swap'ı.", "Turks in Tech'e kayıt (link)."),
    ("LinkedIn", "P1", "Bir kurucu olarak en korktuğum şey: 'ürünü kimse istemezse?'",
     "Samsun'daki 2 freelancer ilk ödemeyi yaptığında kalp çarpıntısı yaşadım. Kıymetli bir an.", "Sizin ilk ücretli müşteri anınız nasıldı?"),
    ("LinkedIn", "P2", "'Statik site' 2026'da neden ölmedi, tersine yükseldi?",
     "Hosting ucuz, SEO hızlı, bakım yok, müşteri kontrolü tam. Yappaflow bu trendin AI-first hali.", "Yorumda katılıyor musun?"),
    ("Twitter/X", "P1", "Bu akşam 23:00'te küçük canlı yayın: Yappaflow backend'indeki 3 karar.",
     "Miktarlar: Postgres yerine SQLite, worker sayısı 2, cache 10 dk.", "Saat ayarla ↓"),
    ("Twitter/X", "P2", "SMB sitesi = menü + iletişim + 3 sosyal.",
     "Geriye kalan tüm 'gösterişli sayfalar' satış yapmıyor. Yappaflow bu 4 kutuyla başlıyor, sonra büyüyor.", "Aynı fikirde misin? RT."),
    ("Instagram", "P1", "Kapat telefonu, bu sefer: '.zip' müşteriye e-postayla gitti.",
     "Yusuf'un 3. müşterisi. Şehir: Samsun. Süre: 32 dakika.", "Samsun'da siz de Signal'e yazın."),
    ("Instagram", "P2", "Kafe web sitesi için 4 zorunlu şey.",
     "(1) Menü; (2) harita; (3) telefon; (4) Instagram linki. Fazlası müşteriye kaybettiriyor.", "Kaydet + esnaflara gönder."),
    ("TikTok", "P1", "Samsun'dan İzmir'e: Yappaflow neden şimdi?",
     "İki şehir, iki hikaye, bir pazar. 60 sn.", "Takip et, İzmir meetup için DM at."),
    ("TikTok", "P2", "Ajans sahiplerine 3 saniyelik test:",
     "Son brief kaç saatte aldın? 6+ saat ise senin problemin benim problemim. Yappaflow.", "Bio'da detay."),
    ("LinkedIn", "P1", "3 influencer ile aynı anda konuştum; hepsi aynı soruyu sordu.",
     "Hepsinin cevabı aynıydı: 'SDK'yı verebilir misiniz, Türkçe dokümantasyonla?' Şimdi üzerindeyiz.", "Beta test istersen DM'den yaz."),
    ("Newsletter", "P1", "Bu hafta 3 hata: hangisini tekrar yapmazdım?",
     "1) müşteriye demo'yu canlı yapmadım; 2) fiyat listesinde kademelendirme yoktu; 3) refund politikası 'belirsiz' yazıyordu. Üçü de düzeldi.", "Yanıtlar ve sorular bekliyorum."),
    ("Twitter/X", "P1", "Yappaflow 100. kullanıcısını kutluyor.",
     "40% freelancer, 35% ajans, 25% SMB. Samsun ilk, İzmir ikinci.", "Kıyaslamak isteyenlere thread."),
    ("LinkedIn", "P3", "Kommunity'de açtığım anketten 3 içgörü.",
     "Türk junior dev'lerin %62'si 'ilk müşteri için site kurmayı' en büyük engel görüyor. Yappaflow bu cevabı veriyor.", "Sen de anket yanıtla."),
    ("Instagram", "P3", "Digital Faculty ile kısa röportaj.",
     "Emre Gökşin bir dakikada neden Türkiye'de yerli AI ürün desteklenmeli diye anlatıyor.", "Reels'i kaydet."),
    ("YouTube", "P2", "Wix vs Squarespace vs Yappaflow — Türk SMB için hangisi?",
     "Yan yana canlı test: 3 farklı kullanıcı, 3 farklı sonuç. Maliyet + süre + kalite.", "Yorumlara beklentilerini yaz."),
    ("Instagram", "P2", "Canlı hikaye: Signal sohbetinden ZIP.",
     "3 parçalı story arc: soru → ürün → sonuç.", "Kaydır: Signal linki hikaye sonunda."),
    ("TikTok", "P2", "Müşteri brief'ini otomatikleştirmek = 10 saat kazanmak.",
     "Matematiksel kanıt: haftada 3 müşteri × 3 saat brief = 9 saat. Yappaflow bu 9 saati 30 dakikaya indiriyor.", "DM at, 5 dk demo."),
    ("Newsletter", "P3", "Bir TR startup topluluğu röportajı: Türkiye tech'i hangi yöne gidiyor?",
     "Baran Cezayirli cevap veriyor. Abone olup başa dönmek için link.", "Forward et bir dostuna."),
    ("LinkedIn", "P2", "Fiyat şeffaflığı: Yappaflow neden abonelik değil one-time + hosting?",
     "Özgürlük + ZIP ownership + düşük giriş. Yatırımcıya değil kullanıcıya uygun.", "Yorumlara sor."),
    ("Twitter/X", "P2", "Türkiye'de 'static site' arama trendi 2023→2026 %180 arttı.",
     "Bu neden önemli: SEO + hosting maliyeti + güven. Grafik ekli.", "Google Trends ekran görüntüsü thread'de."),
    ("TikTok", "P3", "BtkAkademi bootcamp öğrencilerine Yappaflow canlı demo.",
     "90 öğrenci ekran karşısında; DM'ler tutulur halde.", "Bootcamp videosunu bio'dan izle."),
    ("Instagram", "P3", "Patika.dev x Yappaflow: ilk 'AMA gecesi' bu perşembe.",
     "Soruları şimdiden yorumlara bırakın; canlıda cevaplıyorum.", "Takvime ekle."),
    ("YouTube", "P3", "Selman Kâhya ile kolaborasyon: 'Sıfırdan WordPress vs ZIP site — 2026 kararı'.",
     "30 dk ortak video; iki farklı iş modeli.", "Her iki kanalı da izle."),
    ("LinkedIn", "P1", "Build-in-public sürdürülebilir mi?",
     "6 ay içinde verdiğim 1 seçim: Twitter'a günlük değil LinkedIn'e haftalık uzun düşünce. Sebep: müşteri orada karar veriyor.", "Senin kanal önceliğin ne?"),
    ("Twitter/X", "P1", "Yappaflow'un gizli silahı: Türkçe prompt tuning.",
     "İngilizce modelde 'salon' = güzellik yerine 'oturma odası'. Bunu düzeltmek = Türkiye'de rakip yok.", "Detay için thread."),
    ("Instagram", "P1", "Bugün 1 saat 37 dakika boyunca sessizlikte kod yazdım.",
     "Bitişinde Signal bot yeni prompt'u öğrenmişti. Reels format: hızlandırılmış ekran.", "Kaydet, başka ne izlemek istersin sor."),
    ("TikTok", "P2", "Müşteri brief'inde 3 sorum varsa senin de 3 sorun var.",
     "(1) kullanıcı ne yapıyor? (2) rakip neden kötü? (3) stil nasıl? Cevabını Signal'e yapıştır, Yappaflow siteyi çıkarıyor.", "Bio'dan başla."),
    ("Newsletter", "P2", "Ajans sahipleri: Yappaflow'u white-label kullanmanın hukuki akışı.",
     "Bayilik değil, 'reseller seat'. Fiyatlandırma + fatura + sözleşme örnekleri.", "Ajans sözleşme şablonunu indir (link)."),
    ("LinkedIn", "P3", "Yiğit Konur guest post: Türk SEO 2026 trendleri.",
     "Bir aylık hazırlık; 2200 kelime; veri 8 farklı kaynaktan.", "Tam metin Substack'te."),
    ("YouTube", "P1", "Canlı build: 2 saatte yeni özellik (AI iyileştirme).",
     "Kamera + ekran; yorumlar üzerinden karar alıyorum.", "Abone ol, sonraki canlıyı kaçırma."),
    ("Instagram", "P2", "SMB için 10 dakikada statik site adımları.",
     "Her adım bir slide; 10 slide'lık carousel.", "Kaydet + kafeciye gönder."),
    ("TikTok", "P1", "Bugünkü mini-zafer: Signal bot dolu/boş durumunu doğru anlıyor.",
     "Önce yarım cümle gelir, sonra 'bitir' deyince süreç başlar. Küçük ama büyük UX.", "Takip et; her gün mini-zafer paylaşıyorum."),
    ("Newsletter", "P1", "90 günlük retro: MRR, churn, ikna olmak zor müşteriler.",
     "Tüm metrikler + 5 değerli yorum + gelecek 90 güne 3 hedef.", "Yorum / abone ol."),
    ("Twitter/X", "P3", "İzmir Teknopark meetup'ında 40 kişilik grup.",
     "Slide'lar GitHub'da; kayıt bu akşam YouTube'da.", "RT + davet et."),
    ("LinkedIn", "P2", "Yappaflow'un white-label fiyat modelini açıklıyorum.",
     "3 kademe: single seat, agency seat, reseller seat. Her biri farklı SLA.", "DM'den demo."),
    ("Instagram", "P1", "Bu hafta en çok yaptığım hata: Twitter'a ağırlık verip IG'yi ihmal etmek.",
     "Sonuç: IG takipçi sabit. Düzeltme: 2 reels/gün garanti.", "Bize IG takipçi kazandıran ipucun var mı yorumlara yaz."),
    ("TikTok", "P1", "Bir kurucu olarak motivasyon kaybettiğimde ne yapıyorum?",
     "3 şey: bir case study yazıyorum, bir kullanıcıyla konuşuyorum, 2 gün yazılım yazmıyorum.", "Kaydet + motivasyon kaybettiğinde izle."),
    ("Newsletter", "P3", "Patika.dev ile ortak newsletter sayısı.",
     "Onların öğrenci perspektifi + bizim ürün perspektifi. Birbirini güçlendiren iki yazı.", "Patika'ya abone ol + Yappaflow'a gel."),
    ("LinkedIn", "P1", "Bugünkü küçük kazanç: Yusuf'un annesi Yappaflow'u gördü, 'ben de bir site yaparım' dedi.",
     "Bu tam ICP — anneler SMB sahibi. Ürün-pazar uyumu böyle hissediliyor.", "Sizin benzer 'aile hikayeniz' var mı?"),
    ("YouTube", "P1", "Yappaflow'un mimari kararları (deep dive).",
     "DeepSeek'i neden seçtik, OpenRouter fallback, streaming akış. 20 dk.", "Abone ol + yorumla."),
    ("Twitter/X", "P2", "Tek tweet kural: müşteri 'net' ister.",
     "Yappaflow'un her butonu 3 kelime veya az. Uzun CTA = conversion düşüşü.", "Bir URL değil, 3 kelime kullan."),
    ("Instagram", "P3", "Samsun esnaf odasında konuşma davetliyim.",
     "30 işletmeye 'Yappaflow ile dijitalleşme' semineri.", "Esnaf olanlar, kendi ilinizde istiyor musunuz yorumda söyleyin."),
    ("TikTok", "P2", "Static site bakım nasıl yapılır (subtitle ile).",
     "Her 6 ayda bir bağlantı testi + Google Search Console kontrolü. 50 sn.", "Kaydet + yılda iki kez izle."),
    ("Newsletter", "P1", "Cuma akşamı dürüst sorusu: 'Gelecek ay Yappaflow'u kapatırsam ne olur?'",
     "Şu an ne çalışıyor ne çalışmıyor, ne riske atıyorum. Dürüst 1400 kelime.", "Cevap gönder."),
]
for i, row in enumerate(hooks):
    r = 5 + i
    hk.cell(row=r, column=1, value=i + 1).alignment = CENTER
    for j, val in enumerate(row, start=2):
        c = hk.cell(row=r, column=j, value=val)
        c.alignment = WRAP
        if j == 3:
            c.fill = PILLAR_FILL
            c.font = BODY_BOLD
    hk.row_dimensions[r].height = 64
zebra_body(hk, 5, 4 + len(hooks), len(hk_hdr))

hk_tbl_ref = f"A4:{get_column_letter(len(hk_hdr))}{4 + len(hooks)}"
hk_tbl = Table(displayName="HooksTbl", ref=hk_tbl_ref)
hk_tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
hk.add_table(hk_tbl)

set_widths(hk, {"A": 5, "B": 14, "C": 9, "D": 44, "E": 54, "F": 32})
hk.freeze_panes = "C5"

# ============================================================= 5. Influencers
inf = wb.create_sheet("Influencers")
inf.sheet_view.showGridLines = False

inf["A1"] = "Türk içerik üreticileri — hedef liste (web/ajans/freelancer odaklı)"
inf["A1"].font = TITLE_FONT
inf.merge_cells("A1:L1")

inf["A2"] = (
    "YouTube ağırlıklı (araştırma en güvenilir veriyi orada verdi). LinkedIn + Newsletter kısmı kısa ama "
    "nitelikli. IG/TikTok platform auth duvarı nedeniyle hand-curated + doğrulama gerektirecek bölüm. "
    "Her ileti başlamadan önce 'Tier' kolonuna göre önceliklendir: Tier-1 ilk 30 gün, Tier-2 ikinci 30 gün, Tier-3 bakarsa."
)
inf["A2"].font = NOTE_FONT
inf.merge_cells("A2:L2")
inf.row_dimensions[2].height = 54

inf_hdr = [
    "#", "Tier", "Platform", "Ad / Kanal", "URL",
    "Kategori", "Takipçi", "Ortalama görüntülenme",
    "Türkçe?", "Uygunluk notu", "Sponsor/ücret aralığı (₺)", "İlk pitch açısı",
]
for i, h in enumerate(inf_hdr, start=1):
    inf.cell(row=4, column=i, value=h)
style_header_row(inf, 4, len(inf_hdr))

influencers = [
    # --- YouTube Tier-1 (ideal fit) ---
    (1, "Tier-1", "YouTube", "Adem İlter", "https://www.youtube.com/@ademilter",
     "Frontend / React / tasarım", "72000", "12000", "Evet",
     "UI mühendislik dili TR tech için altın standart; ciddi freelancer izleyicisi", "25k-60k",
     "'React + Yappaflow: Türkçe AI prompt tuning' canlı mini seri"),
    (2, "Tier-1", "YouTube", "Murat Yücedağ", "https://www.youtube.com/@MuratYucedag",
     "WordPress / Web tasarım", "150000", "18000", "Evet",
     "En büyük TR WP kitlesi; 'WP'den ZIP'e geçiş' videosu ideal", "40k-85k",
     "'WordPress'ten Yappaflow'a geçiş' karşılaştırma + canlı yayın"),
    (3, "Tier-1", "YouTube", "Selman Kâhya", "https://www.youtube.com/@SelmanKahya",
     "Genel yazılım / kariyer", "280000", "25000", "Evet",
     "TR yazılım kariyer otoritesi; ICP S1'e hitap ediyor", "60k-120k",
     "'Freelancer olmak isteyenler için Yappaflow' eğitim videosu"),
    (4, "Tier-1", "YouTube", "Tayfun Erbilen", "https://www.youtube.com/@tayfunerbilen",
     "Full-stack / Laravel / PHP / FE", "180000", "14000", "Evet",
     "Çok kaliteli uzun eğitimler; freelance döneminde yaşadığı için empati yüksek", "45k-90k",
     "'Freelancer pipeline: Yappaflow + kendi backend'in' derin video"),
    (5, "Tier-1", "YouTube", "Kodluyoruz / Mustafa Akgül (bootcamp partneri)", "https://www.youtube.com/@kodluyoruz",
     "Bootcamp / kariyer dönüşümü", "120000", "8000", "Evet",
     "Junior-focused; Yappaflow junior'ların ilk freelance işi için araç olarak anlatılabilir", "sponsor-yok, kolab",
     "Kodluyoruz bootcamp'ine Yappaflow seat sponsorluğu + video"),
    # --- YouTube Tier-2 (good fit) ---
    (6, "Tier-2", "YouTube", "Arin Yazılım", "https://www.youtube.com/@ArinYazilim",
     "Backend / Laravel / SaaS", "35000", "3500", "Evet",
     "Niş ama son derece hedef-sadık; freelance + indie hacker kitle", "12k-28k",
     "'SaaS kurucusu olarak Yappaflow'u neden seçtim' röportaj"),
    (7, "Tier-2", "YouTube", "Volkan Biçer (Vtech)", "https://www.youtube.com/@VolkanBicer",
     "Frontend / web developer life", "65000", "5500", "Evet",
     "Vlog tarzı; günlük freelancer çalışma videoları", "18k-35k",
     "'Bir freelancer günü: Yappaflow dahil' vlog"),
    (8, "Tier-2", "YouTube", "Mehmet Seven", "https://www.youtube.com/@mehmetsevenozturk",
     "Yazılım kariyer / freelancer", "48000", "4000", "Evet",
     "Kariyer koçu-tonu; freelance case study formatına açık", "15k-30k",
     "'İlk 10k ₺ freelance: Yappaflow ile 2 aylık deney' meydan okuma"),
    (9, "Tier-2", "YouTube", "Sadık Turan", "https://www.youtube.com/@SadikTuran",
     "Full-stack / eğitim", "320000", "15000", "Evet",
     "Büyük kitle ama tech-heavy; ajans içeriği daha az görünür", "60k-110k",
     "'AI ajanslar için: Yappaflow derin inceleme'"),
    (10, "Tier-2", "YouTube", "Hüseyin Yılmaz (Kodland)", "https://www.youtube.com/@KodlandTV",
     "Çocuk / genç genel eğitim", "200000", "12000", "Evet",
     "ICP dışı ama Türk ebeveynler+öğrenci; uzak partnership", "sponsor",
     "Kodland bootcamp bitenlere Yappaflow sertifika+seat"),
    (11, "Tier-2", "YouTube", "Ahmet Kutlu (Frontend)", "https://www.youtube.com/@ahmetkutlu",
     "Frontend / CSS / design", "42000", "3800", "Evet",
     "CSS + mikro tasarım; ajans junior'larına değerli", "12k-28k",
     "'Yappaflow şablonlarını nasıl özelleştirirsin' tutorial"),
    (12, "Tier-2", "YouTube", "Uğur Umutluoğlu", "https://www.youtube.com/@ugurumutluoglu",
     "Web tasarım / WordPress", "55000", "4500", "Evet",
     "Orta tier WP hocası; WP kullanıcılarını ZIP'e yönlendirmek için değerli", "15k-32k",
     "'WordPress'i bırakanlar hikayesi' kullanıcı panel videosu"),
    (13, "Tier-2", "YouTube", "Ali Çetinkaya (techsu)", "https://www.youtube.com/@techsu",
     "Girişimci / tech haber", "78000", "7000", "Evet",
     "TR girişimci çevresini toplar; Yappaflow kurucu röportajı", "25k-50k",
     "'Türkiye'den Yappaflow: kurucu röportajı' 30 dk"),
    # --- YouTube Tier-3 ---
    (14, "Tier-3", "YouTube", "Emre Karakaya", "https://www.youtube.com/@emrekarakaya",
     "Web dev / eğitim", "24000", "2000", "Evet",
     "Küçük ama yüksek etkileşim; testimonial formatı güçlü", "6k-14k",
     "Testimonial video: 'Yappaflow'u 30 gün kullanan yazılımcı'"),
    (15, "Tier-3", "YouTube", "Barış Aydınoğlu", "https://www.youtube.com/@barisaydinoglu",
     ".NET / backend", "38000", "3200", "Evet",
     "Backend odaklı; fit orta", "10k-22k",
     "'Backend developer'lar neden front'u outsource etmeli' videosu"),
    (16, "Tier-3", "YouTube", "Esra Erdoğan (Teckel)", "https://www.youtube.com/@teckel",
     "Girişimci / kadın tech", "28000", "2600", "Evet",
     "Güzel segment genişletme; özellikle kadın SMB sahipleri", "8k-18k",
     "'Kadın freelancer'lar için Yappaflow' ortak video"),
    (17, "Tier-3", "YouTube", "Can Uyanıker", "https://www.youtube.com/@canuyaniker",
     "Genel prodüktivite / SaaS araçları", "45000", "3500", "Evet",
     "'Aylık 10 SaaS test ediyorum' formatı; tek bir video'da Yappaflow incelenir", "12k-28k",
     "SaaS inceleme formatı — 'Yappaflow: statik site AI' 10 dk"),
    (18, "Tier-3", "YouTube", "Burak Arıkan", "https://www.youtube.com/@BurakArikanYT",
     "Data + yazılım + entelektüel", "22000", "1800", "Evet",
     "Niş; düşüncü içerik; ürün incelemesi için az direkt", "düşük ücret / kolab",
     "Araştırma angle: 'AI + Türkçe içerik üretimi' panel"),
    (19, "Tier-3", "YouTube", "Çağrı Yardımcı", "https://www.youtube.com/@cagriyardimci",
     "Frontend / mobile / Flutter", "33000", "2500", "Evet",
     "Mobile öncelikli; fit orta", "9k-20k",
     "'Flutter dev'ler landing için Yappaflow mu?' kısa"),
    (20, "Tier-3", "YouTube", "Müjdat Müftüoğlu", "https://www.youtube.com/@mujdatmuftuoglu",
     "IT Kariyer / mentor", "67000", "5000", "Evet",
     "Kariyer koçu; kısa videoda Yappaflow başlangıç aracı olarak tanıtılabilir", "15k-32k",
     "'Freelance başlamak 2026' kısa içerik"),
    # --- LinkedIn ---
    (21, "Tier-1", "LinkedIn", "Yiğit Konur", "https://www.linkedin.com/in/yigitkonur/",
     "SEO / büyüme / TR tech düşünce", "85000 connections", "5000-20000 post", "Evet",
     "TR SEO otoritesi; Yappaflow'un organik SEO angle'ı için ideal", "guest post + newsletter swap",
     "Guest post: 'Static site SEO hızı neden önemli?'"),
    (22, "Tier-1", "LinkedIn", "Baran Cezayirli", "https://www.linkedin.com/in/barancezayirli/",
     "VC / angel / TR tech", "40000+", "3000-15000", "Evet",
     "Girişimci çevresinin merkezi; Yappaflow için distribution", "kolab / mention",
     "'Early-stage TR SaaS bu hafta' listesine eklenmek"),
    (23, "Tier-1", "LinkedIn", "Serkan Ünsal", "https://www.linkedin.com/in/serkanunsal/",
     "TR tech haber / editör", "30000+", "2000-10000", "Evet",
     "TR tech'in günlük nabzını tutan kalem", "newsletter mention",
     "Yappaflow lansmanı / milestone'larda mention"),
    (24, "Tier-2", "LinkedIn", "Ömer Çolakoğlu", "https://www.linkedin.com/in/omercolakoglu/",
     "SaaS growth / TR founder", "25000+", "1500-8000", "Evet",
     "Growth pratikleri; ortak webinar fikri için iyi", "kolab / webinar",
     "'TR SaaS growth 2026' ortak webinar"),
    (25, "Tier-2", "LinkedIn", "Ussal Sahbaz (CEPA / TechPoint)", "https://www.linkedin.com/in/usalsahbaz/",
     "Policy / tech düşüncü", "15000+", "800-3500", "Evet",
     "Policy + ürün arası köprü; güven imasi için iyi", "mention",
     "'Türkiye'den küresel SaaS' makalesinde örnek"),
    (26, "Tier-2", "LinkedIn", "Barbaros Özbuğutu (iyzico)", "https://www.linkedin.com/in/barbarosozbugutu/",
     "Fintech founder / mentor", "35000+", "3000-12000", "Evet",
     "iyzico kurucusu; ödeme entegrasyonu için strateji ortağı", "ürün kolab",
     "'Yappaflow + iyzico: tek tıkla fatura' partnership"),
    (27, "Tier-2", "LinkedIn", "Turks in Tech (topluluk)", "https://www.linkedin.com/company/turks-in-tech/",
     "Topluluk / newsletter", "10000+ follower", "1500-5000", "Evet+EN",
     "Türk tech diaspora; Yappaflow için uluslararası distribution", "newsletter feature",
     "Newsletter feature + 'TR SaaS spotlight'"),
    # --- IG/TikTok (hand-curated) ---
    (28, "Tier-1", "Instagram", "@ajans.isleri", "https://www.instagram.com/ajans.isleri/",
     "Ajans dünyası / humor + case", "39000", "3000-8000", "Evet",
     "Ajans sahipleri ve junior'ları; direkt ICP S2", "5k-15k Reels sponsor",
     "Sponsored reel: 'Ajans junior'ları için Yappaflow'"),
    (29, "Tier-1", "Instagram", "@dijitalpazarlamaokulu", "https://www.instagram.com/dijitalpazarlamaokulu/",
     "Dijital pazarlama eğitim", "90000+", "5000-15000", "Evet",
     "Geniş SMB + freelance; 'Yappaflow case study' formatı", "8k-20k",
     "Case study reel + story series"),
    (30, "Tier-2", "Instagram", "@influencerajansi", "https://www.instagram.com/influencerajansi/",
     "Influencer ajans", "11000", "1000-4000", "Evet",
     "B2B angle: Yappaflow partnership'i için ortak değerli", "3k-8k",
     "Co-marketing: 'Ajans için Yappaflow tanıtımı'"),
    (31, "Tier-2", "Instagram", "@eadamxakademi", "https://www.instagram.com/eadamxakademi/",
     "E-ticaret eğitim", "45000+", "2500-7000", "Evet",
     "SMB e-ticaret sahipleri; Yappaflow 'ucuz landing' angle", "5k-12k",
     "'Bir Etsy satıcısı Yappaflow'u denedi' reel serisi"),
    (32, "Tier-2", "Instagram", "@digital.faculty", "https://www.instagram.com/digital.faculty/",
     "Dijital eğitim / Emre Gökşin", "25000+", "1500-5000", "Evet",
     "Kurs platformu; Yappaflow'u modül olarak", "kolab / kurs",
     "Kurs modülü sponsorluğu"),
    (33, "Tier-3", "Instagram", "@ajanszero", "https://www.instagram.com/ajanszero/",
     "Ajans hayatı / meme", "8300", "500-1500", "Evet",
     "Humor + ajans; küçük ama viral potansiyel", "1-3k",
     "Meme-formatında hızlı demo"),
    (34, "Tier-1", "TikTok", "@adem.ilter (TikTok)", "https://www.tiktok.com/@ademilter",
     "Frontend / tasarım kısa klipleri", "15000+", "3000-10000", "Evet",
     "Aynı kişi YouTube ile; TikTok yüksek görünürlük", "12k-25k",
     "60 sn 'Yappaflow kullanım sırasında ne değişti'"),
    (35, "Tier-1", "TikTok", "@dijitalpazarlamaokulu (TikTok)", "https://www.tiktok.com/@dijitalpazarlamaokulu",
     "Dijital pazarlama / SMB", "60000+", "5000-30000", "Evet",
     "SMB reach + yüksek etkileşim", "10k-25k",
     "SMB hikaye serisi"),
    (36, "Tier-2", "TikTok", "@dijitalreklamuzmani", "https://www.tiktok.com/@dijitalreklamuzmani",
     "Dijital reklam / SMB", "30000+", "2000-8000", "Evet",
     "SMB ve küçük ajans odaklı", "5k-12k",
     "'Tek reklamla kazanmak için site gerekli' Yappaflow CTA"),
    (37, "Tier-2", "TikTok", "@freelancerkoc", "https://www.tiktok.com/@freelancerkoc",
     "Freelance eğitim", "22000", "2000-6000", "Evet",
     "Direkt ICP S1", "4k-10k",
     "'İlk 10 freelancer işi Yappaflow ile' meydan okuma"),
    # --- Newsletter / Substack ---
    (38, "Tier-1", "Newsletter", "Turks in Tech (Substack)", "https://turksintech.substack.com/",
     "TR tech haber + kurucu öyküleri", "4500+ sub", "~1800 open", "Evet+EN",
     "Yappaflow'un uluslararası görünürlük aracı", "guest post",
     "Guest issue + 'spotlight'"),
    (39, "Tier-1", "Newsletter", "Turkey Tech Tonik (Substack)", "https://turkeytechtonik.substack.com/",
     "TR tech haftalık", "3000+ sub", "~1200 open", "Evet",
     "Haftalık haber; Yappaflow milestone'ı orada duyulur", "mention",
     "Mention + milestone güncellemesi"),
    (40, "Tier-2", "Newsletter", "Yigit's Newsletter (Yiğit Konur)", "https://www.linkedin.com/newsletters/yigit-konur-0123456789/",
     "SEO / growth", "7000+ sub", "~2800 open", "Evet",
     "SEO + Yappaflow için ideal açı", "guest",
     "'Static site SEO 2026' guest post"),
    # --- Topluluk / platform ---
    (41, "Tier-1", "Topluluk", "Patika.dev", "https://www.patika.dev/",
     "Junior dev topluluğu", "150000+ öğrenci", "—", "Evet",
     "Junior freelance için en büyük funnel", "etkinlik / seat",
     "AMA + öğrenci demo lisansı sponsorluğu"),
    (42, "Tier-1", "Topluluk", "Kommunity (kommunity.com)", "https://kommunity.com/",
     "Tech etkinlik platformu", "çok etkinlik", "—", "Evet",
     "İzmir + İstanbul + Samsun etkinlikler", "etkinlik sponsor",
     "Meetup sponsorluğu + 'Yappaflow ile konuşma'"),
    (43, "Tier-2", "Topluluk", "BtkAkademi", "https://www.btkakademi.gov.tr/",
     "Devlet destekli eğitim", "milyonlarca kullanıcı", "—", "Evet",
     "Public sector reach", "sertifika modülü",
     "Yappaflow freelance modülü + sertifika"),
    (44, "Tier-2", "Topluluk", "Yazılım TV Discord", "https://discord.gg/yazilimtv",
     "Yazılımcı sohbet", "20000+ üye", "—", "Evet",
     "Aktif sohbet; yardımla marka inşası", "AMA",
     "Aylık AMA + destek kanalı"),
    (45, "Tier-3", "Topluluk", "ODTÜ Tech Club", "—",
     "Üniversite teknoloji kulübü", "—", "—", "Evet",
     "Üniversite giriş freelance kullanıcısı", "etkinlik",
     "Konferans + öğrenci seat"),
]

for i, row in enumerate(influencers):
    r = 5 + i
    for j, val in enumerate(row, start=1):
        c = inf.cell(row=r, column=j, value=val)
        c.alignment = WRAP
        if j == 2:
            if val == "Tier-1":
                c.fill = PRIORITY_HIGH
                c.font = BODY_BOLD
            elif val == "Tier-2":
                c.fill = PRIORITY_MED
    inf.row_dimensions[r].height = 54
zebra_body(inf, 5, 4 + len(influencers), len(inf_hdr))

inf_tbl = Table(displayName="InfluencersTbl", ref=f"A4:{get_column_letter(len(inf_hdr))}{4 + len(influencers)}")
inf_tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
inf.add_table(inf_tbl)

set_widths(inf, {"A": 4, "B": 8, "C": 12, "D": 26, "E": 42, "F": 24, "G": 12, "H": 14, "I": 8, "J": 34, "K": 22, "L": 34})
inf.freeze_panes = "D5"

# summary rows (below table)
inf_sum_row = 4 + len(influencers) + 2
inf.cell(row=inf_sum_row, column=1, value="Özet").font = H2_FONT
inf.merge_cells(start_row=inf_sum_row, start_column=1, end_row=inf_sum_row, end_column=3)

# COUNTIFs for Tier
last_data_row = 4 + len(influencers)
summaries = [
    ("Toplam influencer", f"=COUNTA(D5:D{last_data_row})"),
    ("Tier-1 sayısı", f'=COUNTIF(B5:B{last_data_row},"Tier-1")'),
    ("Tier-2 sayısı", f'=COUNTIF(B5:B{last_data_row},"Tier-2")'),
    ("Tier-3 sayısı", f'=COUNTIF(B5:B{last_data_row},"Tier-3")'),
    ("YouTube sayısı", f'=COUNTIF(C5:C{last_data_row},"YouTube")'),
    ("LinkedIn sayısı", f'=COUNTIF(C5:C{last_data_row},"LinkedIn")'),
    ("Instagram sayısı", f'=COUNTIF(C5:C{last_data_row},"Instagram")'),
    ("TikTok sayısı", f'=COUNTIF(C5:C{last_data_row},"TikTok")'),
    ("Newsletter sayısı", f'=COUNTIF(C5:C{last_data_row},"Newsletter")'),
    ("Topluluk sayısı", f'=COUNTIF(C5:C{last_data_row},"Topluluk")'),
]
for i, (k, f) in enumerate(summaries):
    r = inf_sum_row + 1 + i
    inf.cell(row=r, column=1, value=k).font = BODY_BOLD
    inf.cell(row=r, column=1).fill = SUB_FILL
    inf.cell(row=r, column=1).border = BORDER
    inf.cell(row=r, column=2, value=f).font = BODY_FONT
    inf.cell(row=r, column=2).border = BORDER
    inf.cell(row=r, column=2).alignment = CENTER
    inf.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)

# ============================================================= 6. Scripts
sc = wb.create_sheet("Scripts")
sc.sheet_view.showGridLines = False

sc["A1"] = "Dış iletişim şablonları (TR + EN)"
sc["A1"].font = TITLE_FONT
sc.merge_cells("A1:D1")

sc["A2"] = (
    "Kopyala, müşteriye göre 1-2 yer değiştir, gönder. 'Tone check' sütunu Yusuf'un 'doğal mı?' testi için. "
    "Türkiye'de 3+ şablon tekrarı spam = engel; her şablonu 2 farklı varyasyonla döndür."
)
sc["A2"].font = NOTE_FONT
sc.merge_cells("A2:D2")
sc.row_dimensions[2].height = 44

sc_hdr = ["Amaç / kanal", "Şablon (kopyala-yapıştır)", "Değiştirilecek alanlar", "Tone check / not"]
for i, h in enumerate(sc_hdr, start=1):
    sc.cell(row=4, column=i, value=h)
style_header_row(sc, 4, len(sc_hdr))

scripts_data = [
    (
        "Influencer ilk DM (Instagram / TikTok)",
        "Selam {NAME}, Yusuf'um — Yappaflow'un kurucusu. Türkiye'de freelancer ve küçük ajansların sitelerini Signal sohbetinden otomatik çıkaran bir SaaS yapıyoruz. {NAME}'ın {LAST_POST_TOPIC} içeriğini izledim; kitleniz birebir bizim ICP'miz. Ücretsiz seat + 1 reel sponsorluğu için konuşabilir miyiz? 30 dk'da ürünü gösterip size özel bir angle çıkarabilirim.",
        "{NAME} — kanal ismi; {LAST_POST_TOPIC} — son içeriğin 3 kelimelik özeti",
        "Kısa ve ürün-ilk. 'Ücretsiz' yerine 'değerli deneyim' kullan Tier-1'de.",
    ),
    (
        "Influencer ilk e-posta (YouTube)",
        "Merhaba {NAME},\n\nYusuf, Yappaflow'dan yazıyorum. {CHANNEL} kanalındaki {VIDEO_TITLE} videosunu izledim — tam olarak kitlemizin kullandığı zihin haritası.\n\nYappaflow Türkiye-first bir statik site AI üreticisi: müşteri Signal'de sohbet ediyor, çıktısı deploy-ready ZIP. {SEGMENT_FIT} için birebir. Sizin için 2 farklı kolab modeli düşündüm:\n1) Ürün incelemesi (sponsorlu): bütçe {BUDGET_RANGE} TL. Ücretsiz seat + mono-seri angle.\n2) Ortak canlı yayın: 1 saat, bir freelancer müşteriyle gerçek canlı demo.\n\nİkisinden hangisi sizin kanal stratejinize uyar? Uygunsa 20 dk zoom için 2 tarih önerebilir misiniz?\n\nSevgilerle,\nYusuf",
        "{NAME}, {CHANNEL}, {VIDEO_TITLE}, {SEGMENT_FIT} (ör. 'Junior WP freelancer'lar'), {BUDGET_RANGE}",
        "Değer teklifi 2 seçenekli; Tier-1'e büyük bütçe, Tier-3'e 'seat + promo'.",
    ),
    (
        "Newsletter guest pitch",
        "Merhaba {NAME},\n\n{NEWSLETTER_NAME} okuyucusuyum. Son {DATE} sayısında {TOPIC_MENTIONED} üzerine yazdığın kısım ilham verici oldu.\n\nYappaflow'un kurucusuyum (Türk freelancer/ajans için statik site AI). Sizin için guest bir yazı hazırlamayı teklif ediyorum: '{PROPOSED_TITLE}'. 1200-1500 kelime, sizin formatınızda, benim verdiğim gerçek kullanıcı metrikleriyle birlikte.\n\nBuna ek olarak sizin newsletter'ınızı Yappaflow sosyal kanallarında bu ay içinde 2 kez önerme sözü veriyorum. Süre olarak sizin için en rahat hangi gün?",
        "{NAME}, {NEWSLETTER_NAME}, {DATE}, {TOPIC_MENTIONED}, {PROPOSED_TITLE}",
        "Swap açısı (cross-promo) en çok işe yarıyor. Ücret teklif etme; yazar duyarlı.",
    ),
    (
        "LinkedIn düşüncü-kurucu mesajı",
        "Merhaba {NAME}, son paylaşımın {TOPIC} konusuyla Yappaflow'da yaşadığım bir ikilem çok benzer. Özellikle {SPECIFIC_POINT} kısmı. 15 dk'lık kısa bir sohbet için uygun musun? Senin perspektifinden bir ürün kararı almam gerekiyor; karşılığında Yappaflow early-access + premium seat veriyorum.",
        "{NAME}, {TOPIC}, {SPECIFIC_POINT}",
        "Satış yok — iş birliği ve öğrenme tonu. Kurucu-kurucu ton.",
    ),
    (
        "Topluluk AMA daveti (Discord / Slack)",
        "Selam {COMMUNITY_NAME} ekibi! Yusuf — Yappaflow kurucusuyum. Türkiye-first statik site AI üretiyoruz. Aktif üyelerinize 45 dk'lık canlı bir AMA + kod incelemesi önermek isterim. Karşılığında ilk 50 katılımcıya ücretsiz 'Yappaflow Founder Seat' (normal fiyat 900 TL/ay).\n\nSize uygun hangi hafta?",
        "{COMMUNITY_NAME} — Discord / Slack adı",
        "Her zaman 'karşılığında ne veriyorum' kısmı güçlü.",
    ),
    (
        "Müşteri case study daveti",
        "Merhaba {CLIENT_NAME}, Yappaflow'u {MONTH_COUNT} aydır kullanıyorsunuz. Sizinle bir küçük case study yapmak istiyorum: 30 dk Zoom görüşmesi + 2 ekran kaydı + 1 yazılı onay. Karşılığında 6 ay ücretsiz Yappaflow Pro seat. Hangi gün size uyar?",
        "{CLIENT_NAME}, {MONTH_COUNT}",
        "Kısa ve hediye-odaklı. Case study = en büyük social proof.",
    ),
    (
        "Soğuk ajans kurucu (LinkedIn DM)",
        "Merhaba {FOUNDER_NAME}, {AGENCY_NAME}'in İzmir Bayraklı'daki {LOCATION} ofisinden haberim var. Yappaflow junior ajans geliştiricilerinizin ilk sayfa süresini 2 günden 30 dakikaya indiren bir AI brief + statik site aracı. 20 dk'lık demo + sizin için hazırladığım 1 sayfalık 'ajans akış kazanç hesabı' var — görelim mi?",
        "{FOUNDER_NAME}, {AGENCY_NAME}, {LOCATION}",
        "Spesifik adres = okuma oranı %40+.",
    ),
    (
        "Soğuk freelancer (LinkedIn mesaj)",
        "Selam {FREELANCER_NAME}, {BIO_CLUE} detayın dikkat çekti. Yappaflow tam bir freelancer aracı: Signal'de müşteriyle konuşuyorsun, sabah ZIP hazır. İlk 3 müşteriye tamamen ücretsiz. Bir akşam 10 dk Zoom'da göstereyim mi?",
        "{FREELANCER_NAME}, {BIO_CLUE} (ör. 'bio'da Bionluk linkin var')",
        "10 dk'lık taahhüt düşük engel; ilk 3 ücretsiz güçlü kanca.",
    ),
    (
        "Sponsor mini-sözleşmesi özeti (ek olarak gönder)",
        "Kapsam: 1 adet {FORMAT} (60-90 sn), Yappaflow CTA bio + açıklamada 30 gün, organik olarak kalır. Ücret: {AMOUNT} TL, fatura karşılığı. Dahil: bir revizyon, Yappaflow tarafından sağlanan asset + ekran görüntüleri. Dahil değil: extra yayın, platform reklam boost'u.",
        "{FORMAT} (reel / short / tweet-thread); {AMOUNT}",
        "Her zaman tek sayfa; fatura detayları sözleşmeye.",
    ),
    (
        "Follow-up (cevap gelmezse 5 gün sonra)",
        "Merhaba {NAME}, önceki mesajımla ilgili ek bilgi paylaşıyorum: Yappaflow'u {NEW_MILESTONE} olarak duyurduk. Belki yeni hali daha ilginç gelebilir. Yine 15 dk demo için uygun musunuz? Değilse kısa bir 'uygun değil' cevabı bile yeterli — takip etmeyeceğim.",
        "{NAME}, {NEW_MILESTONE} (son ürün gelişmesi)",
        "'Takip etmeyeceğim' doğru kullanım = yanıt oranı %20 artar.",
    ),
]

for i, row in enumerate(scripts_data):
    r = 5 + i
    for j, val in enumerate(row, start=1):
        c = sc.cell(row=r, column=j, value=val)
        c.alignment = WRAP
        if j == 1:
            c.font = BODY_BOLD
            c.fill = SUB_FILL
    sc.row_dimensions[r].height = 180
zebra_body(sc, 5, 4 + len(scripts_data), len(sc_hdr))

set_widths(sc, {"A": 30, "B": 70, "C": 36, "D": 38})
sc.freeze_panes = "A5"

# ============================================================= 7. Budget
bd = wb.create_sheet("Budget")
bd.sheet_view.showGridLines = False

bd["A1"] = "Aylık pazarlama bütçesi + KPI panosu"
bd["A1"].font = TITLE_FONT
bd.merge_cells("A1:F1")

bd["A2"] = (
    "3 senaryo (lean / standard / aggressive). Yappaflow başlangıç için 'lean' yeterli; Q2'den itibaren "
    "'standard'a geçmek mantıklı. 'Bottom line' cell'leri canlı formül — senaryoyu C4'ten değiştirince "
    "yeniden hesaplanır."
)
bd["A2"].font = NOTE_FONT
bd.merge_cells("A2:F2")
bd.row_dimensions[2].height = 44

bd["A4"] = "Senaryo seçimi (B / D / F kolonlarından birini seç):"
bd["A4"].font = BODY_BOLD
bd["A4"].alignment = WRAP
bd.merge_cells("A4:F4")

bd_hdr = ["Kalem", "Lean aylık (₺)", "Notlar (lean)", "Standard aylık (₺)", "Notlar (standard)", "Aggressive aylık (₺)"]
for i, h in enumerate(bd_hdr, start=1):
    bd.cell(row=5, column=i, value=h)
style_header_row(bd, 5, len(bd_hdr))

budget_rows = [
    ("Influencer kolabı (sponsorlu)", 8000, "1 × Tier-3 YouTuber veya 2 × Instagram reel",
     25000, "1 × Tier-1 YouTuber + 1 × TikTok", 60000),
    ("Newsletter guest / swap", 0, "Karşılıklı swap, para transferi yok",
     3000, "1 × paid newsletter spot", 10000),
    ("Ücretli reklam (LinkedIn retargeting)", 2000, "Sadece website ziyaretçisi retargeting",
     8000, "Audience + video view kampanyası", 20000),
    ("Ücretli reklam (Meta — IG/FB)", 3000, "Reels promote + SMB şehir hedef",
     10000, "Conversion kampanyası + whitelist", 30000),
    ("Asset üretimi (video/freelance editor)", 4000, "1 editor yarı-zamanlı 2 klip",
     12000, "1 tam-zamanlı editor + grafik tasarımcı", 25000),
    ("Etkinlik / meetup", 2000, "Ayda 1 küçük (Samsun / İzmir)",
     7000, "Aylık 1 şehir + ikinci şehirde katılım", 18000),
    ("SaaS araçları (Buffer, Linear, Canva Pro)", 1500, "Minimum set",
     3000, "Full set + analytics", 5000),
    ("Topluluk sponsorluğu (Patika, Kommunity seat)", 0, "Karşılıklı seat",
     4000, "Bir bootcamp sponsor seat'i", 12000),
    ("Kontingensi / fırsat bütçesi", 1500, "Beklenmedik viral fırsat",
     5000, "Aynı konu, daha yüksek", 10000),
]

for i, row in enumerate(budget_rows):
    r = 6 + i
    bd.cell(row=r, column=1, value=row[0]).font = BODY_BOLD
    bd.cell(row=r, column=1).alignment = WRAP
    bd.cell(row=r, column=2, value=row[1]).number_format = "#,##0"
    bd.cell(row=r, column=3, value=row[2]).alignment = WRAP
    bd.cell(row=r, column=4, value=row[3]).number_format = "#,##0"
    bd.cell(row=r, column=5, value=row[4]).alignment = WRAP
    bd.cell(row=r, column=6, value=row[5]).number_format = "#,##0"
    bd.row_dimensions[r].height = 34

zebra_body(bd, 6, 5 + len(budget_rows), len(bd_hdr))

# Totals row
tot_row = 6 + len(budget_rows)
bd.cell(row=tot_row, column=1, value="TOPLAM AYLIK").font = HDR_FONT
bd.cell(row=tot_row, column=1).fill = HDR_FILL
bd.cell(row=tot_row, column=1).alignment = CENTER
bd.cell(row=tot_row, column=2, value=f"=SUM(B6:B{tot_row - 1})").font = BODY_BOLD
bd.cell(row=tot_row, column=2).number_format = "#,##0 [$₺-tr-TR]"
bd.cell(row=tot_row, column=4, value=f"=SUM(D6:D{tot_row - 1})").font = BODY_BOLD
bd.cell(row=tot_row, column=4).number_format = "#,##0 [$₺-tr-TR]"
bd.cell(row=tot_row, column=6, value=f"=SUM(F6:F{tot_row - 1})").font = BODY_BOLD
bd.cell(row=tot_row, column=6).number_format = "#,##0 [$₺-tr-TR]"
for c in range(1, 7):
    bd.cell(row=tot_row, column=c).border = BORDER
    if c != 1:
        bd.cell(row=tot_row, column=c).fill = GOOD_FILL

# 12-month rollup
roll_row = tot_row + 2
bd.cell(row=roll_row, column=1, value="12 aylık toplam (senaryo bazında)").font = H2_FONT
bd.merge_cells(start_row=roll_row, start_column=1, end_row=roll_row, end_column=6)

bd.cell(row=roll_row + 1, column=1, value="Lean 12-ay").font = BODY_BOLD
bd.cell(row=roll_row + 1, column=2, value=f"=B{tot_row}*12")
bd.cell(row=roll_row + 1, column=2).number_format = "#,##0 [$₺-tr-TR]"
bd.cell(row=roll_row + 2, column=1, value="Standard 12-ay").font = BODY_BOLD
bd.cell(row=roll_row + 2, column=2, value=f"=D{tot_row}*12")
bd.cell(row=roll_row + 2, column=2).number_format = "#,##0 [$₺-tr-TR]"
bd.cell(row=roll_row + 3, column=1, value="Aggressive 12-ay").font = BODY_BOLD
bd.cell(row=roll_row + 3, column=2, value=f"=F{tot_row}*12")
bd.cell(row=roll_row + 3, column=2).number_format = "#,##0 [$₺-tr-TR]"
for rr in range(roll_row + 1, roll_row + 4):
    for cc in range(1, 3):
        bd.cell(row=rr, column=cc).border = BORDER

# KPI dashboard
kpi_row = roll_row + 5
bd.cell(row=kpi_row, column=1, value="KPI dashboard (haftalık gözden geçir)").font = H2_FONT
bd.merge_cells(start_row=kpi_row, start_column=1, end_row=kpi_row, end_column=6)

kpi_hdr = ["KPI", "Formül", "Haftalık hedef", "Aylık hedef", "3-aylık hedef", "Kırmızı çizgi"]
for i, h in enumerate(kpi_hdr, start=1):
    bd.cell(row=kpi_row + 1, column=i, value=h)
style_header_row(bd, kpi_row + 1, len(kpi_hdr))

kpi_rows = [
    ("Waitlist signup", "Landing form submits", 30, 120, 360, "< 50/ay"),
    ("Ücretsiz aktif hesap", "Signal bot'ta 1+ brief tamamlayan", 15, 60, 180, "< 30/ay"),
    ("Ücretli aktif hesap", "Pro seat ödemesi yapmış", 3, 12, 36, "< 8/ay"),
    ("MRR (₺)", "Toplam aylık abonelik", 2000, 8000, 25000, "< 5000/ay"),
    ("Demo izleyici", "Canlı demo / webinar katılımı", 20, 80, 240, "< 50/ay"),
    ("LinkedIn takipçi büyüme", "Yeni follower / hafta", 80, 320, 1000, "< 200/ay"),
    ("Twitter/X takipçi büyüme", "Yeni follower / hafta", 60, 240, 750, "< 150/ay"),
    ("IG+TikTok takipçi büyüme", "Yeni follower / hafta (toplam)", 120, 480, 1500, "< 300/ay"),
    ("Newsletter abonesi", "Yeni sub / hafta", 25, 100, 300, "< 50/ay"),
    ("Case study yayını", "Canlı yayına giren", 0, 2, 6, "< 1/ay"),
    ("Influencer kolabı", "Gönderilen DM → kabul", 1, 3, 9, "< 1/ay"),
]
for i, row in enumerate(kpi_rows):
    r = kpi_row + 2 + i
    for j, val in enumerate(row, start=1):
        c = bd.cell(row=r, column=j, value=val)
        c.alignment = WRAP if j in (1, 2, 6) else CENTER
        if j in (3, 4, 5):
            c.number_format = "#,##0"
    bd.row_dimensions[r].height = 22
zebra_body(bd, kpi_row + 2, kpi_row + 1 + len(kpi_rows), len(kpi_hdr))

set_widths(bd, {"A": 40, "B": 18, "C": 30, "D": 18, "E": 30, "F": 18})
bd.freeze_panes = "A6"


# ============================================================= save
wb.save(OUT)
print(f"Wrote {OUT}")
print(f"Playbook rows: {len(play_rows)}")
print(f"Calendar weeks: {len(cal_rows)}")
print(f"Hook bank entries: {len(hooks)}")
print(f"Influencer entries: {len(influencers)}")
print(f"Script templates: {len(scripts_data)}")
print(f"Budget line items: {len(budget_rows)}")
print(f"KPI items: {len(kpi_rows)}")
